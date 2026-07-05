#!/usr/bin/env python3
import argparse
import ast
import base64
import csv
import hashlib
import json
import os
import re
import shlex
import signal
import socket
import subprocess
import shutil
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib import error, request
import xml.etree.ElementTree as ET


QUESTION_KEYS = (
    "question",
    "prompt",
    "instruction",
    "reformulated_task",
    "requirements",
    "problem",
    "task",
    "query",
    "input",
    "goal",
    "issue",
    "issue_text",
    "description",
    "statement",
    "problem_statement",
    "input_text",
    "query_text",
    "prompt_text",
    "text",
)
ANSWER_KEYS = (
    "answer",
    "answers",
    "rubric",
    "criteria",
    "grading",
    "evaluation_criteria",
    "correct_answer",
    "correct",
    "label",
    "target",
    "output",
    "expected",
    "expected_answer",
    "solution",
    "gold",
    "reference",
    "ground_truth",
    "final_answer",
    "answer_key",
    "correct_option",
    "expected_output",
    "final_response",
    "patch",
    "gold_patch",
    "python_solution",
)
CHOICE_KEYS = ("choices", "options", "candidates", "multiple_choice_targets")
RUBRIC_KEYS = ("rubric", "criteria", "grading", "evaluation_criteria", "answer_rubric")
MODEL_VISIBLE_CONTEXT_PRIORITY_KEYS = (
    "tables",
    "table",
    "exhibit",
    "exhibits",
    "context",
    "contexts",
    "passage",
    "passages",
    "document",
    "documents",
    "article",
    "articles",
    "data",
    "input_data",
    "inputs",
    "facts",
    "fact_pattern",
    "case",
    "scenario",
)
MODEL_VISIBLE_CONTEXT_SKIP_KEYS = set(QUESTION_KEYS) | set(ANSWER_KEYS) | set(CHOICE_KEYS) | {
    "id",
    "qid",
    "question_id",
    "instance_id",
    "topic",
    "category",
    "split",
    "source",
    "metadata",
}
SKIP_DIRS = {".git", "__pycache__", ".venv", "venv", "node_modules", ".mypy_cache", ".pytest_cache"}
MAX_FIELD_CHARS = 6000
MAX_VISIBLE_CONTEXT_CHARS = 8000
MAX_RECORDS_PER_FILE = 500
JUDGE_REQUIRED_KEYS = {"score", "passed", "reason"}
STRICT_STATUSES = {
    "passed",
    "failed_model_answer",
    "failed_model_format",
    "failed_model_missing_artifact",
    "failed_model_tool_use",
    "failed_harness_setup",
    "failed_dataset_extraction",
    "failed_missing_assets",
    "failed_unsupported_capability",
    "skipped_unsupported_capability",
    "failed_grader",
    "failed_token_budget",
    "failed_missing_required_tool",
    "failed_invalid_task_context",
    "failed_model_endpoint",
    "timed_out",
}
INVALID_EVALUATION_STATUSES = {
    "failed_harness_setup",
    "failed_dataset_extraction",
    "failed_missing_assets",
    "failed_unsupported_capability",
    "skipped_unsupported_capability",
    "failed_grader",
    "failed_token_budget",
    "failed_missing_required_tool",
    "failed_invalid_task_context",
    "failed_model_endpoint",
    "timed_out",
}
STATUS_ALIASES = {
    "failed_unsupported_capability": "skipped_unsupported_capability",
    "failed_judge_parse": "failed_grader",
    "skipped_missing_assets": "failed_missing_assets",
}
HTTP_FALLBACK_STATUS_CODES = {400, 404, 422}
MAX_AGENT_TURNS = 8
DEFAULT_AGENT_MAX_TOKENS = 16384
TOOL_RESULT_CHARS = 6000
MAX_REPEATED_IDENTICAL_TOOL_CALLS = 3
DEFAULT_MAX_TOTAL_TOOL_CALLS = 40
MAX_CONSECUTIVE_FAILED_TOOL_CALLS = 5
REPO_PATCH_GRADER_ENV = "AGENT_BENCH_REPO_PATCH_GRADER"
SWELANCER_TEST_COMMAND_ENV = "AGENT_BENCH_SWELANCER_TEST_COMMAND"
SWELANCER_GRADER_TIMEOUT_ENV = "AGENT_BENCH_SWELANCER_GRADER_TIMEOUT"
TARGET_REPO_ROOT_ENV = "AGENT_BENCH_TARGET_REPO_ROOT"
FINTOOLBENCH_EXECUTABLE_TOOLS_ENV = "AGENT_BENCH_FINTOOLBENCH_EXECUTABLE_TOOLS"
FINTOOLBENCH_FIXTURE_ROOT_ENV = "AGENT_BENCH_FINTOOLBENCH_FIXTURE_ROOT"
FINANCE_AGENT_V2_FIXTURE_ROOT_ENV = "AGENT_BENCH_FINANCE_AGENT_V2_FIXTURE_ROOT"
DEFAULT_ASSET_ROOT = "/tmp/agent-bench-assets"
FINANCE_AGENT_V2_REQUIRED_TOOLS = {
    "web_search",
    "edgar_search",
    "parse_html_page",
    "retrieve_information",
    "price_history",
}
FINANCE_AGENT_V2_REQUIRED_ENV = (
    "VALS_API_KEY",
    "TAVILY_API_KEY",
    "SEC_EDGAR_API_KEY",
    "PRICING_DATA_API_KEY",
)
FINANCE_AGENT_V2_CANARY_TERMS = (
    ("ai_native_falcon", ("ai-native", "falcon")),
    ("charlotte_ai", ("charlotte ai", "generative ai")),
    ("aidr", ("ai detection and response",)),
    ("ai_adoption_tailwind", ("ai adoption", "demand")),
    ("adversarial_ai", ("threat actors", "ai")),
    ("ai_regulatory_compliance", ("regulatory", "compliance", "ai")),
    ("rapid_ai_evolution", ("rapidly evolving", "ai")),
)
FINTOOLBENCH_NET_PPE_KEYS = (
    "propertyPlantEquipmentNet",
    "netPropertyPlantEquipment",
    "propertyPlantAndEquipmentNet",
    "netPPE",
    "netPpne",
    "netPPNE",
)
FINTOOLBENCH_REQUIRED_CANARY_ARGS = {
    "companies_balance_sheet_statements": {"symbol": "MMM"},
}
EXPLOITBENCH_REQUIRED_CONFIGS = (
    "benchmarks/v8.yaml",
    "benchmarks/v8-small.yaml",
)
EXPLOITBENCH_REQUIRED_PATHS = EXPLOITBENCH_REQUIRED_CONFIGS + (
    "benchmarks/bench-v8",
)
EXPLOITBENCH_EXCLUDE_PATTERNS = (
    "/docs/",
    "/website/",
    "readme.md",
    "spec.md",
    "citation.cff",
    "contributing.md",
    "security.md",
)
SWELANCER_EXPENSIFY_REPOSITORY = "https://github.com/Expensify/App.git"
SWELANCER_OFFICIAL_PUBLIC_REPOSITORY = "https://github.com/openai/SWELancer-Benchmark.git"
SWELANCER_OFFICIAL_PUBLIC_REF = "3d719bbd5a8cb41295357abc6304fcd29fe68c93"
SWELANCER_CWD_TARGET_REPOS = {
    "/app/expensify": SWELANCER_EXPENSIFY_REPOSITORY,
    "expensify": SWELANCER_EXPENSIFY_REPOSITORY,
}
FINMCP_STATIC_PROMPT = """You are answering a static transcript-reasoning benchmark item.

The transcript below may include prior tool calls and tool outputs from the source dataset.
These tools are not live in this environment.
Use only the provided transcript and data to produce the final answer.

Return only JSON:
{
  "answer": "...",
  "supporting_evidence": "..."
}
"""
STATIC_FINANCE_PROMPT = """You are answering a local static finance benchmark item.

The upstream benchmark may name external finance tools or hosted retrieval services, but those
live backends are not available in this local harness. Use only the benchmark prompt and any
visible record context provided in the task.

Return only compact final-answer JSON.
"""
BINARY_SUFFIXES = {
    ".7z",
    ".bin",
    ".bz2",
    ".dmg",
    ".gif",
    ".gz",
    ".idx",
    ".jpeg",
    ".jpg",
    ".npy",
    ".npz",
    ".pack",
    ".parquet",
    ".pdf",
    ".png",
    ".pkl",
    ".pyc",
    ".rev",
    ".tar",
    ".tgz",
    ".webp",
    ".xlsx",
    ".zip",
}


def normalize_status(status: Any) -> str:
    if not isinstance(status, str):
        return ""
    status = status.strip()
    return STATUS_ALIASES.get(status, status)


def status_count(status_counts: dict[str, Any], status: str) -> int:
    normalized = normalize_status(status)
    for key, value in status_counts.items():
        if normalize_status(key) == normalized and isinstance(value, int):
            return value
    return 0


class ChatCompletionHTTPError(Exception):
    def __init__(self, code: int, body: str) -> None:
        super().__init__(body or f"HTTP {code}")
        self.code = code
        self.body = body


class ChatCompletionTimeoutError(TimeoutError):
    def __init__(self, timeout_seconds: float) -> None:
        super().__init__(f"model request timed out after {timeout_seconds:.1f}s")
        self.timeout_seconds = timeout_seconds


@dataclass(slots=True)
class BenchmarkItem:
    question: str
    expected: str
    source: str
    choices: dict[str, str] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class CapabilitySupport:
    capability: str
    workspace: bool
    tools: bool
    output_collection: bool
    grader: bool
    native: bool = True
    reason: str = ""

    @property
    def supported(self) -> bool:
        return self.workspace and self.tools and self.output_collection and self.grader and self.native

    def to_dict(self) -> dict[str, Any]:
        return {
            "capability": self.capability,
            "workspace": self.workspace,
            "tools": self.tools,
            "output_collection": self.output_collection,
            "grader": self.grader,
            "native": self.native,
            "supported": self.supported,
            "reason": self.reason,
        }


@dataclass(slots=True)
class TaskWorkspace:
    root: Path
    output_dir: Path
    manifest: list[str] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class AgentRun:
    answer: str
    content: str = ""
    usage: dict[str, Any] = field(default_factory=dict)
    tool_trace: list[dict[str, Any]] = field(default_factory=list)
    diagnostics: dict[str, Any] = field(default_factory=dict)
    status_code: int = 200


@dataclass(slots=True)
class OutputBundle:
    answer: str = ""
    patch: str = ""
    artifact_paths: list[str] = field(default_factory=list)
    tool_trace: list[dict[str, Any]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


class AdapterSetupError(Exception):
    def __init__(self, status: str, message: str, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        status = normalize_status(status)
        if status not in STRICT_STATUSES:
            raise ValueError(f"invalid benchmark status: {status}")
        self.status = status
        self.details = details or {}


class BenchmarkAdapter:
    name = "chat_answer"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        return {}

    def supported_capabilities(self) -> set[str]:
        return {
            capability
            for capability, support in self.capability_support().items()
            if support.supported
        }

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        return _annotate_tool_contract(capability_contract_for(required_capabilities, self), required_capabilities, items)

    def prepare_task(self, item: BenchmarkItem) -> TaskWorkspace:
        root = _isolated_workspace_root(item)
        _write_model_visible_task_files(root, item)
        return TaskWorkspace(
            root=root,
            output_dir=_item_output_dir(item),
            manifest=_workspace_manifest(root),
            metadata={"sanitized_workspace": True, "source": item.source},
        )

    def available_tools(self, item: BenchmarkItem, workspace: TaskWorkspace) -> list[dict[str, Any]]:
        return agent_tool_schemas()

    def tools_sent_to_model(self, item: BenchmarkItem, tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return _tools_sent_to_model(item, tools)

    def run_agent_loop(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        with pushd(workspace.root):
            try:
                result = run_agent_loop(benchmark, item, tools=tools)
            except TypeError as exc:
                if "tools" not in str(exc):
                    raise
                result = run_agent_loop(benchmark, item)
        return _agent_run_from_dict(result)

    def collect_outputs(self, run: AgentRun, workspace: TaskWorkspace) -> OutputBundle:
        return OutputBundle(
            answer=run.answer,
            tool_trace=run.tool_trace,
            metadata={"content_sample": run.content[:500], "diagnostics": run.diagnostics},
        )

    def grade(self, benchmark: str, item: BenchmarkItem, outputs: OutputBundle) -> tuple[float, dict[str, Any]]:
        score, grade = grade_answer(benchmark, item, outputs.answer)
        grade.setdefault("output_collection", "answer")
        return score, grade

    def evaluate_item(self, benchmark: str, item: BenchmarkItem) -> dict[str, Any]:
        workspace: TaskWorkspace | None = None
        item_dir = _item_output_dir(item)
        required_capabilities = _item_required_capabilities(item)
        preflight_fields = _evaluation_preflight_fields(item, self)
        _write_item_json(item_dir, "item.json", _item_to_dict(item))
        try:
            workspace = self.prepare_task(item)
            tools = self.available_tools(item, workspace)
            sent_tools = self.tools_sent_to_model(item, tools)
            missing_tools = _missing_required_tools(item, tools)
            if "tool_call" in required_capabilities and not sent_tools:
                missing_tools = sorted(set(missing_tools) or {"tool_call_backend"})
            preflight_fields = _evaluation_preflight_fields(item, self, tools, missing_tools)
            _write_item_json(item_dir, "workspace.json", _workspace_to_dict(workspace))
            _write_item_json(item_dir, "tools.json", _tool_manifest(item, tools, sent_tools))
            if missing_tools:
                reason = f"Missing required tool(s): {', '.join(missing_tools)}"
                payload = evaluation_payload(
                    item,
                    "",
                    0.0,
                    reason,
                    **preflight_fields,
                    status="failed_missing_required_tool",
                    adapter=self.name,
                    capabilities_verified=False,
                    setup_details={
                        "blocker_type": "missing_required_tool",
                        "required_capabilities": sorted(required_capabilities),
                        "required_tools": _item_required_tools(item),
                        "missing_tools": missing_tools,
                        "exposed_tools": _tool_schema_names(tools),
                    },
                )
                _write_item_json(
                    item_dir,
                    "setup_error.json",
                    {
                        "status": "failed_missing_required_tool",
                        "error": reason,
                        "details": payload["setup_details"],
                    },
                )
                _write_item_json(item_dir, "item_result.json", payload)
                return payload
            run = self.run_agent_loop(benchmark, item, workspace, sent_tools)
            _write_item_json(item_dir, "agent_run.json", _agent_run_to_dict(run))
            protocol_status = normalize_status(run.diagnostics.get("status")) if isinstance(run.diagnostics, dict) else ""
            if protocol_status in STRICT_STATUSES and protocol_status != "passed":
                payload = evaluation_payload(
                    item,
                    run.answer,
                    0.0,
                    str(run.diagnostics.get("reason") or protocol_status),
                    status=protocol_status,
                    status_code=run.status_code,
                    content_sample=run.content[:500],
                    usage=run.usage,
                    tool_trace=run.tool_trace,
                    protocol_diagnostics=run.diagnostics,
                    **preflight_fields,
                    adapter=self.name,
                    capabilities_verified=protocol_status not in INVALID_EVALUATION_STATUSES,
                )
                _write_item_json(item_dir, "item_result.json", payload)
                return payload
            outputs = self.collect_outputs(run, workspace)
            _write_item_json(item_dir, "output_bundle.json", _output_bundle_to_dict(outputs))
            protocol = outputs.metadata.get("diagnostics") if isinstance(outputs.metadata, dict) else {}
            if isinstance(protocol, dict) and protocol.get("final_output_contains_tool_syntax"):
                payload = evaluation_payload(
                    item,
                    outputs.answer,
                    0.0,
                    "final output contained tool-call syntax",
                    status="failed_model_format",
                    status_code=run.status_code,
                    content_sample=run.content[:500],
                    usage=run.usage,
                    tool_trace=outputs.tool_trace,
                    protocol_diagnostics=protocol,
                    **preflight_fields,
                    output_bundle={
                        "answer_present": bool(outputs.answer),
                        "patch_present": bool(outputs.patch),
                        "artifact_paths": outputs.artifact_paths,
                        "metadata": outputs.metadata,
                    },
                    adapter=self.name,
                    capabilities_verified=True,
                )
                _write_item_json(item_dir, "item_result.json", payload)
                return payload
            score, grade = self.grade(benchmark, item, outputs)
            _write_item_json(item_dir, "grade.json", grade)
        except ChatCompletionTimeoutError as exc:
            payload = evaluation_payload(
                item,
                "",
                0.0,
                str(exc),
                status="timed_out",
                timed_out=True,
                **preflight_fields,
                adapter=self.name,
                capabilities_verified=False,
            )
            _write_item_json(item_dir, "item_result.json", payload)
            return payload
        except ChatCompletionHTTPError as exc:
            status = "failed_token_budget" if _looks_like_context_error(exc.body) else "failed_model_endpoint"
            payload = evaluation_payload(
                item,
                "",
                0.0,
                exc.body[-2000:] or str(exc),
                status=status,
                status_code=exc.code,
                **preflight_fields,
                adapter=self.name,
                capabilities_verified=False,
            )
            _write_item_json(item_dir, "item_result.json", payload)
            return payload
        except AdapterSetupError as exc:
            payload = evaluation_payload(
                item,
                "",
                0.0,
                str(exc),
                status=exc.status,
                **preflight_fields,
                adapter=self.name,
                capabilities_verified=False,
                setup_details=exc.details,
            )
            _write_item_json(item_dir, "setup_error.json", {"status": exc.status, "error": str(exc), "details": exc.details})
            _write_item_json(item_dir, "item_result.json", payload)
            return payload
        except Exception as exc:
            payload = evaluation_payload(
                item,
                "",
                0.0,
                str(exc),
                status="failed_harness_setup",
                **preflight_fields,
                adapter=self.name,
                capabilities_verified=False,
            )
            _write_item_json(item_dir, "setup_error.json", {"status": "failed_harness_setup", "error": str(exc)})
            _write_item_json(item_dir, "item_result.json", payload)
            return payload

        error_message = ""
        if score < 1.0:
            error_message = str(grade.get("reason") or f"expected {item.expected!r}, got {outputs.answer!r}")
        protocol = outputs.metadata.get("diagnostics") if isinstance(outputs.metadata, dict) else {}
        status_override = grade.get("status") if isinstance(grade.get("status"), str) else None
        if score < 1.0 and (not status_override or normalize_status(status_override) == "failed_model_answer"):
            if isinstance(protocol, dict) and protocol.get("final_output_contains_tool_syntax"):
                status_override = "failed_model_format"
                error_message = error_message or "final output contained tool-call syntax"
            elif isinstance(protocol, dict) and protocol.get("ignored_tool_failure_before_final"):
                status_override = "failed_model_tool_use"
                error_message = error_message or "model finalized after a failed tool call"
        payload = evaluation_payload(
            item,
            outputs.answer or outputs.patch or ", ".join(outputs.artifact_paths),
            score,
            error_message,
            status_code=run.status_code,
            content_sample=run.content[:500],
            usage=run.usage,
            tool_trace=outputs.tool_trace,
            protocol_diagnostics=protocol if isinstance(protocol, dict) else {},
            **preflight_fields,
            output_bundle={
                "answer_present": bool(outputs.answer),
                "patch_present": bool(outputs.patch),
                "artifact_paths": outputs.artifact_paths,
                "metadata": outputs.metadata,
            },
            grade=grade,
            status=status_override,
            timed_out=bool(grade.get("timed_out")),
            judge_parse_repaired=bool(grade.get("judge_parse_repaired")),
            adapter=self.name,
            capabilities_verified=(
                normalize_status(status_override) not in INVALID_EVALUATION_STATUSES
                and not preflight_fields["missing_tools"]
                and ("tool_call" not in required_capabilities or bool(preflight_fields["exposed_tools"]))
            ),
        )
        _write_item_json(item_dir, "item_result.json", payload)
        return payload


class ChatAnswerAdapter(BenchmarkAdapter):
    name = "chat_answer"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        return {
            "chat_answer": CapabilitySupport("chat_answer", True, True, True, True),
            "external_data_required": CapabilitySupport("external_data_required", True, True, True, True),
        }


class ToolCallAdapter(ChatAnswerAdapter):
    name = "tool_call"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        support = super().capability_support()
        support["tool_call"] = CapabilitySupport("tool_call", True, True, True, True)
        return support


class StaticTranscriptReasoningAdapter(ChatAnswerAdapter):
    name = "static_transcript_reasoning"

    def available_tools(self, item: BenchmarkItem, workspace: TaskWorkspace) -> list[dict[str, Any]]:
        return []

    def tools_sent_to_model(self, item: BenchmarkItem, tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return []

    def run_agent_loop(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        static_item = BenchmarkItem(
            question=(
                f"{FINMCP_STATIC_PROMPT}\n\n"
                f"Benchmark question and transcript:\n{item.question}"
            ),
            expected=item.expected,
            source=item.source,
            choices=item.choices,
            metadata={**item.metadata, "live_tools_required": False},
        )
        with pushd(workspace.root):
            return _agent_run_from_dict(run_agent_loop(benchmark, static_item, tools=[]))


class StaticFinanceReasoningAdapter(ToolCallAdapter):
    name = "static_finance_reasoning"

    def available_tools(self, item: BenchmarkItem, workspace: TaskWorkspace) -> list[dict[str, Any]]:
        return []

    def tools_sent_to_model(self, item: BenchmarkItem, tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return []

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        required_tools = sorted({tool for item in items for tool in _item_required_tools(item)})
        for support in contract.values():
            if isinstance(support, dict) and support.get("supported") is True:
                support["mode"] = "static_local_finance"
                support["static_degraded_mode"] = True
                support["native_tool_call_protocol_supported"] = "tool_call" in HARNESS_SUPPORTED_CAPABILITIES
                support["generic_sandbox_tools_available"] = bool(agent_tool_schemas())
                support["benchmark_required_tools_available"] = False
                support["required_benchmark_tools"] = required_tools
                support["tools"] = False
                support["reason"] = "Live finance tool backends are not available; evaluating extracted public prompt rows statically"
        return contract

    def run_agent_loop(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        static_item = BenchmarkItem(
            question=f"{STATIC_FINANCE_PROMPT}\n\nBenchmark question:\n{item.question}",
            expected=item.expected,
            source=item.source,
            choices=item.choices,
            metadata={**item.metadata, "live_tools_required": False},
        )
        with pushd(workspace.root):
            return _agent_run_from_dict(run_agent_loop(benchmark, static_item, tools=[]))


class FinToolBenchAdapter(ToolCallAdapter):
    name = "fintoolbench_live_tools"

    def available_tools(self, item: BenchmarkItem, workspace: TaskWorkspace) -> list[dict[str, Any]]:
        required = set(_item_required_tools(item))
        schemas = _fintoolbench_tool_schemas(Path.cwd(), required_tools=required)
        valid_tools = _fintoolbench_valid_backend_tools(Path.cwd(), required)
        return [schema for schema in schemas if _tool_schema_name(schema) in valid_tools]

    def run_agent_loop(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        previous = os.environ.get(FINTOOLBENCH_EXECUTABLE_TOOLS_ENV)
        os.environ[FINTOOLBENCH_EXECUTABLE_TOOLS_ENV] = json.dumps(sorted(_tool_schema_names(tools)))
        try:
            return super().run_agent_loop(benchmark, item, workspace, tools)
        finally:
            if previous is None:
                os.environ.pop(FINTOOLBENCH_EXECUTABLE_TOOLS_ENV, None)
            else:
                os.environ[FINTOOLBENCH_EXECUTABLE_TOOLS_ENV] = previous

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        required_tools = sorted({tool for item in items for tool in _item_required_tools(item)})
        tool_manifest_path = Path.cwd() / "tools" / "tools_all_annotated.jsonl"
        tool_manifest_count = len(_fintoolbench_tool_manifest(tool_manifest_path)) if tool_manifest_path.is_file() else 0
        schema_tools = set(_tool_schema_names(_fintoolbench_tool_schemas(Path.cwd(), required_tools=set(required_tools))))
        backend_canaries = _fintoolbench_backend_canaries(Path.cwd(), set(required_tools))
        invalid_backend_tools = sorted(
            tool
            for tool, canary in backend_canaries.items()
            if tool in schema_tools and canary.get("passed") is not True
        )
        available_tools = schema_tools - set(invalid_backend_tools)
        missing_schema_tools = sorted(set(required_tools) - schema_tools)
        missing_tools = sorted(set(missing_schema_tools) | set(invalid_backend_tools))
        backend_available = not missing_tools and bool(available_tools)
        support = contract.get("tool_call")
        if isinstance(support, dict):
            support["mode"] = "live_financial_tools"
            support["tool_manifest"] = "tools/tools_all_annotated.jsonl"
            support["tool_manifest_count"] = tool_manifest_count
            support["backend"] = "agent_bench_fixture_deterministic"
            support["backend_available"] = backend_available
            support["required_benchmark_tools"] = required_tools
            support["schema_tools"] = sorted(schema_tools)
            support["exposed_tools"] = sorted(available_tools)
            support["executable_tools"] = sorted(available_tools) if backend_available else []
            support["missing_tools"] = missing_tools
            support["missing_schema_tools"] = missing_schema_tools
            support["missing_tool_backends"] = invalid_backend_tools
            support["backend_canaries"] = backend_canaries
            support["benchmark_required_tools_available"] = backend_available
            if not schema_tools:
                support["supported"] = False
                support["tools"] = False
                support["reason"] = "FinToolBench tool manifest tools/tools_all_annotated.jsonl did not expose callable schemas"
            elif missing_schema_tools:
                support["supported"] = False
                support["tools"] = False
                support["reason"] = "FinToolBench required tool schemas are missing: " + ", ".join(missing_tools)
            elif invalid_backend_tools:
                support["supported"] = False
                support["tools"] = False
                support["reason"] = (
                    "FinToolBench required tool backend failed semantic canary: "
                    + ", ".join(invalid_backend_tools)
                )
        return contract


class FinanceAgentV2Adapter(ToolCallAdapter):
    name = "finance_agent_v2_fixture_tools"

    def available_tools(self, item: BenchmarkItem, workspace: TaskWorkspace) -> list[dict[str, Any]]:
        return _finance_agent_v2_available_tool_schemas(workspace.root)

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        backend = _finance_agent_v2_backend_status(Path.cwd())
        missing_env = _finance_agent_v2_missing_environment()
        exposed_tools = _tool_schema_names(_finance_agent_v2_tool_schemas() if backend["ready"] else [])
        missing_tools = sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS - set(exposed_tools))
        for capability in ("tool_call", "external_data_required"):
            support = contract.get(capability)
            if not isinstance(support, dict):
                continue
            support["mode"] = "deterministic_fixture_finance_agent_v2_tools"
            support["backend"] = "agent_bench_fixture_deterministic"
            support["backend_available"] = bool(backend["ready"])
            support["backend_canary"] = backend
            support["required_benchmark_tools"] = sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS)
            support["required_environment"] = [] if backend["ready"] else list(FINANCE_AGENT_V2_REQUIRED_ENV)
            support["optional_live_environment"] = list(FINANCE_AGENT_V2_REQUIRED_ENV)
            support["missing_environment"] = [] if backend["ready"] else missing_env
            support["missing_optional_live_environment"] = missing_env
            support["exposed_tools"] = exposed_tools
            support["missing_tools"] = missing_tools
            support["benchmark_required_tools_available"] = not missing_tools
            if missing_tools:
                support["supported"] = False
                support["tools"] = False
                support["reason"] = _finance_agent_v2_backend_unavailable_reason(backend, missing_tools, missing_env)
        return contract


class BrowserGuiAdapter(ToolCallAdapter):
    name = "browser_or_gui"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        support = super().capability_support()
        support["browser_or_gui"] = CapabilitySupport("browser_or_gui", True, True, True, True)
        return support

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        support = contract.get("browser_or_gui")
        if isinstance(support, dict):
            support["mode"] = "text_task_adapter"
            support["reason"] = "Browser/GUI tasks are evaluated from extracted task data and repository files"
        return contract


class FileArtifactAdapter(ChatAnswerAdapter):
    name = "file_artifact"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        support = super().capability_support()
        support["file_artifact"] = CapabilitySupport("file_artifact", True, True, True, True)
        support["office_document_editing"] = CapabilitySupport("office_document_editing", True, True, True, True)
        return support

    def prepare_task(self, item: BenchmarkItem) -> TaskWorkspace:
        source_root = Path.cwd()
        asset_paths = _artifact_asset_paths(item, source_root)
        if not asset_paths:
            raise AdapterSetupError(
                "failed_missing_assets",
                "file_artifact task did not expose required input files or dataset assets",
                {"source": item.source},
            )
        workspace_root = _isolated_workspace_root(item)
        _write_model_visible_task_files(workspace_root, item)
        input_assets = _materialize_artifact_assets(asset_paths, workspace_root, item)
        relative_output_dir = Path("agent_bench_outputs") / _safe_slug(item.source)
        output_dir = workspace_root / relative_output_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        _expose_input_files_at_workspace_root(workspace_root, input_assets)
        _write_task_file_manifest(workspace_root, input_assets)
        return TaskWorkspace(
            root=workspace_root,
            output_dir=output_dir,
            manifest=_workspace_manifest(workspace_root),
            metadata={
                "sanitized_workspace": True,
                "input_assets": input_assets,
                "artifact_output_dir": str(relative_output_dir),
                "item_output_dir": str(_item_output_dir(item)),
                "source": item.source,
                "source_repository_root": str(source_root),
            },
        )

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        if "file_artifact" not in required_capabilities:
            return contract
        assetful_items = sum(1 for item in items if _artifact_asset_paths(item, Path.cwd()))
        support = contract.get("file_artifact")
        if isinstance(support, dict):
            support["assets_provisioned_count"] = assetful_items
            support["required_item_count"] = len(items)
            if not items or assetful_items <= 0:
                support["supported"] = False
                support["workspace"] = False
                support["reason"] = "No required file-artifact assets were provisioned for extracted items"
            else:
                canary = _file_artifact_canary()
                support["canary"] = canary
                if not canary.get("passed"):
                    support["supported"] = False
                    support["tools"] = False
                    support["output_collection"] = False
                    support["reason"] = str(canary.get("reason") or "file_artifact canary failed")
        return contract

    def run_agent_loop(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        augmented = BenchmarkItem(
            question=(
                f"{item.question}\n\n"
                "Generate the required deliverable files under "
                f"{workspace.metadata['artifact_output_dir']}. "
                "Write one or more files inside that directory, for example "
                f"{workspace.metadata['artifact_output_dir']}/response.md; do not write to the directory path itself. "
                "Do not stop after analysis. Before your final answer, create at least one deliverable file with "
                "write_text_file or write_base64_file. If a native workbook/document is not feasible, write a "
                "concrete response.md with the best calculations, summary, code, or plan you can produce."
            ),
            expected=item.expected,
            source=item.source,
            choices=item.choices,
            metadata={**item.metadata, "_file_artifact_task": True},
        )
        return self._run_with_file_artifact_turn_budget(benchmark, augmented, workspace, tools)

    def _run_with_file_artifact_turn_budget(
        self,
        benchmark: str,
        item: BenchmarkItem,
        workspace: TaskWorkspace,
        tools: list[dict[str, Any]],
    ) -> AgentRun:
        previous = os.environ.get("AGENT_BENCH_AGENT_TURNS")
        current = _bounded_int(previous, MAX_AGENT_TURNS, 1, 24) if previous is not None else MAX_AGENT_TURNS
        if current < 12:
            os.environ["AGENT_BENCH_AGENT_TURNS"] = "12"
        try:
            return super().run_agent_loop(benchmark, item, workspace, tools)
        finally:
            if current < 12:
                if previous is None:
                    os.environ.pop("AGENT_BENCH_AGENT_TURNS", None)
                else:
                    os.environ["AGENT_BENCH_AGENT_TURNS"] = previous

    def collect_outputs(self, run: AgentRun, workspace: TaskWorkspace) -> OutputBundle:
        raw_item_dir = str(workspace.metadata.get("item_output_dir") or "")
        if raw_item_dir:
            item_dir = Path(raw_item_dir)
        else:
            item_dir = _item_output_dir(BenchmarkItem("", "", str(workspace.metadata.get("source", "item"))))
        artifact_paths = _collect_artifacts_to_output(workspace.output_dir, item_dir / "artifacts")
        return OutputBundle(
            answer=run.answer,
            artifact_paths=artifact_paths,
            tool_trace=run.tool_trace,
            metadata={
                "content_sample": run.content[:500],
                "diagnostics": run.diagnostics,
                "artifact_output_dir": workspace.metadata.get("artifact_output_dir", ""),
                "artifact_collection_dir": str(item_dir / "artifacts"),
                "input_assets": workspace.metadata.get("input_assets", []),
            },
        )

    def grade(self, benchmark: str, item: BenchmarkItem, outputs: OutputBundle) -> tuple[float, dict[str, Any]]:
        if not outputs.artifact_paths:
            if outputs.answer.strip():
                score, grade = judge_answer(benchmark, item, outputs.answer, "file_artifact")
                grade.setdefault("output_collection", "text_answer_fallback")
                grade.setdefault("artifact_paths", [])
                return score, grade
            return 0.0, {
                "method": "artifact_presence",
                "score": 0.0,
                "status": "failed_model_answer",
                "reason": "model did not produce any files in the task output directory",
            }
        allowed_artifact_root = outputs.metadata.get("artifact_collection_dir") if isinstance(outputs.metadata, dict) else ""
        integrity_errors = _artifact_integrity_errors(
            outputs.artifact_paths,
            allowed_root=Path(allowed_artifact_root) if isinstance(allowed_artifact_root, str) and allowed_artifact_root else None,
        )
        if integrity_errors:
            return 0.0, {
                "method": "artifact_integrity",
                "score": 0.0,
                "status": "failed_model_tool_use",
                "reason": "generated artifact integrity check failed",
                "artifact_errors": integrity_errors,
            }
        artifact_answer = (
            f"Generated artifacts:\n{json.dumps(outputs.artifact_paths, ensure_ascii=False)}\n\n"
            f"Artifact previews:\n{_artifact_previews(outputs.artifact_paths)}\n\n"
            f"Final answer:\n{outputs.answer}"
        )
        score, grade = judge_answer(benchmark, item, artifact_answer, "file_artifact")
        grade.setdefault("output_collection", "file_artifact")
        return score, grade


class RepoPatchAdapter(ChatAnswerAdapter):
    name = "repo_patch"

    def capability_support(self) -> dict[str, CapabilitySupport]:
        support = super().capability_support()
        support["repo_patch"] = CapabilitySupport("repo_patch", True, True, True, True)
        return support

    def capability_contract(self, required_capabilities: set[str], items: list[BenchmarkItem]) -> dict[str, Any]:
        contract = capability_contract_for(required_capabilities, self)
        if "repo_patch" not in required_capabilities:
            return contract
        missing_metadata = [
            item.source
            for item in items
            if not _repo_patch_target_repo(item) or not _repo_patch_base_commit(item)
        ]
        support = contract.get("repo_patch")
        if isinstance(support, dict):
            support["missing_metadata_count"] = len(missing_metadata)
            support["required_item_count"] = len(items)
            support["missing_metadata_sources"] = missing_metadata[:20]
            grader_command = os.environ.get(REPO_PATCH_GRADER_ENV, "").strip()
            swelancer_grader = _swelancer_official_grader_for_items(items)
            if missing_metadata:
                support["supported"] = False
                support["workspace"] = False
                support["reason"] = "repo_patch items are missing target repo/base_commit metadata"
            elif not _repo_patch_checkout_available(items):
                support["supported"] = False
                support["workspace"] = False
                support["reason"] = (
                    "repo_patch target repository checkout is not materialized; "
                    f"set {TARGET_REPO_ROOT_ENV} or AGENT_BENCH_ALLOW_TARGET_CHECKOUT=1"
                )
            else:
                if grader_command:
                    support["official_grader"] = True
                    support["official_equivalent"] = True
                    support["score_mode"] = "official_repo_patch_grader"
                    support["official_grader_command"] = grader_command
                elif swelancer_grader.get("available"):
                    support["official_grader"] = True
                    support["official_equivalent"] = True
                    support["score_mode"] = "official_swelancer_task_tests"
                    support["official_grader_command"] = swelancer_grader.get("command", "")
                    support["swelancer_task_tests"] = swelancer_grader
                else:
                    support["official_grader"] = False
                    support["official_equivalent"] = False
                    support["score_mode"] = "smoke_fallback"
                    support["fallback_grader"] = "model_judge_task_compliance"
                    support["reason"] = (
                        "repo_patch official patch/test grader is not configured; "
                        "using model-judge task-compliance fallback"
                    )
                canary = _repo_patch_canary()
                support["canary"] = canary
                if not canary.get("passed"):
                    support["supported"] = False
                    support["workspace"] = False
                    support["tools"] = False
                    support["output_collection"] = False
                    support["reason"] = str(canary.get("reason") or "repo_patch canary failed")
        return contract

    def prepare_task(self, item: BenchmarkItem) -> TaskWorkspace:
        target_repo = _repo_patch_target_repo(item)
        base_commit = _repo_patch_base_commit(item)
        if not target_repo or not base_commit:
            raise AdapterSetupError(
                "failed_invalid_task_context",
                "repo_patch item is missing target repo/base_commit metadata",
                {"source": item.source, "metadata_keys": sorted(item.metadata)},
            )
        if normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ") == "swe lancer":
            swelancer_error = validate_swelancer_item(item, Path.cwd())
            if swelancer_error is not None:
                status, reason, details = swelancer_error
                raise AdapterSetupError(status, reason, details)
        root = _prepare_target_repo_checkout(item, target_repo, base_commit)
        return TaskWorkspace(
            root=root,
            output_dir=_item_output_dir(item),
            manifest=_workspace_manifest(root),
            metadata={
                "target_repo": target_repo,
                "base_commit": base_commit,
                "target_checkout_path": str(root),
                "instance_id": str(item.metadata.get("instance_id", "")),
            },
        )

    def collect_outputs(self, run: AgentRun, workspace: TaskWorkspace) -> OutputBundle:
        diff = _git_diff(workspace.root)
        patch_path = workspace.output_dir / "model.patch"
        workspace.output_dir.mkdir(parents=True, exist_ok=True)
        patch_path.write_text(diff, encoding="utf-8")
        return OutputBundle(
            answer=run.answer,
            patch=diff,
            artifact_paths=[str(patch_path)],
            tool_trace=run.tool_trace,
            metadata={
                "content_sample": run.content[:500],
                "target_repo": workspace.metadata.get("target_repo", ""),
                "base_commit": workspace.metadata.get("base_commit", ""),
                "target_checkout_path": workspace.metadata.get("target_checkout_path", ""),
                "patch_path": str(patch_path),
            },
        )

    def grade(self, benchmark: str, item: BenchmarkItem, outputs: OutputBundle) -> tuple[float, dict[str, Any]]:
        official_grader = _repo_patch_official_grader_config(benchmark, item)
        if not outputs.patch.strip():
            return 0.0, {
                "score": 0.0,
                "status": "failed_model_missing_artifact",
                "reason": "empty file: model.patch contains no repository diff",
                **_repo_patch_artifact_check_grade_fields(
                    official_grader,
                    fallback_score_mode="smoke_patch_presence",
                    not_run_reason="empty_patch",
                ),
            }
        if not _looks_like_unified_diff(outputs.patch):
            return 0.0, {
                "score": 0.0,
                "status": "failed_model_missing_artifact",
                "reason": "invalid patch format: model.patch is not a unified diff",
                **_repo_patch_artifact_check_grade_fields(
                    official_grader,
                    fallback_score_mode="smoke_patch_presence",
                    not_run_reason="invalid_patch_format",
                ),
            }
        grader_command = os.environ.get(REPO_PATCH_GRADER_ENV, "").strip()
        if grader_command:
            return _run_repo_patch_grader(grader_command, item, outputs)
        if official_grader.get("kind") == "swelancer_task_tests":
            return _run_swelancer_official_grader(item, outputs, official_grader)
        patch_answer = (
            f"Model patch:\n{outputs.patch}\n\n"
            f"Final answer:\n{outputs.answer}"
        )
        score, grade = judge_answer(benchmark, item, patch_answer, "task_compliance")
        grade["method"] = "task_compliance_fallback"
        grade.setdefault("output_collection", "git_diff")
        grade["official_grader"] = False
        grade["official_grader_configured"] = False
        grade["official_equivalent"] = False
        grade["score_mode"] = "smoke_task_compliance_fallback"
        grade["included_in_official_score"] = False
        grade["fallback_grader"] = "model_judge_task_compliance"
        return score, grade


class UnsupportedCapabilityAdapter(BenchmarkAdapter):
    name = "unsupported"

    def __init__(self, capability: str, reason: str) -> None:
        self.capability = capability
        self.reason = reason

    def capability_support(self) -> dict[str, CapabilitySupport]:
        return {
            "chat_answer": CapabilitySupport("chat_answer", True, True, True, True),
            "external_data_required": CapabilitySupport("external_data_required", True, True, True, True),
            self.capability: CapabilitySupport(
                self.capability,
                workspace=False,
                tools=False,
                output_collection=False,
                grader=False,
                native=False,
                reason=self.reason,
            )
        }


ADAPTERS: tuple[BenchmarkAdapter, ...] = (
    ChatAnswerAdapter(),
    StaticTranscriptReasoningAdapter(),
    StaticFinanceReasoningAdapter(),
    FinToolBenchAdapter(),
    FinanceAgentV2Adapter(),
    ToolCallAdapter(),
    BrowserGuiAdapter(),
    FileArtifactAdapter(),
    RepoPatchAdapter(),
)
UNSUPPORTED_CAPABILITY_REASONS = {
    "grader_side_gold_labels": "Gold labels or rubrics are not kept grader-side only",
    "kaggle_competition_submission": "No Kaggle competition execution/submission adapter is implemented",
}
HARNESS_SUPPORTED_CAPABILITIES = {
    capability
    for adapter in ADAPTERS
    for capability in adapter.supported_capabilities()
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Run local benchmark records from a cloned public benchmark.")
    parser.add_argument("--benchmark", required=True)
    parser.add_argument("--kind", default="repository")
    args = parser.parse_args()

    cwd = Path.cwd()
    files = sorted(str(path.relative_to(cwd)) for path in cwd.rglob("*") if path.is_file())[:200]
    markers = [
        name
        for name in (
            "README.md",
            "README.rst",
            "pyproject.toml",
            "requirements.txt",
            "Dockerfile",
            "LICENSE",
            "LICENSE.md",
        )
        if (cwd / name).exists()
    ]
    output_dir = Path(os.environ.get("AGENT_BENCH_OUTPUT_DIR", "/outputs"))
    output_dir.mkdir(parents=True, exist_ok=True)
    required_capabilities = _required_capabilities(args.benchmark)
    adapter = select_adapter(required_capabilities)
    capability_contract = capability_contract_for(required_capabilities, adapter)
    unsupported_capabilities = sorted(required_capabilities - HARNESS_SUPPORTED_CAPABILITIES)
    if unsupported_capabilities:
        payload = _base_result_payload(
            args=args,
            files=files,
            markers=markers,
            sample_limit=_sample_limit(),
            required_capabilities=sorted(required_capabilities),
            unsupported_capabilities=unsupported_capabilities,
            adapter=adapter,
            capability_contract=capability_contract,
        )
        payload.update(
            {
                "score": 0.0,
                "raw_score": 0.0,
                "valid_score": 0.0,
                "status": "skipped_unsupported_capability",
                "error": (
                    "Benchmark requires unsupported capability/capabilities: "
                    + ", ".join(unsupported_capabilities)
                ),
                "extracted_task_count": 0,
                "evaluated_task_count": 0,
                "valid_evaluated_task_count": 0,
                "evaluation_passed_count": 0,
                "skipped_task_count": 1,
                "fallback_task_count": 0,
                "extraction_errors": [],
                "extraction_sources": [],
                "model_evals": [],
                "model_eval": {},
                "status_counts": {"skipped_unsupported_capability": 1},
                "capabilities_verified": False,
                "included_in_official_score": False,
            }
        )
        _write_result(output_dir, payload)
        return 0
    sample_limit = _sample_limit()
    items, extraction_errors = extract_benchmark_items(cwd, sample_limit)
    if not items and _allow_readiness_fallback():
        fallback = readiness_fallback_item(cwd, args.benchmark)
        if fallback is not None:
            items.append(fallback)
            extraction_errors.append(
                "Used repository-readiness fallback because no machine-readable benchmark rows were found"
            )
    capability_contract = adapter.capability_contract(required_capabilities, items)
    preflight_status, preflight_reason = _preflight_failure_from_contract(required_capabilities, capability_contract)
    if items and preflight_status:
        evaluations = preflight_failed_evaluations(items, adapter, preflight_status, preflight_reason, capability_contract)
    else:
        valid_items, preflight_evaluations = validate_items_preflight(
            items,
            adapter,
            required_capabilities,
            capability_contract,
        )
        model_evaluations = run_model_evaluations(args.benchmark, valid_items, adapter) if valid_items else []
        evaluations = preflight_evaluations + model_evaluations
    evaluated_count = len(evaluations)
    valid_evaluations = _valid_evaluations(evaluations)
    raw_score = _average_score(evaluations)
    score = _average_score(valid_evaluations)
    error_message = ""
    if not items:
        error_message = extraction_errors[0] if extraction_errors else "No benchmark task records were found in the cloned repository"
    elif not evaluations and _is_remote_provider():
        error_message = "No model evaluations completed"
    status, status_error = _overall_status_and_error(evaluations, error_message)
    if status_error:
        error_message = status_error

    payload = _base_result_payload(
        args=args,
        files=files,
        markers=markers,
        sample_limit=sample_limit,
        required_capabilities=sorted(required_capabilities),
        unsupported_capabilities=sorted(
            set(
                _payload_unsupported_capabilities(
                    required_capabilities,
                    capability_contract,
                    status,
                    error_message,
                    args.benchmark,
                )
            )
            | set(_evaluation_unsupported_capabilities(evaluations))
        ),
        adapter=adapter,
        capability_contract=capability_contract,
    )
    payload.update(
        {
            "score": score,
            "raw_score": raw_score,
            "valid_score": score,
            "status": status,
            "extracted_task_count": len(items),
            "evaluated_task_count": evaluated_count,
            "valid_evaluated_task_count": len(valid_evaluations),
            "evaluation_passed_count": sum(1 for item in evaluations if item.get("passed")),
            "skipped_task_count": sum(
                1
                for item in evaluations
                if normalize_status(item.get("status"))
                in {"failed_missing_assets", "skipped_unsupported_capability"}
            ),
            "grader_failure_count": sum(
                1 for item in evaluations if normalize_status(item.get("status")) == "failed_grader"
            ),
            "judge_parse_failure_count": sum(
                1 for item in evaluations if normalize_status(item.get("status")) == "failed_grader"
            ),
            "judge_retry_count": sum(int(item.get("judge_retry_count") or 0) for item in evaluations),
            "judge_parse_repaired_count": sum(1 for item in evaluations if item.get("judge_parse_repaired")),
            "fallback_task_count": sum(1 for item in items if item.metadata.get("fallback")),
            "extraction_errors": extraction_errors[:20],
            "extraction_sources": sorted({item.source for item in items}),
            "model_evals": evaluations,
            "model_eval": summarize_evaluations(evaluations),
            "status_counts": _status_counts(evaluations) or ({status: 1} if status != "completed" else {}),
            "capabilities_verified": (
                bool(evaluations)
                and all(bool(item.get("capabilities_verified", True)) for item in evaluations)
                and _capability_profile(required_capabilities, capability_contract)[
                    "benchmark_required_tools_available"
                ]
            ),
        }
    )
    if error_message:
        payload["error"] = error_message
    _promote_evaluation_tool_fields(payload, evaluations)
    _promote_evaluation_workspace_fields(payload, evaluations)
    _promote_evaluation_official_score_fields(payload, evaluations)
    payload["included_in_official_score"] = (
        normalize_status(status) not in INVALID_EVALUATION_STATUSES
        and bool(payload.get("capabilities_verified"))
        and not _evaluation_excluded_from_official_score(evaluations)
    )

    _write_result(output_dir, payload)
    return 0 if status == "completed" else 2


def _promote_evaluation_tool_fields(payload: dict[str, Any], evaluations: list[dict[str, Any]]) -> None:
    for key in ("required_tools", "exposed_tools", "missing_tools", "missing_env", "missing_environment"):
        values = _flatten_string_values(payload.get(key))
        for evaluation in evaluations:
            if isinstance(evaluation, dict):
                values.extend(_flatten_string_values(evaluation.get(key)))
                setup = evaluation.get("setup_details")
                if isinstance(setup, dict):
                    values.extend(_flatten_string_values(setup.get(key)))
                    details = setup.get("details")
                    if isinstance(details, dict):
                        values.extend(_flatten_string_values(details.get(key)))
        payload[key] = sorted({value for value in values if value})


def _promote_evaluation_workspace_fields(payload: dict[str, Any], evaluations: list[dict[str, Any]]) -> None:
    for key in ("target_checkout_path", "target_repo", "base_commit"):
        if payload.get(key):
            continue
        for evaluation in evaluations:
            if not isinstance(evaluation, dict):
                continue
            output_bundle = evaluation.get("output_bundle")
            metadata = output_bundle.get("metadata") if isinstance(output_bundle, dict) else None
            if isinstance(metadata, dict) and metadata.get(key):
                payload[key] = metadata[key]
                break
            item_metadata = evaluation.get("metadata")
            if isinstance(item_metadata, dict) and item_metadata.get(key):
                payload[key] = item_metadata[key]
                break


def _promote_evaluation_official_score_fields(payload: dict[str, Any], evaluations: list[dict[str, Any]]) -> None:
    score_modes = sorted(
        {
            str(evaluation.get("score_mode"))
            for evaluation in evaluations
            if isinstance(evaluation, dict) and str(evaluation.get("score_mode") or "").strip()
        }
    )
    if score_modes:
        payload["score_modes"] = score_modes
        payload["score_mode"] = score_modes[0] if len(score_modes) == 1 else "mixed"
    official_equivalent_values = [
        evaluation.get("official_equivalent")
        for evaluation in evaluations
        if isinstance(evaluation, dict) and isinstance(evaluation.get("official_equivalent"), bool)
    ]
    if official_equivalent_values:
        payload["official_equivalent"] = all(official_equivalent_values)


def _evaluation_excluded_from_official_score(evaluations: list[dict[str, Any]]) -> bool:
    return any(
        isinstance(evaluation, dict) and evaluation.get("included_in_official_score") is False
        for evaluation in evaluations
    )


def _base_result_payload(
    *,
    args: argparse.Namespace,
    files: list[str],
    markers: list[str],
    sample_limit: int,
    required_capabilities: list[str],
    unsupported_capabilities: list[str],
    adapter: BenchmarkAdapter,
    capability_contract: dict[str, Any],
) -> dict[str, Any]:
    capability_profile = _capability_profile(set(required_capabilities), capability_contract)
    required_tools = _contract_required_tools(capability_contract)
    exposed_tools = _contract_exposed_tools(capability_contract)
    missing_tools = _contract_missing_tools(capability_contract)
    missing_env = _contract_missing_env(capability_contract)
    return {
        "benchmark": args.benchmark,
        "group": os.environ.get("AGENT_BENCH_BENCHMARK_GROUP") or "Benchmarks",
        "kind": args.kind,
        "adapter": adapter.name,
        "repository": os.environ.get("AGENT_BENCH_REPOSITORY", ""),
        "repository_ref": os.environ.get("AGENT_BENCH_REPOSITORY_REF", ""),
        "subdir": os.environ.get("AGENT_BENCH_SUBDIR", ""),
        "homepage": os.environ.get("AGENT_BENCH_BENCHMARK_HOMEPAGE", ""),
        "license": os.environ.get("AGENT_BENCH_BENCHMARK_LICENSE", ""),
        "credit": os.environ.get("AGENT_BENCH_BENCHMARK_CREDIT", ""),
        "citation": os.environ.get("AGENT_BENCH_BENCHMARK_CITATION", ""),
        "dataset_id": os.environ.get("AGENT_BENCH_DATASET_ID", ""),
        "model": os.environ.get("AGENT_BENCH_MODEL", ""),
        "repository_ready": bool(files),
        "file_count_sampled": len(files),
        "markers": markers,
        "sample_files": files,
        "sample_limit": sample_limit,
        "required_capabilities": required_capabilities,
        "supported_capabilities": sorted(
            capability
            for capability in required_capabilities
            if isinstance(capability_contract.get(capability), dict)
            and capability_contract[capability].get("supported") is True
            and capability_contract[capability].get("benchmark_required_tools_available") is not False
        ),
        "required_tools": required_tools,
        "exposed_tools": exposed_tools,
        "missing_tools": missing_tools,
        "missing_env": missing_env,
        "native_tool_call_protocol_supported": capability_profile["native_tool_call_protocol_supported"],
        "generic_sandbox_tools_available": capability_profile["generic_sandbox_tools_available"],
        "benchmark_required_tools_available": capability_profile["benchmark_required_tools_available"],
        "static_degraded_mode": capability_profile["static_degraded_mode"],
        "capability_contract": capability_contract,
        "unsupported_capabilities": unsupported_capabilities,
    }


def _capability_profile(
    required_capabilities: set[str],
    capability_contract: dict[str, Any],
) -> dict[str, bool]:
    requires_tool_call = "tool_call" in required_capabilities
    contract_values = [value for value in capability_contract.values() if isinstance(value, dict)]
    benchmark_required_tools_available = all(
        value.get("benchmark_required_tools_available") is not False for value in contract_values
    )
    return {
        "native_tool_call_protocol_supported": (not requires_tool_call)
        or "tool_call" in HARNESS_SUPPORTED_CAPABILITIES,
        "generic_sandbox_tools_available": bool(agent_tool_schemas()),
        "benchmark_required_tools_available": benchmark_required_tools_available,
        "static_degraded_mode": any(value.get("static_degraded_mode") is True for value in contract_values),
    }


def _payload_unsupported_capabilities(
    required_capabilities: set[str],
    capability_contract: dict[str, Any],
    status: str,
    error_message: str,
    benchmark: str,
) -> list[str]:
    unsupported = set(_contract_unsupported_capabilities(required_capabilities, capability_contract))
    if status == "skipped_unsupported_capability":
        unsupported.update(_implicit_unsupported_capabilities(benchmark, error_message))
    return sorted(unsupported)


def _contract_unsupported_capabilities(
    required_capabilities: set[str],
    capability_contract: dict[str, Any],
) -> list[str]:
    unsupported: list[str] = []
    for capability in sorted(required_capabilities):
        support = capability_contract.get(capability)
        if isinstance(support, dict) and (
            support.get("supported") is False
            or support.get("benchmark_required_tools_available") is False
        ):
            unsupported.append(capability)
    return unsupported


def _implicit_unsupported_capabilities(benchmark: str, error_message: str) -> list[str]:
    lowered_benchmark = normalize_text(benchmark).replace("-", " ")
    lowered_error = error_message.lower()
    if lowered_benchmark == "biomystery bench" and "scoring is disabled" in lowered_error:
        return ["grader_side_gold_labels"]
    return []


def _evaluation_unsupported_capabilities(evaluations: list[dict[str, Any]]) -> list[str]:
    capabilities: set[str] = set()
    for item in evaluations:
        setup = item.get("setup_details")
        if not isinstance(setup, dict):
            continue
        for container in (setup, setup.get("details")):
            if not isinstance(container, dict):
                continue
            raw = container.get("unsupported_capability") or container.get("unsupported_capabilities")
            if isinstance(raw, str) and raw.strip():
                capabilities.add(raw.strip())
            elif isinstance(raw, list):
                capabilities.update(str(value).strip() for value in raw if str(value).strip())
    return sorted(capabilities)


def _write_result(output_dir: Path, payload: dict[str, Any]) -> None:
    (output_dir / "agent_bench_result.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def select_adapter(required_capabilities: set[str]) -> BenchmarkAdapter:
    benchmark = normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ")
    if benchmark == "finmcp bench":
        return StaticTranscriptReasoningAdapter()
    if benchmark in {"fintoolbench", "finance agent v2"} and _static_conversion_live_tools_disabled():
        return StaticFinanceReasoningAdapter()
    if benchmark == "fintoolbench":
        return FinToolBenchAdapter()
    if benchmark == "finance agent v2":
        return FinanceAgentV2Adapter()
    if "browser_or_gui" in required_capabilities:
        return BrowserGuiAdapter()
    if "tool_call" in required_capabilities:
        return ToolCallAdapter()
    if "repo_patch" in required_capabilities:
        return RepoPatchAdapter()
    if "file_artifact" in required_capabilities or "office_document_editing" in required_capabilities:
        return FileArtifactAdapter()
    return ChatAnswerAdapter()


def capability_contract_for(required_capabilities: set[str], adapter: BenchmarkAdapter) -> dict[str, Any]:
    support = adapter.capability_support()
    contract: dict[str, Any] = {}
    for capability in sorted(required_capabilities):
        if capability in support:
            contract[capability] = support[capability].to_dict()
            continue
        reason = UNSUPPORTED_CAPABILITY_REASONS.get(capability, "No adapter advertises this capability")
        contract[capability] = CapabilitySupport(
            capability,
            workspace=False,
            tools=False,
            output_collection=False,
            grader=False,
            native=False,
            reason=reason,
        ).to_dict()
    return contract


def _annotate_tool_contract(
    contract: dict[str, Any],
    required_capabilities: set[str],
    items: list[BenchmarkItem],
) -> dict[str, Any]:
    required_tools = sorted(
        {
            tool
            for item in items
            for tool in _item_required_tools(item)
        }
        | set(_descriptor_required_tools())
    )
    if not required_tools and "tool_call" not in required_capabilities:
        return contract
    exposed_tools = _tool_schema_names(agent_tool_schemas())
    missing_tools = sorted(set(required_tools) - set(exposed_tools))
    support = contract.get("tool_call")
    if not isinstance(support, dict):
        return contract
    support.setdefault("mode", "generic_agent_tools")
    support["required_benchmark_tools"] = required_tools
    support["exposed_tools"] = exposed_tools
    support["missing_tools"] = missing_tools
    support["benchmark_required_tools_available"] = not missing_tools
    if missing_tools:
        support["supported"] = False
        support["tools"] = False
        support["reason"] = "Benchmark required tool schemas are missing: " + ", ".join(missing_tools)
    elif "tool_call" in required_capabilities and not exposed_tools:
        support["supported"] = False
        support["tools"] = False
        support["benchmark_required_tools_available"] = False
        support["missing_tools"] = ["tool_call_backend"]
        support["reason"] = "Benchmark requires tool_call but no callable tool backend is exposed"
    return contract


def _required_capabilities(benchmark: str) -> set[str]:
    configured = _capabilities_from_env()
    if configured:
        return configured
    normalized = normalize_text(benchmark).replace("-", " ")
    capability_map = {
        "swe bench": {"repo_patch"},
        "swe bench verified": {"repo_patch"},
        "swe lancer": {"repo_patch"},
        "gdpval": {"file_artifact", "external_data_required", "office_document_editing"},
        "paperbench": {"file_artifact"},
        "automationbench": {"tool_call", "external_data_required"},
        "osworld": {"browser_or_gui"},
        "exploitbench": {"tool_call"},
        "finmcp bench": {"chat_answer"},
        "fintoolbench": {"tool_call", "external_data_required"},
        "finance agent v2": {"tool_call", "external_data_required"},
    }
    for key, capabilities in capability_map.items():
        if normalized == key:
            return set(capabilities)
    return {"chat_answer"}


def _item_required_capabilities(item: BenchmarkItem) -> set[str]:
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    configured = metadata.get("required_capabilities")
    if isinstance(configured, list):
        values = {str(value).strip() for value in configured if str(value).strip()}
        if values:
            return values
    return _required_capabilities(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", ""))


def _evaluation_preflight_fields(
    item: BenchmarkItem,
    adapter: BenchmarkAdapter,
    tools: list[dict[str, Any]] | None = None,
    missing_tools: list[str] | None = None,
    capability_contract: dict[str, Any] | None = None,
) -> dict[str, Any]:
    required_capabilities = _item_required_capabilities(item)
    exposed_tools = _tool_schema_names(tools or [])
    return {
        "required_capabilities": sorted(required_capabilities),
        "supported_capabilities": _supported_capabilities_for_item(required_capabilities, adapter, capability_contract),
        "required_tools": _item_required_tools(item),
        "exposed_tools": exposed_tools,
        "missing_tools": sorted(set(missing_tools or [])),
    }


def _supported_capabilities_for_item(
    required_capabilities: set[str],
    adapter: BenchmarkAdapter,
    capability_contract: dict[str, Any] | None = None,
) -> list[str]:
    if capability_contract:
        return sorted(
            capability
            for capability in required_capabilities
            if isinstance(capability_contract.get(capability), dict)
            and capability_contract[capability].get("supported") is True
            and capability_contract[capability].get("benchmark_required_tools_available") is not False
        )
    supported = adapter.supported_capabilities()
    return sorted(capability for capability in required_capabilities if capability in supported)


def _static_conversion_live_tools_disabled() -> bool:
    benchmark = _benchmark_descriptor()
    adapter_mode = str(benchmark.get("adapter_mode") or benchmark.get("mode") or "").strip().lower()
    if adapter_mode in {
        "static",
        "static_gold_answer",
        "static_public_prompt",
        "static_transcript",
        "static_transcript_reasoning",
    }:
        return True
    if benchmark.get("live_tools_required") is False:
        return True
    static_conversion = benchmark.get("static_conversion")
    if isinstance(static_conversion, dict):
        static_mode = str(static_conversion.get("adapter_mode") or static_conversion.get("mode") or "").strip().lower()
        if static_mode in {
            "static",
            "static_gold_answer",
            "static_public_prompt",
            "static_transcript",
            "static_transcript_reasoning",
        }:
            return True
        return static_conversion.get("live_tools_required") is False
    return False


def _benchmark_descriptor() -> dict[str, Any]:
    for env_name in ("AGENT_BENCH_BENCHMARK_JSON", "AGENT_BENCH_MANIFEST_JSON"):
        raw = os.environ.get(env_name, "")
        if not raw.strip():
            continue
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict):
            return payload
    return {}


def _descriptor_required_tools() -> list[str]:
    descriptor = _benchmark_descriptor()
    tools: list[str] = []
    for key in ("required_tools", "tool_names"):
        tools.extend(_flatten_string_values(descriptor.get(key)))
    declared_tools = descriptor.get("tools")
    if isinstance(declared_tools, list):
        for entry in declared_tools:
            if isinstance(entry, str):
                tools.append(entry)
            elif isinstance(entry, dict):
                for key in ("name", "tool", "tool_name", "api_name"):
                    value = entry.get(key)
                    if isinstance(value, str) and value.strip():
                        tools.append(value.strip())
                function = entry.get("function")
                if isinstance(function, dict) and isinstance(function.get("name"), str):
                    tools.append(function["name"].strip())
    dataset_source = descriptor.get("dataset_source")
    if isinstance(dataset_source, dict):
        tools.extend(_flatten_string_values(dataset_source.get("required_tools")))
    assets = descriptor.get("assets")
    if isinstance(assets, list):
        for asset in assets:
            if not isinstance(asset, dict):
                continue
            tools.extend(_flatten_string_values(asset.get("required_tools")))
            rules = asset.get("validation_rules")
            if isinstance(rules, dict):
                tools.extend(_flatten_string_values(rules.get("required_tools")))
    return sorted({tool.strip() for tool in tools if tool.strip()})


def _capabilities_from_env() -> set[str]:
    raw = os.environ.get("AGENT_BENCH_REQUIRED_CAPABILITIES", "")
    return {item.strip() for item in raw.split(",") if item.strip()}


def extract_benchmark_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    specialized_items, specialized_errors = extract_specialized_items(root, limit - len(items))
    items.extend(specialized_items)
    errors.extend(specialized_errors)
    if _specialized_extraction_is_terminal() and (items or errors):
        return items[:limit], errors[:20]
    if not _skip_huggingface_dataset_loader():
        hf_items, hf_errors = extract_huggingface_items(limit - len(items))
        items.extend(hf_items)
        errors.extend(hf_errors)
    for path in _candidate_files(root):
        if len(items) >= limit:
            break
        try:
            new_items = extract_items_from_file(path, root, limit - len(items))
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        items.extend(new_items)
    return items, errors


def _specialized_extraction_is_terminal() -> bool:
    benchmark = normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ")
    return benchmark in {
        "biomystery bench",
        "codeneedle",
        "paperbench",
        "stockbench",
        "exploitbench",
        "finmcp bench",
        "fintoolbench",
        "finance agent v2",
        "swe lancer",
        "quantcode bench",
    }


def readiness_fallback_item(root: Path, benchmark: str) -> BenchmarkItem | None:
    excerpt_path = _first_existing_text(
        root,
        (
            "README.md",
            "README.rst",
            "docs/README.md",
        ),
    )
    if excerpt_path is None:
        return None
    text = excerpt_path.read_text(encoding="utf-8", errors="replace").strip()
    if len(text) < 80:
        return None
    question = (
        f"{benchmark} public benchmark local-readiness task.\n"
        "The lightweight Docker adapter did not find a standalone answer-key file in this clone. "
        "Using the public benchmark excerpt below, summarize the benchmark task format, local execution "
        "requirements, and what a valid agent submission is expected to produce.\n\n"
        f"Source: {excerpt_path.relative_to(root)}\n\n"
        f"{truncate(text, 4200)}"
    )
    return BenchmarkItem(
        question=truncate(question),
        expected=(
            "Candidate answer should accurately summarize the public benchmark task format, local "
            "execution requirements, and expected agent deliverable from the provided source excerpt."
        ),
        source=f"{excerpt_path.relative_to(root)}:readiness-fallback",
        metadata={"grading": "task_compliance", "fallback": True, "expected_key": "readiness_summary"},
    )


def _allow_readiness_fallback() -> bool:
    return os.environ.get("AGENT_BENCH_ALLOW_READINESS_FALLBACK", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def _first_existing_text(root: Path, relative_paths: tuple[str, ...]) -> Path | None:
    for relative_path in relative_paths:
        path = root / relative_path
        if path.is_file():
            return path
    return None


def _candidate_files(root: Path) -> list[Path]:
    suffix_priority = {
        ".jsonl": 0,
        ".json": 1,
        ".csv": 2,
        ".parquet": 3,
        ".py": 4,
        ".yaml": 5,
        ".yml": 5,
        ".md": 6,
        ".txt": 7,
    }
    candidates: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in suffix_priority:
            continue
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        if _skip_candidate_file(path):
            continue
        try:
            if path.stat().st_size == 0:
                continue
        except OSError:
            continue
        candidates.append(path)
    return sorted(
        candidates,
        key=lambda path: (
            _path_relevance(path),
            suffix_priority[path.suffix.lower()],
            len(path.parts),
            str(path),
        ),
    )


def _skip_candidate_file(path: Path) -> bool:
    text = str(path).lower()
    filename = path.name.lower()
    if filename in {"package-lock.json", "pnpm-lock.yaml", "uv.lock", "poetry.lock", "benchmark_plan.md"}:
        return True
    if path.suffix.lower() == ".py" and filename != "tasks.py":
        return True
    if path.suffix.lower() in {".yaml", ".yml"} and "/prompt/" not in text and "prompt" not in filename:
        return True
    if path.suffix.lower() in {".md", ".txt", ".yaml", ".yml"} and not _is_prompt_text_file(path):
        return True
    return any(
        marker in text
        for marker in (
            "model_pricing",
            "pricing.json",
            "cache",
            ".pytest_cache",
            "node_modules",
            "package.json",
            "tsconfig",
        )
    )


def _is_prompt_text_file(path: Path) -> bool:
    text = str(path).lower()
    filename = path.name.lower()
    if filename in {"readme.md", "license", "security.md", "contributing.md", "requirements.txt"}:
        return False
    return any(
        marker in text
        for marker in (
            "description.md",
            "description_obfuscated.md",
            "spec.md",
            "/prompts/",
            "/prompt/",
            "research_problem.txt",
            "instructions.txt",
            "task",
            "question",
            "problem",
            "benchmark",
        )
    )


def _path_relevance(path: Path) -> int:
    text = str(path).lower()
    if any(word in text for word in ("test", "validation", "eval", "benchmark", "question", "problem", "task", "prompt")):
        return 0
    if any(word in text for word in ("description.md", "spec.md", "instructions.txt")):
        return 1
    if any(word in text for word in ("config", "pricing", "cache", "knowledge", "example")):
        return 2
    return 1


def extract_huggingface_items(limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    if limit <= 0:
        return [], []
    dataset_ids = _huggingface_dataset_ids()
    if not dataset_ids:
        return [], []
    try:
        from datasets import load_dataset
    except ImportError as exc:
        return [], [f"Hugging Face streaming requires the datasets package ({exc})"]

    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for dataset_id in dataset_ids:
        if len(items) >= limit:
            break
        extracted, dataset_errors = _extract_huggingface_dataset_items(
            load_dataset,
            dataset_id,
            limit - len(items),
        )
        items.extend(extracted)
        errors.extend(dataset_errors)
    return items, errors


def _huggingface_dataset_ids() -> list[str]:
    candidates: list[str] = []
    explicit = os.environ.get("AGENT_BENCH_DATASET_ID", "").strip()
    if explicit:
        candidates.append(explicit)
    repository = os.environ.get("AGENT_BENCH_REPOSITORY", "")
    marker = "huggingface.co/datasets/"
    if marker in repository:
        dataset_id = repository.split(marker, 1)[1].strip("/")
        if dataset_id:
            candidates.append(dataset_id)
    unique: list[str] = []
    for candidate in candidates:
        dataset_id = candidate.split("?", 1)[0].split("#", 1)[0].strip("/")
        if dataset_id and dataset_id not in unique:
            unique.append(dataset_id)
    return unique


def _skip_huggingface_dataset_loader() -> bool:
    benchmark = os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")
    return normalize_text(benchmark).replace("-", " ") in {"biomystery bench"}


def _extract_huggingface_dataset_items(
    load_dataset: Any,
    dataset_id: str,
    limit: int,
) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for split in ("test", "validation", "dev", "train"):
        if len(items) >= limit:
            break
        try:
            dataset = load_dataset(dataset_id, split=split, streaming=True)
        except Exception as exc:
            errors.append(f"{dataset_id}/{split}: {exc}")
            continue
        items.extend(
            _items_from_huggingface_iterable(
                dataset,
                f"huggingface:{dataset_id}/{split}",
                limit - len(items),
            )
        )
    if items:
        return items, errors
    try:
        dataset_dict = load_dataset(dataset_id, streaming=True)
    except Exception as exc:
        errors.append(f"{dataset_id}: {exc}")
        return items, errors
    if hasattr(dataset_dict, "items"):
        for split, dataset in dataset_dict.items():
            if len(items) >= limit:
                break
            items.extend(
                _items_from_huggingface_iterable(
                    dataset,
                    f"huggingface:{dataset_id}/{split}",
                    limit - len(items),
                )
            )
    else:
        items.extend(_items_from_huggingface_iterable(dataset_dict, f"huggingface:{dataset_id}", limit))
    return items, errors


def _items_from_huggingface_iterable(
    dataset: Any,
    source_prefix: str,
    limit: int,
) -> list[BenchmarkItem]:
    items: list[BenchmarkItem] = []
    for index, row in enumerate(dataset):
        if len(items) >= limit or index >= MAX_RECORDS_PER_FILE:
            break
        item = item_from_record(row, f"{source_prefix}:{index + 1}")
        if item is not None:
            items.append(item)
    return items


def extract_specialized_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    benchmark = os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")
    if limit <= 0:
        return [], []
    if benchmark == "BioMystery Bench":
        return extract_biomystery_items(root, limit)
    if benchmark == "codeneedle":
        return extract_codeneedle_items(root, limit)
    if benchmark == "StockBench":
        return extract_stockbench_items(root, limit)
    if benchmark == "ExploitBench":
        return extract_exploitbench_items(root, limit)
    if benchmark == "FinMCP-Bench":
        return extract_finmcp_static_items(root, limit)
    if benchmark == "FinToolBench":
        return extract_fintoolbench_items(root, limit)
    if benchmark == "Finance Agent v2":
        return extract_finance_agent_v2_items(root, limit)
    if benchmark == "InvestorBench":
        return extract_investorbench_items(root, limit)
    if benchmark == "PaperBench":
        return extract_paperbench_items(root, limit)
    if benchmark == "SWE-Lancer":
        return extract_swelancer_items(root, limit)
    if benchmark == "QuantCode-Bench":
        return extract_quantcode_items(root, limit)
    if benchmark == "Humanity's Last Exam":
        return [], ["Humanity's Last Exam requires accessible dataset records; format-only fallback is disabled"]
    return [], []


def extract_biomystery_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    seen: set[str] = set()
    data_sources = 0

    for archive_path in sorted(root.rglob("*.zip")):
        if len(items) >= limit:
            break
        data_sources += 1
        try:
            with zipfile.ZipFile(archive_path) as archive:
                for member in sorted(archive.namelist()):
                    if len(items) >= limit:
                        break
                    if member.endswith("/"):
                        continue
                    suffix = Path(member).suffix.lower()
                    if suffix not in {".json", ".jsonl", ".csv"}:
                        continue
                    try:
                        text = archive.read(member).decode("utf-8", errors="replace")
                    except Exception as exc:
                        errors.append(f"{archive_path.relative_to(root)}!{member}: {exc}")
                        continue
                    source = f"{archive_path.relative_to(root)}!{member}"
                    _extend_unique_items(items, _biomystery_items_from_text(text, source, limit - len(items)), seen)
        except zipfile.BadZipFile as exc:
            errors.append(f"{archive_path.relative_to(root)}: invalid zip archive: {exc}")
        except Exception as exc:
            errors.append(f"{archive_path.relative_to(root)}: {exc}")

    for path in sorted(root.glob("problems.*")) + sorted(root.glob("*biomystery*.*")):
        if len(items) >= limit:
            break
        if not path.is_file():
            continue
        data_sources += 1
        try:
            _extend_unique_items(items, extract_items_from_file(path, root, limit - len(items)), seen)
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")

    if items:
        return items[:limit], errors[:20]
    if data_sources <= 0:
        return [], ["BioMystery Bench local data files were not found"]
    return [], (errors or ["BioMystery Bench records with answer_rubric were not found"])[:20]


def _biomystery_items_from_text(text: str, source: str, limit: int) -> list[BenchmarkItem]:
    suffix = Path(source).suffix.lower()
    if suffix == ".json":
        payload = json.loads(text)
        return _items_from_json_payload(payload, Path(source), limit)
    if suffix == ".jsonl":
        items: list[BenchmarkItem] = []
        for index, line in enumerate(text.splitlines(), 1):
            if len(items) >= limit or index > MAX_RECORDS_PER_FILE:
                break
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            item = item_from_record(record, f"{source}:{index}")
            if item is not None:
                items.append(item)
        return items
    if suffix == ".csv":
        items = []
        reader = csv.DictReader(text.splitlines())
        for index, row in enumerate(reader, 2):
            if len(items) >= limit or index > MAX_RECORDS_PER_FILE + 1:
                break
            item = item_from_record(row, f"{source}:{index}")
            if item is not None:
                items.append(item)
        return items
    return []


def _extend_unique_items(target: list[BenchmarkItem], candidates: list[BenchmarkItem], seen: set[str]) -> None:
    for item in candidates:
        key = hashlib.sha256(f"{item.question}\0{item.expected}".encode("utf-8")).hexdigest()
        if key in seen:
            continue
        seen.add(key)
        target.append(item)


def _data_record_files(root: Path) -> list[Path]:
    suffixes = {".jsonl", ".json", ".csv"}
    paths: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in suffixes:
            continue
        lowered = str(path.relative_to(root)).lower()
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        if any(marker in lowered for marker in ("readme", "package-lock", "pricing", "model_pricing")):
            continue
        paths.append(path)
    return sorted(paths, key=lambda path: (_path_relevance(path), len(path.parts), str(path)))


def _records_from_data_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        return _records_from_jsonl(path)
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        return [record for record in walk_records(payload) if isinstance(record, dict)][:MAX_RECORDS_PER_FILE]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)][:MAX_RECORDS_PER_FILE]
    return []


def _records_from_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            if len(records) >= MAX_RECORDS_PER_FILE:
                break
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(record, dict):
                records.append(record)
    return records


def _required_tools_from_record(record: dict[str, Any]) -> list[str]:
    tools: list[str] = []
    lower_to_key = {str(key).lower(): key for key in record}
    for key in (
        "required_tools",
        "select_tools",
        "selected_tools",
        "selected_tool",
        "tools",
        "tool_names",
        "gold_tools",
        "api_name",
        "api_names",
    ):
        original = lower_to_key.get(key)
        if original is not None:
            tools.extend(_flatten_string_values(record.get(original)))
    return sorted({tool.strip() for tool in tools if tool.strip()})


def extract_paperbench_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    rubric_paths = sorted(root.rglob("rubric.json"))
    for path in rubric_paths:
        if len(items) >= limit:
            break
        try:
            payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        nodes = _paperbench_rubric_nodes(payload)
        for index, node in enumerate(nodes, 1):
            if len(items) >= limit:
                break
            record = dict(node["record"])
            record["paperbench_parent_requirements"] = node["parents"]
            item = _paperbench_item_from_record(record, f"{path.relative_to(root)}:leaf-{index:06d}")
            if item is not None:
                items.append(item)
    return items, errors


def _paperbench_item_from_record(record: dict[str, Any], source: str) -> BenchmarkItem | None:
    question = _text_from_value(record.get("requirements")).strip()
    if not question:
        return None
    expected = "Candidate answer should satisfy the task requirements from the benchmark prompt."
    metadata = _record_metadata(record, source, "task_compliance", "", expected)
    return BenchmarkItem(
        question=truncate(question),
        expected=expected,
        source=source,
        metadata=metadata,
    )


def _paperbench_rubric_nodes(payload: Any) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    order = 0

    def visit(value: Any, parents: list[str], depth: int) -> None:
        nonlocal order
        if len(nodes) >= MAX_RECORDS_PER_FILE:
            return
        if isinstance(value, list):
            for item in value:
                visit(item, parents, depth)
            return
        if not isinstance(value, dict):
            return
        requirements = str(value.get("requirements") or "").strip()
        children = [item for item in value.get("sub_tasks", []) if isinstance(item, dict)]
        if requirements:
            order += 1
            nodes.append(
                {
                    "record": value,
                    "parents": parents,
                    "depth": depth,
                    "leaf": not children,
                    "order": order,
                }
            )
        next_parents = parents + ([requirements] if requirements else [])
        for child in children:
            visit(child, next_parents, depth + 1)

    visit(payload, [], 0)
    return sorted(nodes, key=lambda item: (not item["leaf"], -int(item["depth"]), int(item["order"])))


def extract_codeneedle_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for path in sorted((root / "fixtures").glob("*")):
        if len(items) >= limit:
            break
        if not path.is_file() or _is_binary_like_file(path):
            continue
        text = path.read_text(encoding="utf-8", errors="replace")
        needle = _codeneedle_target_line(text)
        if not needle:
            continue
        question = (
            "codeneedle retrieval task. Read the code context below and return the exact target line "
            "that defines or starts the most salient function/API entry point.\n\n"
            f"Source: fixtures/{path.name}\n\n"
            f"{truncate(text, 5000)}"
        )
        items.append(
            BenchmarkItem(
                question=truncate(question),
                expected=needle,
                source=f"fixtures/{path.name}:codeneedle",
                metadata={"grading": "exact", "expected_key": "fixture_needle"},
            )
        )
    if not items:
        errors.append("codeneedle fixture/context files were not found")
    return items, errors[:20]


def _codeneedle_target_line(text: str) -> str:
    for line in text.splitlines():
        stripped = line.strip()
        if re.match(r"(?:async\s+)?def\s+\w+|function\s+\w+|class\s+\w+", stripped):
            return stripped
    for line in text.splitlines():
        stripped = line.strip()
        if stripped and not stripped.startswith(("#", "//", "/*", "*")):
            return stripped
    return ""


def extract_stockbench_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    cache_root = root / "storage" / "cache"
    if not cache_root.is_dir():
        return [], ["StockBench cache/task input directory was not found"]
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for path in sorted((cache_root / "financials").glob("*.annual.json")):
        if len(items) >= limit:
            break
        try:
            payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        symbol = path.name.split(".", 1)[0]
        compact = truncate(json.dumps(payload, ensure_ascii=False, sort_keys=True), 3500)
        question = (
            f"StockBench concrete financial-analysis task for {symbol}.\n"
            "Using the compact cached annual financial JSON below, identify whether the company appears "
            "fundamentally stronger, weaker, or mixed based on the latest available annual data. "
            "Return one label: stronger, weaker, or mixed, with a short rationale.\n\n"
            f"Input JSON ({path.relative_to(root)}):\n{compact}"
        )
        items.append(
            BenchmarkItem(
                question=truncate(question),
                expected=(
                    "Answer should choose stronger, weaker, or mixed and justify the decision using the "
                    "provided cached financial input."
                ),
                source=f"{path.relative_to(root)}:stockbench",
                choices={"A": "stronger", "B": "weaker", "C": "mixed"},
                metadata={
                    "grading": "task_compliance",
                    "expected_key": "stockbench_cached_financials",
                    "input_files": [str(path.relative_to(root))],
                },
            )
        )
    if not items:
        errors.append("StockBench concrete cached financial records were not found")
    return items, errors[:20]


def extract_exploitbench_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    errors: list[str] = []
    missing_paths = [relative for relative in EXPLOITBENCH_REQUIRED_PATHS if not (root / relative).exists()]
    if missing_paths:
        return [], [
            "ExploitBench upstream challenge environment/configuration was not fully materialized",
            f"missing required path(s): {', '.join(missing_paths)}",
        ]
    items: list[BenchmarkItem] = []
    for relative in EXPLOITBENCH_REQUIRED_CONFIGS:
        path = root / relative
        text = path.read_text(encoding="utf-8", errors="replace")
        target_images = _container_images_from_text(text)
        if not target_images:
            errors.append(f"{relative}: no target container/image reference found")
            continue
        items.append(
            BenchmarkItem(
                question=truncate(
                    "ExploitBench upstream V8 capability-ladder challenge configuration.\n"
                    "Use the upstream ExploitBench runner and deterministic oracle for this target; "
                    "do not answer from documentation or benchmark methodology text.\n\n"
                    f"Config: {relative}\n"
                    f"Target images: {', '.join(target_images[:5])}\n"
                    f"Config excerpt:\n{truncate(text, 3200)}"
                ),
                expected="Use the upstream ExploitBench deterministic capability/oracle outputs for grading.",
                source=relative,
                metadata={
                    "grading": "task_compliance",
                    "expected_key": "upstream_deterministic_oracle",
                    "benchmark_config": relative,
                    "target_image": target_images[0],
                    "target_images": target_images,
                    "grader": "upstream_deterministic_oracle",
                    "oracle": "upstream_capability_oracle",
                    "required_tools": ["exploitbench"],
                    "capability_flags": ["v8_capability_ladder"],
                },
            )
        )
        if len(items) >= limit:
            break
    if not items and not errors:
        errors.append("ExploitBench concrete upstream challenge configs were not found")
    return items[:limit], errors[:20]


def _container_images_from_text(text: str) -> list[str]:
    images: set[str] = set()
    for match in re.finditer(r"(?:ghcr\.io|docker\.io|quay\.io)/[A-Za-z0-9._/@:-]+", text):
        images.add(match.group(0).rstrip("',\")]}"))
    for match in re.finditer(r"\b(?:image|target_image|docker_image)\s*:\s*['\"]?([^'\"\s]+)", text):
        value = match.group(1).strip().rstrip(",")
        if "/" in value and ":" in value:
            images.add(value)
    return sorted(images)


def extract_finmcp_static_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for path in _data_record_files(root):
        if len(items) >= limit:
            break
        try:
            records = _records_from_data_file(path)
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        for index, record in enumerate(records, 1):
            if len(items) >= limit:
                break
            if not isinstance(record, dict):
                continue
            transcript = _finmcp_transcript(record)
            question = _first_text(record, QUESTION_KEYS)
            if not question or not transcript:
                continue
            expected_key, expected_value = _first_value_with_key(record, ANSWER_KEYS)
            expected = stringify_expected(expected_value, preserve_rubric=True) if expected_value is not None else (
                "Answer should use only the supplied historical transcript and data."
            )
            items.append(
                BenchmarkItem(
                    question=truncate(f"User query:\n{question}\n\nStatic transcript:\n{transcript}", 9000),
                    expected=truncate(expected, 3000),
                    source=f"{path.relative_to(root)}:record-{index:06d}",
                    metadata={
                        "grading": "task_compliance" if expected_value is None else "rubric",
                        "expected_key": expected_key or "static_transcript_answer",
                        "source_dataset": "DianJin/FinMCP-Bench",
                        "live_tools_required": False,
                        "required_capabilities": ["chat_answer"],
                        "messages": record.get("messages"),
                        "transcript": transcript,
                    },
                )
            )
    if not items:
        errors.append("FinMCP-Bench records with static transcript/messages were not found")
    return items[:limit], errors[:20]


def _finmcp_transcript(record: dict[str, Any]) -> str:
    for key in ("messages", "transcript", "conversation", "trajectory", "dialogue"):
        value = record.get(key)
        if not value:
            continue
        text = _visible_value_to_text(value)
        if text and ("tool" in text.lower() or len(text) >= 80):
            return truncate(text, 7000)
    return ""


def extract_fintoolbench_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    question_path = root / "data" / "question" / "select_data_real_remove_duplicates.jsonl"
    tool_path = root / "tools" / "tools_all_annotated.jsonl"
    errors: list[str] = []
    if not question_path.is_file():
        errors.append("FinToolBench question set data/question/select_data_real_remove_duplicates.jsonl was not found")
    if not tool_path.is_file():
        errors.append("FinToolBench tool manifest tools/tools_all_annotated.jsonl was not found")
    if errors:
        return [], errors
    tool_manifest = _fintoolbench_tool_manifest(tool_path)
    items: list[BenchmarkItem] = []
    for index, record in enumerate(_records_from_jsonl(question_path), 1):
        if len(items) >= limit:
            break
        item = item_from_record(record, f"{question_path.relative_to(root)}:{index}")
        if item is None:
            continue
        required_tools = _required_tools_from_record(record)
        item.metadata.update(
            {
                "expected_key": item.metadata.get("expected_key") or "fintoolbench_tool_required_query",
                "required_tools": required_tools,
                "live_tools_required": True,
                "local_evaluation_mode": "live_tool_call",
                "tool_manifest": str(tool_path.relative_to(root)),
                "tool_manifest_count": len(tool_manifest),
            }
        )
        items.append(item)
    if not items:
        errors.append("FinToolBench concrete tool-required questions were not found")
    return items, errors[:20]


def _fintoolbench_tool_manifest(path: Path) -> dict[str, dict[str, Any]]:
    manifest: dict[str, dict[str, Any]] = {}
    for record in _records_from_jsonl_unbounded(path):
        if not isinstance(record, dict):
            continue
        name = _first_text(record, ("name", "tool_name", "api_name", "function_name"))
        if name:
            manifest[name] = record
    return manifest


def _records_from_jsonl_unbounded(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(payload, dict):
                records.append(payload)
    return records


def extract_finance_agent_v2_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    data_files = [path for path in _data_record_files(root) if "task.md" not in path.name.lower()]
    task_md_files = sorted(path for path in root.rglob("TASK.md") if path.is_file())
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for path in data_files:
        if len(items) >= limit:
            break
        try:
            records = _records_from_data_file(path)
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        for index, record in enumerate(records, 1):
            if len(items) >= limit:
                break
            item = item_from_record(record, f"{path.relative_to(root)}:record-{index:06d}")
            if item is None:
                continue
            item.metadata.setdefault("required_tools", sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS))
            item.metadata["live_tools_required"] = True
            item.metadata["local_evaluation_mode"] = "live_tool_call"
            items.append(item)
    if not items and task_md_files:
        path = task_md_files[0]
        text = path.read_text(encoding="utf-8", errors="replace")
        items.append(
            BenchmarkItem(
                question=truncate(text, 6000),
                expected="Finance Agent v2 requires official data/platform evidence, not TASK.md-only prompts.",
                source=str(path.relative_to(root)),
                metadata={
                    "grading": "task_compliance",
                    "task_md_only": True,
                    "required_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                    "live_tools_required": True,
                },
            )
        )
    if not items:
        errors.append("Finance Agent v2 official data records or evidence packets were not found")
    return items[:limit], errors[:20]


def extract_swelancer_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    csv_path = _swelancer_task_csv(root)
    if csv_path is None:
        return [], ["SWE-Lancer task CSV was not found"]
    items: list[BenchmarkItem] = []
    with csv_path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        for index, row in enumerate(csv.DictReader(handle), 2):
            if len(items) >= limit:
                break
            issue_id = str(row.get("question_id") or "").strip()
            issue_dir = root / "issues" / issue_id if issue_id else None
            issue_data = _load_json_file(issue_dir / "issue_data.json") if issue_dir else {}
            prompt_parts = [
                str(row.get("title") or "").strip(),
                str(row.get("description") or "").strip(),
                _text_from_value(row.get("prompt")),
                _text_from_value(issue_data.get("problem_statement")),
                _text_from_value(issue_data.get("issue")),
                _text_from_value(issue_data.get("issue_repo_steps")),
            ]
            question = "\n\n".join(part for part in prompt_parts if part).strip()
            if not question:
                continue
            target_context = _swelancer_target_context(row, issue_data, issue_dir)
            metadata = {
                "grading": "task_compliance",
                "expected_key": "repo_patch_tests",
                "instance_id": issue_id,
                "issue_dir": str(issue_dir.relative_to(root)) if issue_dir and issue_dir.exists() else "",
                "catalog_root": str(root),
                "upstream_public_repository": SWELANCER_OFFICIAL_PUBLIC_REPOSITORY,
                "upstream_public_ref": SWELANCER_OFFICIAL_PUBLIC_REF,
            }
            metadata.update(target_context)
            if not target_context.get("target_repo") or not target_context.get("base_commit"):
                metadata["invalid_task_context_reason"] = (
                    "SWE-Lancer public catalog row did not expose target_repo/base_commit metadata"
                )
            items.append(
                BenchmarkItem(
                    question=truncate(question),
                    expected="A repository patch should satisfy the SWE-Lancer issue tests.",
                    source=f"{csv_path.relative_to(root)}:{index}",
                    metadata=metadata,
                )
            )
    if not items:
        return [], ["SWE-Lancer CSV did not contain concrete issue tasks"]
    return items, []


def _swelancer_task_csv(root: Path) -> Path | None:
    for relative in ("all_swelancer_tasks.csv", "swelancer_tasks.csv", "swelancer_tasks_lite.csv"):
        path = root / relative
        if path.is_file():
            return path
    return None


def _swelancer_target_context(
    row: dict[str, Any],
    issue_data: dict[str, Any],
    issue_dir: Path | None,
) -> dict[str, str]:
    containers = [row, issue_data]
    context: dict[str, str] = {}
    target_repo = _first_keyed_text(
        containers,
        (
            "target_repo",
            "target_repository",
            "repo",
            "repo_url",
            "repository",
            "github_repo",
            "git_repo",
        ),
    )
    base_commit = _first_keyed_text(
        containers,
        (
            "base_commit",
            "base_sha",
            "base_revision",
            "base_ref",
            "commit",
            "revision",
        ),
    )
    cwd = _first_keyed_text(containers, ("cwd", "workdir", "working_directory", "workspace"))
    issue_commit_id = ""
    target_repo_source = "record_metadata" if target_repo else ""
    base_commit_source = "record_metadata" if base_commit else ""
    if issue_dir is not None:
        file_target_repo = _first_existing_file_text(
            issue_dir,
            ("target_repo.txt", "target_repository.txt", "repo.txt", "repository.txt"),
        )
        if file_target_repo and not target_repo:
            target_repo = file_target_repo
            target_repo_source = f"{issue_dir.name}/target_repo.txt"
        file_base_commit = _first_existing_file_text(
            issue_dir,
            ("base_commit.txt", "base_sha.txt", "base_revision.txt"),
        )
        if file_base_commit and not base_commit:
            base_commit = file_base_commit
            base_commit_source = f"{issue_dir.name}/base_commit.txt"
        issue_commit_id = _first_existing_file_text(issue_dir, ("commit_id.txt",))
        if issue_commit_id and not base_commit:
            base_commit = issue_commit_id
            base_commit_source = f"{issue_dir.name}/commit_id.txt"
    inferred_target_repo = _swelancer_target_repo_from_cwd(cwd)
    if inferred_target_repo and not target_repo:
        target_repo = inferred_target_repo
        target_repo_source = f"cwd:{cwd}"
    if cwd:
        context["official_workspace_cwd"] = cwd
    if target_repo:
        context["target_repo"] = target_repo
        if target_repo_source:
            context["target_repo_source"] = target_repo_source
    if base_commit:
        context["base_commit"] = base_commit
        if base_commit_source:
            context["base_commit_source"] = base_commit_source
    if issue_commit_id:
        context["issue_commit_id"] = issue_commit_id
    if inferred_target_repo:
        context["metadata_association"] = "official_swelancer_workspace"
    return context


def _swelancer_target_repo_from_cwd(cwd: str) -> str:
    normalized = cwd.strip().replace("\\", "/").rstrip("/").lower()
    return SWELANCER_CWD_TARGET_REPOS.get(normalized, "")


def _first_keyed_text(containers: list[dict[str, Any]], keys: tuple[str, ...]) -> str:
    normalized_keys = {key.lower(): key for key in keys}
    for container in containers:
        lower_to_key = {str(key).lower(): key for key in container}
        for lowered in normalized_keys:
            original = lower_to_key.get(lowered)
            if original is None:
                continue
            value = container.get(original)
            if isinstance(value, (str, int, float)) and str(value).strip():
                return str(value).strip()
    return ""


def _first_existing_file_text(root: Path, filenames: tuple[str, ...]) -> str:
    for filename in filenames:
        path = root / filename
        if path.is_file():
            text = path.read_text(encoding="utf-8", errors="replace").strip()
            if text:
                return text
    return ""


def extract_quantcode_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    path = root / "data" / "benchmark_tasks_multiframe.json"
    if not path.is_file():
        return [], ["QuantCode-Bench task JSON was not found"]
    base_commit = _git_head(root)
    if not base_commit:
        return [], ["QuantCode-Bench repository base commit could not be identified"]
    try:
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception as exc:
        return [], [f"{path.relative_to(root)}: {exc}"]
    items: list[BenchmarkItem] = []
    for index, record in enumerate(walk_records(payload), 1):
        if len(items) >= limit:
            break
        question = _first_text(record, QUESTION_KEYS)
        if not question:
            continue
        instance_id = str(record.get("id") or record.get("source") or f"record-{index:06d}")
        items.append(
            BenchmarkItem(
                question=truncate(question),
                expected="A repository patch should implement the requested QuantCode-Bench strategy/task.",
                source=f"{path.relative_to(root)}:record-{index:06d}",
                metadata={
                    "grading": "task_compliance",
                    "expected_key": "repo_patch_tests",
                    "target_repo": str(root),
                    "base_commit": base_commit,
                    "instance_id": instance_id,
                },
            )
        )
    if not items:
        return [], ["QuantCode-Bench JSON did not contain concrete task records"]
    return items, []


def _items_from_archive_member(text: str, member: str, source_prefix: str, limit: int) -> list[BenchmarkItem]:
    suffix = Path(member).suffix.lower()
    if suffix == ".json":
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return []
        return _items_from_records(walk_records(payload), source_prefix, limit)
    if suffix == ".jsonl":
        records = []
        for line in text.splitlines():
            if len(records) >= min(limit, MAX_RECORDS_PER_FILE):
                break
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(payload, dict):
                records.append(payload)
        return _items_from_records(records, source_prefix, limit)
    if suffix == ".csv":
        reader = csv.DictReader(text.splitlines())
        return _items_from_records(list(reader)[: min(limit, MAX_RECORDS_PER_FILE)], source_prefix, limit)
    return []


def _items_from_records(records: list[dict[str, Any]], source_prefix: str, limit: int) -> list[BenchmarkItem]:
    items: list[BenchmarkItem] = []
    for index, record in enumerate(records):
        if len(items) >= limit or index >= MAX_RECORDS_PER_FILE:
            break
        item = item_from_record(record, f"{source_prefix}:{index + 1}")
        if item is not None:
            items.append(item)
    return items


def extract_investorbench_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    data_dir = root / "data"
    if not data_dir.is_dir():
        return [], ["InvestorBench data directory was not found"]
    items: list[BenchmarkItem] = []
    errors: list[str] = []
    for path in sorted(data_dir.glob("*.json")):
        if len(items) >= limit:
            break
        try:
            series = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except Exception as exc:
            errors.append(f"{path.relative_to(root)}: {exc}")
            continue
        if not isinstance(series, dict):
            continue
        dates = sorted(series)
        for index, current_date in enumerate(dates[:-1]):
            if len(items) >= limit:
                break
            current = series.get(current_date)
            future = series.get(dates[index + 1])
            if not isinstance(current, dict) or not isinstance(future, dict):
                continue
            current_price = _as_float(current.get("prices"))
            future_price = _as_float(future.get("prices"))
            if current_price is None or future_price is None:
                continue
            symbol = path.stem.upper()
            movement = (future_price - current_price) / current_price
            if movement > 0.0025:
                expected = "buy"
            elif movement < -0.0025:
                expected = "sell"
            else:
                expected = "hold"
            news = current.get("news") if isinstance(current.get("news"), list) else []
            news_text = "\n".join(str(item) for item in news[:5])
            question = (
                f"InvestorBench trading decision for {symbol} on {current_date}.\n"
                f"Current price: {current_price:.4f}. Recent news:\n{news_text}\n\n"
                "Choose one action for the next trading step: buy, sell, or hold. "
                "Return only the action and a short rationale."
            )
            items.append(
                BenchmarkItem(
                    question=truncate(question),
                    expected=expected,
                    source=f"{path.relative_to(root)}:{current_date}",
                    choices={"A": "buy", "B": "sell", "C": "hold"},
                    metadata={"grading": "exact", "expected_key": "next_price_direction"},
                )
            )
    return items, errors


def extract_hle_items(root: Path, limit: int) -> tuple[list[BenchmarkItem], list[str]]:
    readme = root / "README.md"
    runner = root / "hle_eval" / "run_model_predictions.py"
    if not readme.is_file():
        return [], ["HLE README.md was not found"]
    readme_text = readme.read_text(encoding="utf-8", errors="replace")
    runner_text = runner.read_text(encoding="utf-8", errors="replace") if runner.is_file() else ""
    match = re.search(r"SYSTEM_PROMPT\s*=\s*([\"'])(.*?)\1", runner_text, flags=re.DOTALL)
    system_prompt = match.group(2) if match else "Provide an explanation, answer, and confidence."
    question = (
        "Humanity's Last Exam evaluation-format task from the public benchmark repository.\n"
        f"{system_prompt}\n\n"
        "Summarize the required answer format and then answer a closed-ended academic benchmark item "
        "using that format if a question is supplied."
    )
    item = BenchmarkItem(
        question=truncate(question),
        expected=(
            "Response should follow the HLE public evaluation format with Explanation, Answer, "
            "and Confidence fields for closed-ended academic questions."
        ),
        source="hle_eval/run_model_predictions.py",
        metadata={"grading": "rubric", "expected_key": "public_evaluation_format"},
    )
    return [item][:limit], []


def _as_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_json_file(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _git_head(root: Path) -> str:
    completed = subprocess.run(
        ["git", "-C", str(root), "rev-parse", "HEAD"],
        text=True,
        capture_output=True,
        timeout=30,
        check=False,
    )
    return completed.stdout.strip() if completed.returncode == 0 else ""


def _is_financemath_benchmark() -> bool:
    return normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ") == "financemath"


def _numeric_answer_from_python_solution(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    for pattern in (
        r"return\s+(-?\d+(?:\.\d+)?(?:e[+-]?\d+)?)",
        r"answer\s*=\s*(-?\d+(?:\.\d+)?(?:e[+-]?\d+)?)",
        r"(-?\d+(?:\.\d+)?(?:e[+-]?\d+)?)\s*$",
    ):
        match = re.search(pattern, value.strip(), flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def extract_items_from_file(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        return _items_from_json_payload(payload, path.relative_to(root), limit)
    if suffix == ".jsonl":
        return _items_from_jsonl(path, root, limit)
    if suffix == ".csv":
        return _items_from_csv(path, root, limit)
    if suffix == ".parquet":
        return _items_from_parquet(path, root, limit)
    if suffix == ".py":
        return _items_from_python(path, root, limit)
    if suffix in {".md", ".txt", ".yaml", ".yml"}:
        return _items_from_text_file(path, root, limit)
    return []


def _items_from_python(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    source = path.read_text(encoding="utf-8", errors="replace")
    tree = ast.parse(source)
    constants = _module_constants(tree)
    items: list[BenchmarkItem] = []
    for node in ast.walk(tree):
        if len(items) >= limit:
            break
        if not isinstance(node, ast.Return) or node.value is None:
            continue
        try:
            record = _literal_from_ast(node.value, constants)
        except ValueError:
            continue
        for candidate in _records_from_python_literal(record):
            if len(items) >= limit:
                break
            item = item_from_record(candidate, f"{path.relative_to(root)}:{node.lineno}")
            if item is not None:
                items.append(item)
    return items


def _module_constants(tree: ast.Module) -> dict[str, Any]:
    constants: dict[str, Any] = {}
    for node in tree.body:
        if not isinstance(node, ast.Assign) or len(node.targets) != 1:
            continue
        target = node.targets[0]
        if not isinstance(target, ast.Name):
            continue
        try:
            constants[target.id] = _literal_from_ast(node.value, constants)
        except ValueError:
            continue
    return constants


def _literal_from_ast(node: ast.AST, constants: dict[str, Any]) -> Any:
    if isinstance(node, ast.Constant):
        return node.value
    if isinstance(node, ast.Name):
        if node.id in constants:
            return constants[node.id]
        raise ValueError(node.id)
    if isinstance(node, (ast.List, ast.Tuple)):
        return [_literal_from_ast(item, constants) for item in node.elts]
    if isinstance(node, ast.Dict):
        return {
            _literal_from_ast(key, constants): _literal_from_ast(value, constants)
            for key, value in zip(node.keys, node.values)
            if key is not None
        }
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
        value = _literal_from_ast(node.operand, constants)
        if isinstance(value, (int, float)):
            return -value
    raise ValueError(type(node).__name__)


def _records_from_python_literal(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        return [value]
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def _items_from_text_file(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    text = path.read_text(encoding="utf-8", errors="replace").strip()
    if len(text) < 80:
        return []
    record = {
        "prompt": truncate(text),
        "rubric": "Candidate answer should satisfy the benchmark task or prompt described in this source file.",
    }
    item = item_from_record(record, str(path.relative_to(root)))
    return [item] if item is not None and limit > 0 else []


def _items_from_jsonl(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    items: list[BenchmarkItem] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for index, line in enumerate(handle):
            if len(items) >= limit or index >= MAX_RECORDS_PER_FILE:
                break
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            item = item_from_record(record, f"{path.relative_to(root)}:{index + 1}")
            if item is not None:
                items.append(item)
    return items


def _items_from_csv(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    items: list[BenchmarkItem] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader):
            if len(items) >= limit or index >= MAX_RECORDS_PER_FILE:
                break
            item = item_from_record(row, f"{path.relative_to(root)}:{index + 2}")
            if item is not None:
                items.append(item)
    return items


def _items_from_parquet(path: Path, root: Path, limit: int) -> list[BenchmarkItem]:
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise RuntimeError("pyarrow is required to read parquet benchmark data") from exc
    table = pq.read_table(path)
    items: list[BenchmarkItem] = []
    for index, row in enumerate(table.slice(0, min(MAX_RECORDS_PER_FILE, table.num_rows)).to_pylist()):
        if len(items) >= limit:
            break
        item = item_from_record(row, f"{path.relative_to(root)}:{index + 1}")
        if item is not None:
            items.append(item)
    return items


def _items_from_json_payload(payload: Any, source: Path, limit: int) -> list[BenchmarkItem]:
    items: list[BenchmarkItem] = []
    for index, record in enumerate(walk_records(payload), 1):
        if len(items) >= limit:
            break
        item = item_from_record(record, f"{source}:record-{index:06d}")
        if item is not None:
            items.append(item)
    return items


def walk_records(payload: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    def visit(value: Any) -> None:
        if len(records) >= MAX_RECORDS_PER_FILE:
            return
        if isinstance(value, dict):
            records.append(value)
            for nested in value.values():
                if isinstance(nested, (dict, list)):
                    visit(nested)
        elif isinstance(value, list):
            for nested in value[:MAX_RECORDS_PER_FILE]:
                visit(nested)

    visit(payload)
    return records


def item_from_record(record: Any, source: str) -> BenchmarkItem | None:
    if not isinstance(record, dict):
        return None
    question = _first_text(record, QUESTION_KEYS)
    expected_key, expected_value = _first_value_with_key(record, ANSWER_KEYS)
    if not question:
        return None
    grading = "exact"
    if expected_value is None:
        if not _is_task_like_record(record, source, question):
            return None
        expected = "Candidate answer should satisfy the task requirements from the benchmark prompt."
        grading = "task_compliance"
    else:
        grading = "rubric" if expected_key in RUBRIC_KEYS else "exact"
        if _is_financemath_benchmark() and expected_key == "python_solution":
            solution_answer = _numeric_answer_from_python_solution(expected_value)
            if solution_answer:
                expected_value = solution_answer
        canonical_answer = (
            extract_canonical_answer_from_rubric(str(expected_value))
            if expected_key in RUBRIC_KEYS and isinstance(expected_value, str)
            else ""
        )
        if canonical_answer:
            grading = "exact"
            expected = canonical_answer
        else:
            expected = stringify_expected(expected_value, preserve_rubric=grading == "rubric")
        if not expected and _is_task_like_record(record, source, question):
            grading = "task_compliance"
            expected = "Candidate answer should satisfy the task requirements from the benchmark prompt."
        if grading == "exact" and _looks_like_reference_file(expected):
            grading = "task_compliance"
            expected = "Candidate answer should satisfy the task requirements using the referenced benchmark files."
    if _is_financemath_benchmark():
        if _numeric_value(expected) is None:
            return None
        grading = "numeric"
    if not expected:
        return None
    choices = _choices_from_record(record)
    metadata = _record_metadata(record, source, grading, expected_key, expected)
    return BenchmarkItem(
        question=truncate(question),
        expected=truncate(expected, 3000 if grading == "rubric" else 1000),
        choices=choices,
        source=source,
        metadata=metadata,
    )


def _record_metadata(
    record: dict[str, Any],
    source: str,
    grading: str,
    expected_key: str,
    expected: str,
) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "keys": sorted(str(key) for key in record.keys())[:30],
        "grading": grading,
        "expected_key": expected_key or "",
    }
    lower_to_key = {str(key).lower(): key for key in record}
    for key in (
        "target_repo",
        "repo",
        "repo_url",
        "git_repo",
        "repository",
        "base_commit",
        "commit",
        "base_sha",
        "revision",
        "instance_id",
        "test_patch",
        "patch",
        "gold_patch",
    ):
        original = lower_to_key.get(key)
        if original is None:
            continue
        value = record.get(original)
        if isinstance(value, (str, int, float, bool)):
            metadata[key] = str(value)
    if expected_key in {"patch", "gold_patch"} and expected:
        metadata["reference_patch"] = expected
    for key in (
        "input_assets",
        "input_files",
        "asset_paths",
        "attachments",
        "files",
        "reference_files",
        "documents",
        "dataset_files",
    ):
        original = lower_to_key.get(key)
        if original is not None:
            metadata[key] = record.get(original)
    if source.startswith("huggingface:"):
        metadata["dataset_source"] = source
    visible_context = _model_visible_record_context(record, expected_key)
    if visible_context:
        metadata["visible_context"] = visible_context
    return metadata


def _model_visible_record_context(record: dict[str, Any], expected_key: str) -> str:
    sections: list[str] = []
    consumed: set[str] = set()
    lowered_to_original = {str(key).lower(): key for key in record}
    hidden_keys = {key.lower() for key in MODEL_VISIBLE_CONTEXT_SKIP_KEYS}
    if expected_key:
        hidden_keys.add(expected_key.lower())

    for key in MODEL_VISIBLE_CONTEXT_PRIORITY_KEYS:
        original = lowered_to_original.get(key)
        if original is None:
            continue
        consumed.add(str(original).lower())
        text = _format_visible_context_value(str(original), record.get(original))
        if text:
            sections.append(text)

    for original, value in record.items():
        lowered = str(original).lower()
        if lowered in consumed or lowered in hidden_keys:
            continue
        if not _is_useful_visible_context_value(value):
            continue
        text = _format_visible_context_value(str(original), value)
        if text:
            sections.append(text)

    if not sections:
        return ""
    return truncate("\n\n".join(sections), MAX_VISIBLE_CONTEXT_CHARS)


def _is_useful_visible_context_value(value: Any) -> bool:
    if isinstance(value, str):
        stripped = value.strip()
        return len(stripped) >= 20 and not _looks_like_reference_file(stripped)
    if isinstance(value, (int, float, bool)):
        return False
    if isinstance(value, list):
        return any(_is_useful_visible_context_value(item) for item in value)
    if isinstance(value, dict):
        return any(_is_useful_visible_context_value(item) for item in value.values())
    return False


def _format_visible_context_value(key: str, value: Any) -> str:
    text = _visible_value_to_text(value)
    if not text:
        return ""
    return f"{key}:\n{text}"


def _visible_value_to_text(value: Any) -> str:
    if isinstance(value, str):
        return truncate(value.strip(), MAX_FIELD_CHARS)
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        if not value:
            return ""
        if all(not isinstance(item, (dict, list)) for item in value):
            return "\n".join(f"- {truncate(str(item), 1000)}" for item in value)
        return truncate(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True), MAX_FIELD_CHARS)
    if isinstance(value, dict):
        return truncate(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True), MAX_FIELD_CHARS)
    return ""


def _is_task_like_record(record: dict[str, Any], source: str, question: str) -> bool:
    if len(question.strip()) < 40:
        return False
    source_text = source.lower()
    key_text = " ".join(str(key).lower() for key in record.keys())
    if any(word in source_text for word in ("task", "question", "problem", "benchmark", "eval", "test", "validation")):
        return True
    return any(word in key_text for word in ("task", "question", "problem", "prompt", "requirement"))


def _first_text(record: dict[str, Any], keys: tuple[str, ...]) -> str:
    _, value = _first_value_with_key(record, keys)
    text = _text_from_value(value)
    if text:
        return text
    return ""


def _text_from_value(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, dict):
                content = item.get("content")
                if isinstance(content, str):
                    parts.append(content.strip())
            elif isinstance(item, str):
                parts.append(item.strip())
        return "\n".join(part for part in parts if part).strip()
    if isinstance(value, dict):
        for key in ("content", "text", "prompt", "instruction", "question"):
            if isinstance(value.get(key), str):
                return value[key].strip()
    return ""


def _first_value(record: dict[str, Any], keys: tuple[str, ...]) -> Any:
    _, value = _first_value_with_key(record, keys)
    return value


def _first_value_with_key(record: dict[str, Any], keys: tuple[str, ...]) -> tuple[str, Any]:
    lowered = {str(key).lower(): key for key in record}
    for candidate in keys:
        if candidate in lowered:
            return candidate, record[lowered[candidate]]
    for key, value in record.items():
        key_text = str(key).lower()
        if any(_key_matches(candidate, key_text) for candidate in keys):
            return key_text, value
    return "", None


def _key_matches(candidate: str, key_text: str) -> bool:
    if len(candidate) <= 5:
        return False
    return key_text.endswith(candidate) or key_text.startswith(candidate + "_")


def _choices_from_record(record: dict[str, Any]) -> dict[str, str]:
    value = _first_value(record, CHOICE_KEYS)
    choices: dict[str, str] = {}
    if isinstance(value, dict):
        for key, item in value.items():
            label = normalize_choice_label(str(key), len(choices))
            choices[label] = truncate(str(item), 1000)
    elif isinstance(value, list):
        for index, item in enumerate(value):
            label = normalize_choice_label("", index)
            choices[label] = truncate(str(item), 1000)
    else:
        for label in ("A", "B", "C", "D", "E"):
            if label in record:
                choices[label] = truncate(str(record[label]), 1000)
            elif label.lower() in record:
                choices[label] = truncate(str(record[label.lower()]), 1000)
    return choices


def normalize_choice_label(value: str, index: int) -> str:
    cleaned = value.strip().upper()
    if re.fullmatch(r"[A-Z]", cleaned):
        return cleaned
    if cleaned.isdigit():
        numeric = int(cleaned)
        if 0 <= numeric <= 25:
            return chr(ord("A") + numeric)
        if 1 <= numeric <= 26:
            return chr(ord("A") + numeric - 1)
    return chr(ord("A") + index)


def stringify_expected(value: Any, preserve_rubric: bool = False) -> str:
    if isinstance(value, str):
        value = value.strip()
        return value if preserve_rubric else extract_expected_from_rubric(value)
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        if not value:
            return ""
        if preserve_rubric:
            return stringify_rubric(value)
        return str(value[0]).strip()
    if isinstance(value, dict):
        if preserve_rubric:
            return stringify_rubric(value)
        for key in ("answer", "text", "label", "value"):
            if key in value:
                return stringify_expected(value[key])
    return ""


def stringify_rubric(value: Any) -> str:
    if isinstance(value, str):
        text = value.strip()
        parsed = parse_json_object(text)
        if parsed is not None:
            return stringify_rubric(parsed)
        return text
    if isinstance(value, list):
        parts = [stringify_rubric(item) for item in value]
        return "\n".join(f"- {part}" for part in parts if part)
    if isinstance(value, dict):
        if "criteria" in value:
            return str(value["criteria"]).strip()
        if "requirements" in value:
            return str(value["requirements"]).strip()
        return "; ".join(f"{key}: {item}" for key, item in value.items() if item not in (None, ""))
    return str(value).strip()


def _looks_like_reference_file(value: str) -> bool:
    text = value.strip().lower()
    if not text:
        return False
    if re.search(r"\.(csv|json|jsonl|xlsx|xls|pdf|txt|png|jpg|jpeg|parquet|zip)(?:$|[\s#?])", text):
        return True
    return text.startswith(("reference_files/", "references/", "attachments/", "files/"))


def extract_expected_from_rubric(value: str) -> str:
    canonical = extract_canonical_answer_from_rubric(value)
    if canonical:
        return canonical
    for pattern in (
        r"(?:expected answer|answer)\s+is\s+([^.\n]+)",
        r"(?:expected|gold|target)\s*[:\-]\s*([^.\n]+)",
    ):
        match = re.search(pattern, value, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return value


def extract_canonical_answer_from_rubric(value: str) -> str:
    text = " ".join(value.strip().split())
    for pattern in (
        r"\bThe answer is\s+(.+?)\s+Score\s+1(?:\.0)?\b",
        r"\banswer is\s+(.+?)(?:\.|;)?\s+Score\s+1(?:\.0)?\b",
        r"\bcorrect answer is\s+(.+?)(?:\.|;)?\s+Score\s+1(?:\.0)?\b",
        r"^(.{1,120}?)\s+Score\s+1(?:\.0)?\b",
    ):
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip(" .;:,'\"")
    return ""


def run_model_evaluations(
    benchmark: str,
    items: list[BenchmarkItem],
    adapter: BenchmarkAdapter | None = None,
) -> list[dict[str, Any]]:
    if not _is_remote_provider():
        return []
    adapter = adapter or select_adapter(_required_capabilities(benchmark))
    evaluations: list[dict[str, Any]] = []
    for item in items:
        evaluations.append(run_model_on_item(benchmark, item, adapter))
    return evaluations


def run_model_on_item(
    benchmark: str,
    item: BenchmarkItem,
    adapter: BenchmarkAdapter | None = None,
) -> dict[str, Any]:
    adapter = adapter or select_adapter(_required_capabilities(benchmark))
    return adapter.evaluate_item(benchmark, item)


def _preflight_failure_from_contract(
    required_capabilities: set[str],
    capability_contract: dict[str, Any],
) -> tuple[str, str]:
    for capability in sorted(required_capabilities):
        support = capability_contract.get(capability)
        if not isinstance(support, dict) or support.get("supported") is not False:
            continue
        reason = str(support.get("reason") or f"{capability} capability preflight failed")
        if support.get("benchmark_required_tools_available") is False or support.get("missing_tools"):
            return "failed_missing_required_tool", reason
        if capability == "file_artifact":
            return "failed_missing_assets", reason
        if capability == "repo_patch":
            if "missing target repo/base_commit metadata" in reason:
                return "failed_invalid_task_context", reason
            if "official patch/test grader" in reason:
                return "skipped_unsupported_capability", reason
            return "failed_harness_setup", reason
        return "skipped_unsupported_capability", reason
    return "", ""


def preflight_failed_evaluations(
    items: list[BenchmarkItem],
    adapter: BenchmarkAdapter,
    status: str,
    reason: str,
    capability_contract: dict[str, Any],
) -> list[dict[str, Any]]:
    evaluations: list[dict[str, Any]] = []
    missing_tools = _contract_missing_tools(capability_contract)
    exposed_tools = _contract_exposed_tools(capability_contract)
    for item in items:
        item_dir = _item_output_dir(item)
        _write_item_json(item_dir, "item.json", _item_to_dict(item))
        setup_payload = {
            "status": status,
            "error": reason,
            "capability_contract": capability_contract,
            "missing_tools": missing_tools,
            "exposed_tools": exposed_tools,
        }
        _write_item_json(item_dir, "setup_error.json", setup_payload)
        payload = evaluation_payload(
            item,
            "",
            0.0,
            reason,
            **_evaluation_preflight_fields(item, adapter, [], missing_tools, capability_contract),
            status=status,
            adapter=adapter.name,
            capabilities_verified=False,
            setup_details=setup_payload,
        )
        _write_item_json(item_dir, "item_result.json", payload)
        evaluations.append(payload)
    return evaluations


def validate_items_preflight(
    items: list[BenchmarkItem],
    adapter: BenchmarkAdapter,
    required_capabilities: set[str],
    capability_contract: dict[str, Any],
) -> tuple[list[BenchmarkItem], list[dict[str, Any]]]:
    valid: list[BenchmarkItem] = []
    invalid: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_dirs: set[Path] = set()
    for item in items:
        status, reason, details = _item_preflight_failure(
            item,
            required_capabilities,
            seen_ids,
            seen_dirs,
        )
        if not status:
            valid.append(item)
            continue
        item_dir = _item_output_dir(item)
        _write_item_json(item_dir, "item.json", _item_to_dict(item))
        setup_payload = {
            "status": status,
            "error": reason,
            "details": details,
            "capability_contract": capability_contract,
        }
        _write_item_json(item_dir, "setup_error.json", setup_payload)
        missing_tools = _flatten_string_values(details.get("missing_tools")) or _flatten_string_values(details.get("required_tools"))
        payload = evaluation_payload(
            item,
            "",
            0.0,
            reason,
            **_evaluation_preflight_fields(item, adapter, [], missing_tools, capability_contract),
            status=status,
            adapter=adapter.name,
            capabilities_verified=False,
            setup_details=setup_payload,
        )
        _write_item_json(item_dir, "item_result.json", payload)
        invalid.append(payload)
    return valid, invalid


def _contract_missing_tools(capability_contract: dict[str, Any]) -> list[str]:
    tools: list[str] = []
    for support in capability_contract.values():
        if not isinstance(support, dict):
            continue
        tools.extend(_flatten_string_values(support.get("missing_tools")))
    return sorted({tool for tool in tools if tool})


def _contract_missing_env(capability_contract: dict[str, Any]) -> list[str]:
    values: list[str] = []
    for support in capability_contract.values():
        if not isinstance(support, dict):
            continue
        values.extend(_flatten_string_values(support.get("missing_env")))
        values.extend(_flatten_string_values(support.get("missing_environment")))
    return sorted({value for value in values if value})


def _contract_required_tools(capability_contract: dict[str, Any]) -> list[str]:
    tools: list[str] = []
    for support in capability_contract.values():
        if not isinstance(support, dict):
            continue
        tools.extend(_flatten_string_values(support.get("required_benchmark_tools")))
    return sorted({tool for tool in tools if tool})


def _contract_exposed_tools(capability_contract: dict[str, Any]) -> list[str]:
    tools: list[str] = []
    for support in capability_contract.values():
        if not isinstance(support, dict):
            continue
        tools.extend(_flatten_string_values(support.get("exposed_tools")))
    return sorted({tool for tool in tools if tool})


def _item_preflight_failure(
    item: BenchmarkItem,
    required_capabilities: set[str],
    seen_ids: set[str],
    seen_dirs: set[Path],
) -> tuple[str, str, dict[str, Any]]:
    item_id = _item_id(item)
    item_dir = _item_output_dir(item)
    details: dict[str, Any] = {"item_id": item_id, "item_output_dir": str(item_dir)}
    if item_id in seen_ids:
        return "failed_dataset_extraction", f"duplicate benchmark item id: {item_id}", details
    if item_dir in seen_dirs:
        return "failed_dataset_extraction", f"duplicate item output directory: {item_dir.name}", details
    seen_ids.add(item_id)
    seen_dirs.add(item_dir)

    if not item.question.strip():
        return "failed_dataset_extraction", "extracted record has an empty question", details
    if not item.expected.strip():
        return "failed_dataset_extraction", "extracted record has no expected answer or rubric", details
    if _looks_like_incomplete_prompt_template(item):
        return (
            "failed_dataset_extraction",
            "extracted prompt appears to be a template without the referenced task input",
            details,
        )

    benchmark = normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ")
    if benchmark == "finance agent v2":
        finance_tools = _finance_agent_v2_available_tool_schemas(Path.cwd())
        finance_error = validate_finance_agent_v2_item(item, Path.cwd(), finance_tools)
        if finance_error is not None:
            return finance_error
    if benchmark == "swe lancer":
        swelancer_error = validate_swelancer_item(item, Path.cwd())
        if swelancer_error is not None:
            return swelancer_error
    if benchmark == "finmcp bench":
        finmcp_error = validate_finmcp_static_item(item)
        if finmcp_error is not None:
            return finmcp_error
    if benchmark == "exploitbench":
        exploit_error = validate_exploitbench_item(item)
        if exploit_error is not None:
            return exploit_error

    missing_refs = _missing_referenced_paths(item, Path.cwd())
    if missing_refs:
        details["missing_references"] = missing_refs
        status = "failed_missing_assets" if "file_artifact" in required_capabilities else "failed_dataset_extraction"
        return status, "extracted record references missing input files", details

    if "file_artifact" in required_capabilities:
        asset_errors = _artifact_asset_validation_errors(item, Path.cwd())
        if asset_errors:
            details["asset_errors"] = asset_errors
            details["blocker_type"] = str(asset_errors[0].get("blocker_type") or "missing_asset")
            reason = _asset_validation_error_message(asset_errors)
            return "failed_missing_assets", reason, details

    if _would_exceed_context_budget(item):
        details["estimated_prompt_tokens"] = _estimated_prompt_tokens(item)
        details["context_limit_tokens"] = _model_context_limit()
        return "failed_token_budget", "estimated task prompt exceeds model context budget with reserve", details

    if item.metadata.get("grading") in {"exact", "numeric"} and score_answer(item.expected, item.expected, item.choices) < 1.0:
        return "failed_grader", "exact-answer grader canary failed on the gold answer", details

    return "", "", {}


def validate_exploitbench_item(item: BenchmarkItem) -> tuple[str, str, dict[str, Any]] | None:
    details = {"blocker_type": "missing_reference_dataset", "item_id": _item_id(item), "source": item.source}
    source_lower = item.source.lower()
    if item.source.endswith((".yaml", ".yml", ".json", ".toml")) is False:
        return "failed_invalid_task_context", "ExploitBench scored item must come from a real upstream config", details
    if any(pattern in source_lower for pattern in EXPLOITBENCH_EXCLUDE_PATTERNS):
        return "failed_invalid_task_context", "ExploitBench docs/spec files are not valid scored items", details
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    if not (metadata.get("target_image") or metadata.get("environment")):
        return "failed_invalid_task_context", "ExploitBench item has no target image/environment", details
    if not (metadata.get("grader") or metadata.get("oracle")):
        return "failed_invalid_task_context", "ExploitBench item has no upstream grader/oracle", details
    upstream_error = _exploitbench_upstream_error(item)
    if upstream_error is not None:
        return upstream_error
    return None


def _exploitbench_upstream_ready(item: BenchmarkItem) -> bool:
    return _exploitbench_upstream_error(item) is None


def _exploitbench_upstream_error(item: BenchmarkItem) -> tuple[str, str, dict[str, Any]] | None:
    if os.environ.get("AGENT_BENCH_EXPLOITBENCH_UPSTREAM_READY", "").strip().lower() in {"1", "true", "yes"}:
        return None
    root = Path.cwd()
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    details = {
        "blocker_type": "missing_reference_dataset",
        "item_id": _item_id(item),
        "target_image": metadata.get("target_image"),
    }
    missing_configs = [relative for relative in EXPLOITBENCH_REQUIRED_CONFIGS if not (root / relative).is_file()]
    benchmark_config = metadata.get("benchmark_config")
    if isinstance(benchmark_config, str) and benchmark_config and not (root / benchmark_config).is_file():
        missing_configs.append(benchmark_config)
    if missing_configs:
        details["missing_configs"] = sorted(set(missing_configs))
        return (
            "failed_dataset_extraction",
            "ExploitBench upstream challenge configuration was not materialized",
            details,
        )
    if not (root / "benchmarks" / "bench-v8").exists():
        details["missing_paths"] = ["benchmarks/bench-v8"]
        return (
            "failed_dataset_extraction",
            "ExploitBench upstream deterministic oracle assets were not materialized",
            details,
        )
    runner = shutil.which("exploitbench")
    if not runner:
        return (
            "failed_missing_required_tool",
            "ExploitBench upstream runner/oracle backend is not installed or not on PATH",
            {
                **details,
                "blocker_type": "missing_required_tool_backend",
                "required_tools": ["exploitbench"],
                "missing_tools": ["exploitbench"],
                "exposed_tools": [],
            },
        )
    completed = subprocess.run(
        [runner, "--help"],
        text=True,
        capture_output=True,
        cwd=root,
        timeout=30,
        check=False,
    )
    if completed.returncode != 0:
        return (
            "failed_missing_required_tool",
            "ExploitBench upstream runner/oracle backend failed its readiness check",
            {
                **details,
                "blocker_type": "missing_required_tool_backend",
                "required_tools": ["exploitbench"],
                "missing_tools": ["exploitbench"],
                "exposed_tools": [],
                "exit_code": completed.returncode,
                "stderr": completed.stderr[-1000:],
            },
        )
    return None


def validate_finmcp_static_item(item: BenchmarkItem) -> tuple[str, str, dict[str, Any]] | None:
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    details = {"blocker_type": "invalid_task_context", "item_id": _item_id(item)}
    if metadata.get("source_dataset") != "DianJin/FinMCP-Bench":
        return "failed_invalid_task_context", "FinMCP-Bench static item missing source dataset provenance", details
    if metadata.get("live_tools_required") is not False:
        return "failed_invalid_task_context", "FinMCP-Bench static item must not require live MCP tools", details
    capabilities = metadata.get("required_capabilities", [])
    if isinstance(capabilities, list) and "tool_call" in capabilities:
        return "failed_invalid_task_context", "FinMCP-Bench static item still advertises tool_call", details
    if not item.question.strip():
        return "failed_invalid_task_context", "FinMCP-Bench static item has no user query", details
    if not (metadata.get("transcript") or metadata.get("messages")):
        return "failed_invalid_task_context", "FinMCP-Bench static item lacks transcript/messages", details
    return None


def validate_finance_agent_v2_item(
    item: BenchmarkItem,
    workspace: Path,
    tools: list[dict[str, Any]],
) -> tuple[str, str, dict[str, Any]] | None:
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    if (
        metadata.get("live_tools_required") is False
        and not metadata.get("task_md_only")
        and _static_conversion_live_tools_disabled()
    ):
        return None
    missing_env = [key for key in FINANCE_AGENT_V2_REQUIRED_ENV if not os.environ.get(key)]
    documents = metadata.get("required_documents")
    has_documents = False
    if isinstance(documents, list) and documents:
        has_documents = all(
            isinstance(doc, dict)
            and isinstance(doc.get("path"), str)
            and (workspace / doc["path"]).is_file()
            for doc in documents
        )
    exposed_tools = set(_tool_schema_names(tools))
    has_retrieval_tools = FINANCE_AGENT_V2_REQUIRED_TOOLS.issubset(exposed_tools)
    missing_tools = sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS - exposed_tools)
    backend = _finance_agent_v2_backend_status(workspace)
    if metadata.get("task_md_only") and not has_documents:
        return (
            "failed_missing_required_tool",
            "Finance Agent v2 required tools are not exposed: " + ", ".join(missing_tools),
            {
                "blocker_type": "missing_required_tool_backend",
                "item_id": _item_id(item),
                "required_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                "missing_tools": missing_tools,
                "exposed_tools": sorted(exposed_tools),
            },
        )
    if has_retrieval_tools and not backend["ready"]:
        return (
            "failed_missing_required_tool",
            _finance_agent_v2_backend_unavailable_reason(backend, missing_tools, missing_env),
            {
                "blocker_type": "missing_required_tool_backend",
                "missing_environment": missing_env,
                "required_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                "missing_tools": missing_tools,
                "exposed_tools": sorted(exposed_tools),
                "backend_canary": backend,
            },
        )
    if missing_env and not has_documents and not backend["ready"]:
        return (
            "failed_missing_required_tool",
            _finance_agent_v2_backend_unavailable_reason(backend, sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS), missing_env),
            {
                "blocker_type": "missing_required_tool_backend",
                "missing_environment": missing_env,
                "required_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                "missing_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                "exposed_tools": sorted(exposed_tools),
                "backend_canary": backend,
            },
        )
    if not has_documents and not has_retrieval_tools:
        return (
            "failed_missing_required_tool",
            _finance_agent_v2_backend_unavailable_reason(backend, missing_tools, missing_env),
            {
                "blocker_type": "missing_required_tool_backend",
                "item_id": _item_id(item),
                "required_tools": sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS),
                "missing_tools": missing_tools,
                "exposed_tools": sorted(exposed_tools),
                "backend_canary": backend,
            },
        )
    return None


def validate_swelancer_item(item: BenchmarkItem, catalog_root: Path) -> tuple[str, str, dict[str, Any]] | None:
    target_repo = _repo_patch_target_repo(item)
    base_commit = _repo_patch_base_commit(item)
    details = {
        "blocker_type": "invalid_task_context",
        "item_id": _item_id(item),
        "target_repo": target_repo,
        "base_commit": base_commit,
    }
    if not target_repo or not base_commit:
        return (
            "failed_invalid_task_context",
            "SWE-Lancer task metadata does not identify a target repository and base commit",
            details,
        )
    if _looks_like_swelancer_catalog_path(target_repo):
        return (
            "failed_invalid_task_context",
            "SWE-Lancer catalog checkout is not the target repository",
            {**details, "catalog_path": str(catalog_root.resolve()), "resolved_target_repo": target_repo},
        )
    catalog = catalog_root.resolve()
    target_path = Path(target_repo).expanduser()
    if target_path.exists():
        try:
            resolved = target_path.resolve()
        except OSError:
            resolved = target_path
        if resolved == catalog or catalog in resolved.parents:
            return (
                "failed_invalid_task_context",
                "SWE-Lancer catalog checkout is not the target repository",
                {**details, "catalog_path": str(catalog), "resolved_target_repo": str(resolved)},
            )
    return None


def _looks_like_swelancer_catalog_path(value: str) -> bool:
    normalized = value.strip().replace("\\", "/").rstrip("/").lower()
    return normalized in {
        "frontier-evals/project/swelancer",
        "project/swelancer",
        "/workspace/repo/project/swelancer",
    } or normalized.endswith("/frontier-evals/project/swelancer")


def _artifact_asset_validation_errors(item: BenchmarkItem, root: Path) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    for path in _artifact_asset_files(item, root):
        error = _validate_required_asset_file(path)
        if error:
            errors.append({"path": str(path), **error})
    return errors


def _validate_required_asset_file(path: Path) -> dict[str, str]:
    if not path.exists():
        return {"reason": "missing", "blocker_type": "missing_asset"}
    if not path.is_file():
        return {}
    try:
        size = path.stat().st_size
    except OSError as exc:
        return {"reason": f"stat_failed: {exc}", "blocker_type": "missing_asset"}
    if _is_git_lfs_pointer_file(path):
        return {"reason": "git_lfs_pointer_stub", "blocker_type": "git_lfs_pointer_stub"}
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        try:
            if path.read_bytes()[:5] != b"%PDF-":
                return {"reason": "not_pdf", "blocker_type": "missing_asset"}
        except OSError as exc:
            return {"reason": f"read_failed: {exc}", "blocker_type": "missing_asset"}
    if suffix in {".xlsx", ".xlsm", ".docx"} and not zipfile.is_zipfile(path):
        return {"reason": f"{suffix}_not_zip_container", "blocker_type": "missing_asset"}
    if suffix in {".pdf", ".xlsx", ".xlsm", ".docx"} and size < 128:
        return {"reason": "too_small", "blocker_type": "missing_asset"}
    return {}


def _asset_validation_error_message(errors: list[dict[str, str]]) -> str:
    if any(error.get("reason") == "git_lfs_pointer_stub" for error in errors):
        return "Required artifact assets are Git LFS pointer stubs, not materialized files"
    return "Required artifact assets are missing, corrupt, or incomplete"


def _artifact_asset_files(item: BenchmarkItem, root: Path) -> list[Path]:
    files: list[Path] = []
    for path in _artifact_asset_paths(item, root):
        if path.is_file():
            files.append(path)
        elif path.is_dir():
            files.extend(child for child in path.rglob("*") if child.is_file())
    return sorted(set(files))


def _is_git_lfs_pointer_file(path: Path) -> bool:
    try:
        if path.stat().st_size > 512:
            return False
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    return text.startswith("version https://git-lfs.github.com/spec/")


def _looks_like_incomplete_prompt_template(item: BenchmarkItem) -> bool:
    text = item.question.strip()
    lowered = text.lower()
    benchmark = normalize_text(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "")).replace("-", " ")
    if benchmark in {"edinet bench", "stockbench"} and not _has_referenced_or_inferred_assets(item):
        source = item.source.lower()
        if "prompt" in source or "template" in source or re.search(r"\b(agent|system)\s+prompt\b", lowered):
            return True
    template_endings = (
        "the report is as follows:",
        "report is as follows:",
        "input is as follows:",
        "data is as follows:",
        "json is as follows:",
        "as follows:",
        "provided below:",
    )
    if any(lowered.endswith(ending) for ending in template_endings):
        return not _has_referenced_or_inferred_assets(item)
    if re.search(r"\{\{\s*[^}]+\s*\}\}", text) or re.search(r"\{[a-zA-Z_][a-zA-Z0-9_]+\}", text):
        return not _has_referenced_or_inferred_assets(item)
    return False


def _has_referenced_or_inferred_assets(item: BenchmarkItem) -> bool:
    if _artifact_asset_paths(item, Path.cwd()):
        return True
    return bool(_flatten_string_values(item.metadata.get("input_files")))


def _missing_referenced_paths(item: BenchmarkItem, root: Path) -> list[str]:
    missing: list[str] = []
    for raw in _referenced_path_values(item):
        candidate = (root / raw).resolve()
        try:
            inside_root = candidate == root.resolve() or root.resolve() in candidate.parents
        except OSError:
            inside_root = False
        if not inside_root:
            missing.append(raw)
            continue
        if not candidate.exists():
            missing.append(raw)
    return sorted(set(missing))


def _referenced_path_values(item: BenchmarkItem) -> list[str]:
    values: list[str] = []
    for key in (
        "input_assets",
        "input_files",
        "asset_paths",
        "attachments",
        "files",
        "reference_files",
        "documents",
        "dataset_files",
    ):
        for raw in _flatten_string_values(item.metadata.get(key)):
            if _looks_like_path_value(raw):
                values.append(raw)
    return values


def _looks_like_path_value(value: str) -> bool:
    if not value or "\n" in value:
        return False
    if value.startswith(("http://", "https://", "s3://", "gs://")):
        return False
    return "/" in value or "\\" in value or bool(Path(value).suffix)


def _would_exceed_context_budget(item: BenchmarkItem) -> bool:
    context_limit = _model_context_limit()
    reserve = max(1024, int(context_limit * 0.05))
    return _estimated_prompt_tokens(item) + _max_answer_tokens(item) + reserve > context_limit


def _estimated_prompt_tokens(item: BenchmarkItem) -> int:
    choice_chars = sum(len(label) + len(value) + 4 for label, value in item.choices.items())
    total_chars = len(_model_visible_question(item)) + choice_chars + 1200
    return max(1, total_chars // 4)


def _estimated_payload_tokens(payload: dict[str, Any]) -> int:
    text = json.dumps(payload.get("messages", []), ensure_ascii=False, sort_keys=True)
    if "tools" in payload:
        text += json.dumps(payload.get("tools", []), ensure_ascii=False, sort_keys=True)
    return max(1, len(text) // 4)


def _token_budget_failure(payload: dict[str, Any]) -> dict[str, Any] | None:
    context_limit = _model_context_limit()
    reserve = max(1024, int(context_limit * 0.05))
    max_tokens = _bounded_int(payload.get("max_tokens"), _max_answer_tokens(BenchmarkItem("", "", "")), 1, 1_000_000)
    estimated = _estimated_payload_tokens(payload)
    if estimated + max_tokens + reserve <= context_limit:
        return None
    return {
        "status": "failed_token_budget",
        "reason": "estimated model prompt exceeds context budget with reserve",
        "estimated_prompt_tokens": estimated,
        "max_answer_tokens": max_tokens,
        "context_limit_tokens": context_limit,
        "reserve_tokens": reserve,
    }


def _model_context_limit() -> int:
    raw = os.environ.get("AGENT_BENCH_MODEL_CONTEXT_LIMIT", os.environ.get("AGENT_BENCH_CONTEXT_LIMIT", "32768"))
    return _bounded_int(raw, 32768, 2048, 2_000_000)


def run_agent_loop(
    benchmark: str,
    item: BenchmarkItem,
    tools: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    base_url = os.environ.get("AGENT_BENCH_BASE_URL", "").rstrip("/")
    model = os.environ.get("AGENT_BENCH_MODEL", "")
    headers = chat_completion_headers("target")

    direct_answer = _is_direct_answer_item(item)
    if tools is not None and not tools:
        direct_answer = True
    choice_text = ""
    if item.choices:
        choice_text = "\nChoices:\n" + "\n".join(f"{key}. {value}" for key, value in sorted(item.choices.items()))
        response_instruction = (
            'Return only compact JSON exactly like {"answer":"A","confidence":0.0}. '
            "The answer must be one listed choice label or choice text."
        )
    elif direct_answer:
        response_instruction = (
            'Return only compact JSON exactly like {"answer":"your final answer","confidence":0.0}. '
            "The answer value must be the exact final value with no explanation."
        )
    else:
        response_instruction = 'When finished, return JSON exactly like {"answer":"your final answer","confidence":0.0}.'
    messages: list[dict[str, Any]] = [
        {
            "role": "system",
            "content": _agent_system_prompt(direct_answer),
        },
        {
            "role": "user",
            "content": (
                "/no_think\n"
                f"Benchmark: {benchmark}\n"
                f"Question/task from benchmark data:\n{_model_visible_question(item)}"
                f"{choice_text}\n"
                f"{response_instruction}"
            ),
        },
    ]
    if tools is None:
        tools = agent_tool_schemas()
    tool_trace: list[dict[str, Any]] = []
    usage: dict[str, Any] = {}
    native_tools: bool | None = None
    last_content = ""
    final_answer_count = 0
    first_final_answer = ""
    direct_answer_retries = 0

    for _turn in range(_max_agent_turns()):
        payload: dict[str, Any] = {
            "model": model,
            "messages": messages,
            "temperature": 0,
            "top_p": 1,
            "max_tokens": _max_answer_tokens(item),
            "stream": False,
        }
        if direct_answer:
            payload["response_format"] = {"type": "json_object"}
        elif native_tools is not False and tools:
            payload["tools"] = tools
            payload["tool_choice"] = "auto"
        budget_failure = _token_budget_failure(payload)
        if budget_failure is not None:
            return _agent_loop_result(
                "",
                last_content,
                usage,
                tool_trace,
                final_answer_count,
                status="failed_token_budget",
                reason=str(budget_failure["reason"]),
                extra=budget_failure,
            )
        response_body, used_payload = post_chat_completion_with_variant(base_url, payload, headers)
        native_tools = "tools" in used_payload
        parsed = json.loads(response_body)
        if isinstance(parsed.get("usage"), dict):
            usage = _merge_usage(usage, parsed["usage"])
        message = extract_openai_message(parsed)
        finish_reason = extract_openai_finish_reason(parsed)
        content = extract_openai_content(parsed)
        tool_calls = extract_openai_tool_calls(parsed)
        if not content and not tool_calls and _message_has_hidden_reasoning(message):
            return _agent_loop_result(
                "",
                "",
                usage,
                tool_trace,
                final_answer_count,
                status="failed_model_format",
                reason="model produced hidden reasoning but no final content",
                extra={
                    "finish_reason": finish_reason,
                    "hidden_reasoning_no_final": True,
                    "no_final_content": True,
                },
            )
        last_content = content

        if tool_calls and native_tools:
            messages.append({"role": "assistant", "content": content, "tool_calls": tool_calls})
            final = handle_native_tool_calls(tool_calls, messages, tool_trace)
            if final is not None:
                final_answer_count = _final_answer_trace_count(tool_trace)
                if final and not first_final_answer:
                    first_final_answer = final
                answer = first_final_answer or final
                return _agent_loop_result(answer, final, usage, tool_trace, final_answer_count)
            protocol_failure = _protocol_limit_failure(tool_trace, final_answer_count, content)
            if protocol_failure:
                return _agent_loop_result(
                    "",
                    content,
                    usage,
                    tool_trace,
                    final_answer_count,
                    status="failed_model_tool_use",
                    reason=protocol_failure,
                )
            continue

        text_call = parse_text_tool_request(content)
        if text_call is not None:
            tool_name, arguments = text_call
            if tool_name == "final_answer":
                final_answer_count += 1
                answer = str(arguments.get("answer", ""))
                if answer and not first_final_answer:
                    first_final_answer = answer
                return _agent_loop_result(first_final_answer or answer, content, usage, tool_trace, final_answer_count)
            if not tools:
                return _agent_loop_result(
                    "",
                    content,
                    usage,
                    tool_trace,
                    final_answer_count,
                    status="failed_model_tool_use",
                    reason=f"model requested unavailable tool: {tool_name}",
                )
            result = execute_agent_tool(tool_name, arguments)
            tool_trace.append(
                {
                    "tool": tool_name,
                    "arguments": arguments,
                    "result": result[:1000],
                    "failed": _tool_result_failed(tool_name, result),
                }
            )
            protocol_failure = _protocol_limit_failure(tool_trace, final_answer_count, content)
            if protocol_failure:
                return _agent_loop_result(
                    "",
                    content,
                    usage,
                    tool_trace,
                    final_answer_count,
                    status="failed_model_tool_use",
                    reason=protocol_failure,
                )
            messages.append({"role": "assistant", "content": content})
            messages.append(
                {
                    "role": "user",
                    "content": f"Tool result for {tool_name}:\n{result}\n\nContinue or provide the final JSON answer.",
                }
            )
            native_tools = False
            continue

        if not direct_answer and _looks_like_intermediate_agent_message(content):
            messages.append({"role": "assistant", "content": content})
            messages.append(
                {
                    "role": "user",
                    "content": (
                        "Continue the task now. If work remains, call exactly one available tool. "
                        "If the deliverable or answer is complete, return only compact final JSON."
                    ),
                }
            )
            continue

        answer = extract_answer(content)
        if answer:
            if direct_answer and direct_answer_retries < 1 and _needs_direct_answer_retry(answer, item):
                direct_answer_retries += 1
                messages.append({"role": "assistant", "content": content})
                messages.append(
                    {
                        "role": "user",
                        "content": _direct_answer_retry_instruction(item),
                    }
                )
                continue
            return _agent_loop_result(answer, content, usage, tool_trace, final_answer_count)
        messages.append({"role": "assistant", "content": content})
        messages.append({"role": "user", "content": "The previous response was empty. Provide the final JSON answer."})

    return {
        "answer": extract_answer(last_content),
        "content": last_content,
        "usage": usage,
        "tool_trace": tool_trace,
        "diagnostics": protocol_diagnostics(tool_trace, last_content, final_answer_count),
    }


def _tools_sent_to_model(item: BenchmarkItem, tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not tools:
        return []
    if "tool_call" in _item_required_capabilities(item) or _item_required_tools(item):
        return tools
    if _is_direct_answer_item(item):
        return []
    return tools


def _tool_manifest(
    item: BenchmarkItem,
    available_tools: list[dict[str, Any]],
    sent_tools: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "tools": sent_tools,
        "sent_to_model": sent_tools,
        "available_in_runner": available_tools,
        "suppressed_for_direct_answer": bool(available_tools) and not sent_tools and _is_direct_answer_item(item),
    }


def _agent_loop_result(
    answer: str,
    content: str,
    usage: dict[str, Any],
    tool_trace: list[dict[str, Any]],
    final_answer_count: int,
    status: str = "",
    reason: str = "",
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    diagnostics = protocol_diagnostics(tool_trace, content, final_answer_count)
    if status:
        diagnostics["status"] = status
    if reason:
        diagnostics["reason"] = reason
    if extra:
        diagnostics.update(extra)
    return {
        "answer": answer,
        "content": content,
        "usage": usage,
        "tool_trace": tool_trace,
        "diagnostics": diagnostics,
    }


def protocol_diagnostics(
    tool_trace: list[dict[str, Any]],
    final_content: str,
    final_answer_count: int,
) -> dict[str, Any]:
    failed_tools = [
        item
        for item in tool_trace
        if isinstance(item, dict) and bool(item.get("failed"))
    ]
    tool_names = [str(item.get("tool", "")) for item in tool_trace if isinstance(item, dict)]
    repeated_identical_calls = 0
    seen_calls: set[str] = set()
    for item in tool_trace:
        if not isinstance(item, dict):
            continue
        signature = json.dumps(
            {"tool": item.get("tool"), "arguments": item.get("arguments")},
            ensure_ascii=False,
            sort_keys=True,
        )
        if signature in seen_calls:
            repeated_identical_calls += 1
        seen_calls.add(signature)
    return {
        "tool_call_count": len(tool_trace),
        "tool_failure_count": len(failed_tools),
        "ignored_tool_failure_before_final": bool(failed_tools and final_content.strip()),
        "final_answer_count": final_answer_count,
        "repeated_identical_tool_call_count": repeated_identical_calls,
        "final_output_contains_tool_syntax": _contains_tool_syntax(final_content),
        "tool_names": tool_names,
    }


def _protocol_limit_failure(
    tool_trace: list[dict[str, Any]],
    final_answer_count: int,
    final_content: str,
) -> str:
    if final_answer_count > 1:
        return "model attempted to submit more than one final answer"
    if len(tool_trace) > _max_total_tool_calls():
        return f"model exceeded maximum total tool calls ({_max_total_tool_calls()})"
    repeated = _max_identical_tool_call_count(tool_trace)
    if repeated > MAX_REPEATED_IDENTICAL_TOOL_CALLS:
        return (
            "model repeated an identical tool call "
            f"{repeated} times; maximum is {MAX_REPEATED_IDENTICAL_TOOL_CALLS}"
        )
    failed = _consecutive_failed_tool_calls(tool_trace)
    if failed > MAX_CONSECUTIVE_FAILED_TOOL_CALLS:
        return (
            "model made too many consecutive failed tool calls "
            f"({failed}; maximum is {MAX_CONSECUTIVE_FAILED_TOOL_CALLS})"
        )
    if _contains_tool_syntax(final_content) and not tool_trace:
        return "model emitted tool-call syntax without a valid tool request"
    return ""


def _max_identical_tool_call_count(tool_trace: list[dict[str, Any]]) -> int:
    counts: dict[str, int] = {}
    max_count = 0
    for item in tool_trace:
        if not isinstance(item, dict):
            continue
        if item.get("tool") == "final_answer":
            continue
        signature = json.dumps(
            {"tool": item.get("tool"), "arguments": item.get("arguments")},
            ensure_ascii=False,
            sort_keys=True,
        )
        counts[signature] = counts.get(signature, 0) + 1
        max_count = max(max_count, counts[signature])
    return max_count


def _consecutive_failed_tool_calls(tool_trace: list[dict[str, Any]]) -> int:
    count = 0
    for item in reversed(tool_trace):
        if not isinstance(item, dict) or item.get("tool") == "final_answer":
            continue
        if not item.get("failed"):
            break
        count += 1
    return count


def _final_answer_trace_count(tool_trace: list[dict[str, Any]]) -> int:
    return sum(1 for item in tool_trace if isinstance(item, dict) and item.get("tool") == "final_answer")


def _max_total_tool_calls() -> int:
    return _bounded_int(
        os.environ.get("AGENT_BENCH_MAX_TOTAL_TOOL_CALLS"),
        DEFAULT_MAX_TOTAL_TOOL_CALLS,
        1,
        10_000,
    )


def _tool_result_failed(tool_name: str, result: str) -> bool:
    prefix = result.strip().lower()
    if prefix.startswith("tool ") and " failed:" in prefix[:120]:
        return True
    if prefix.startswith("unknown tool:"):
        return True
    if tool_name == "run_command" and re.search(r"\bexit_code=(?!0\b)\d+", result):
        return True
    parsed = parse_json_object(result)
    if isinstance(parsed, dict):
        payload = parsed.get("result")
        if isinstance(payload, dict) and payload.get("fixture_valid") is False:
            return True
        if isinstance(payload, dict) and payload.get("error"):
            return True
    return any(
        prefix.startswith(marker)
        for marker in (
            "missing ",
            "invalid ",
            "unable to ",
            "file not found:",
            "path does not exist:",
            "refusing to ",
            "run_command expects ",
            "run_command legacy ",
            "unsupported ",
        )
    )


def _contains_tool_syntax(content: str) -> bool:
    lowered = content.lower()
    return any(
        marker in lowered
        for marker in (
            "<tool_call",
            "<|tool_call",
            "</tool_call",
            "<tool_code",
            "</tool_code",
            '"tool":',
            '"tool_name":',
        )
    )


def _fintoolbench_tool_schemas(
    root: Path,
    *,
    required_tools: set[str] | None = None,
) -> list[dict[str, Any]]:
    tool_path = root / "tools" / "tools_all_annotated.jsonl"
    if not tool_path.is_file():
        return []
    manifest = _fintoolbench_tool_manifest(tool_path)
    selected = sorted(required_tools or manifest)
    schemas: list[dict[str, Any]] = []
    for name in selected:
        record = manifest.get(name)
        if record is None:
            continue
        schemas.append(_fintoolbench_tool_schema(name, record))
    return schemas


def _fintoolbench_tool_schema(name: str, record: dict[str, Any]) -> dict[str, Any]:
    description = _first_text(
        record,
        (
            "description",
            "tool_description",
            "api_description",
            "doc",
            "documentation",
            "summary",
        ),
    )
    parameters = _tool_parameters_schema(record)
    return {
        "type": "function",
        "function": {
            "name": name,
            "description": truncate(description or f"Run the FinToolBench financial tool {name}.", 900),
            "parameters": parameters,
        },
    }


def _tool_parameters_schema(record: dict[str, Any]) -> dict[str, Any]:
    for key in ("parameters", "args", "arguments", "input_schema", "schema"):
        raw = record.get(key)
        parsed = _parse_tool_schema_value(raw)
        if isinstance(parsed, dict) and parsed.get("type") == "object":
            parsed.setdefault("additionalProperties", True)
            return parsed
        if isinstance(parsed, dict) and parsed:
            return {
                "type": "object",
                "properties": {
                    str(name): _json_schema_for_tool_parameter(value)
                    for name, value in parsed.items()
                    if str(name).strip()
                },
                "additionalProperties": True,
            }
        if isinstance(parsed, list) and parsed:
            properties: dict[str, Any] = {}
            required: list[str] = []
            for item in parsed:
                if isinstance(item, str):
                    properties[item] = {"type": "string"}
                    continue
                if not isinstance(item, dict):
                    continue
                name = _first_text(item, ("name", "parameter", "param", "key"))
                if not name:
                    continue
                properties[name] = _json_schema_for_tool_parameter(item)
                if item.get("required") is True:
                    required.append(name)
            if properties:
                schema = {"type": "object", "properties": properties, "additionalProperties": True}
                if required:
                    schema["required"] = sorted(set(required))
                return schema
    return {
        "type": "object",
        "properties": {
            "query": {"type": "string", "description": "Natural-language tool query."},
            "symbol": {"type": "string", "description": "Ticker, company, or instrument identifier."},
            "date": {"type": "string", "description": "Optional date or date range."},
        },
        "additionalProperties": True,
    }


def _parse_tool_schema_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            names = re.findall(r"[A-Za-z_][A-Za-z0-9_]*", value)
            return {name: {"type": "string"} for name in names[:12]}
    return None


def _json_schema_for_tool_parameter(value: Any) -> dict[str, Any]:
    if isinstance(value, dict) and isinstance(value.get("schema"), dict):
        return value["schema"]
    if isinstance(value, dict):
        raw_type = str(value.get("type") or value.get("data_type") or "string").lower()
        json_type = {
            "str": "string",
            "string": "string",
            "int": "integer",
            "integer": "integer",
            "float": "number",
            "number": "number",
            "bool": "boolean",
            "boolean": "boolean",
            "array": "array",
            "list": "array",
            "object": "object",
            "dict": "object",
        }.get(raw_type, "string")
        schema: dict[str, Any] = {"type": json_type}
        description = _first_text(value, ("description", "desc", "help"))
        if description:
            schema["description"] = truncate(description, 500)
        return schema
    return {"type": "string"}


def _finance_agent_v2_missing_environment() -> list[str]:
    return [key for key in FINANCE_AGENT_V2_REQUIRED_ENV if not os.environ.get(key)]


def _finance_agent_v2_available_tool_schemas(root: Path | None = None) -> list[dict[str, Any]]:
    return _finance_agent_v2_tool_schemas() if _finance_agent_v2_backend_status(root)["ready"] else []


def _finance_agent_v2_tool_schemas() -> list[dict[str, Any]]:
    descriptions = {
        "web_search": "Search the local/upstream finance evidence corpus for public web information.",
        "edgar_search": "Search SEC/EDGAR filings and filing metadata.",
        "parse_html_page": "Extract readable text from a local HTML page or cached web page.",
        "retrieve_information": "Retrieve relevant snippets from local benchmark evidence files.",
        "price_history": "Return deterministic cached price history for a ticker or symbol.",
    }
    return [
        {
            "type": "function",
            "function": {
                "name": name,
                "description": descriptions[name],
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "symbol": {"type": "string"},
                        "url": {"type": "string"},
                        "start_date": {"type": "string"},
                        "end_date": {"type": "string"},
                    },
                    "additionalProperties": True,
                },
            },
        }
        for name in sorted(FINANCE_AGENT_V2_REQUIRED_TOOLS)
    ]


def _finance_agent_v2_backend_unavailable_reason(
    backend: dict[str, Any],
    missing_tools: list[str],
    missing_env: list[str],
) -> str:
    reason = str(backend.get("reason") or "Finance Agent v2 fixture backend is unavailable")
    details: list[str] = [reason]
    if missing_tools:
        details.append("missing benchmark tools: " + ", ".join(missing_tools))
    if missing_env:
        details.append("live credentials absent: " + ", ".join(missing_env))
    details.append("live service backend is not implemented/configured")
    return "; ".join(details)


def _finance_agent_v2_backend_status(root: Path | None = None) -> dict[str, Any]:
    store = _finance_agent_v2_fixture_store(root)
    searched = [str(path) for path in _finance_agent_v2_fixture_roots(root)]
    if store is None:
        return {
            "ready": False,
            "mode": "deterministic_fixture_finance_agent_v2_tools",
            "reason": "Finance Agent v2 deterministic fixture store was not found",
            "searched_roots": searched,
        }
    manifest = _read_json_payload(store / "manifest.json")
    errors: list[str] = []
    if not isinstance(manifest, dict):
        errors.append("manifest.json is missing or malformed")
        manifest = {}
    ticker = str(manifest.get("ticker") or "").upper()
    if ticker != "CRWD":
        errors.append("manifest ticker must be CRWD")
    validated_files, file_errors = _finance_agent_v2_validate_manifest_files(store, manifest)
    errors.extend(file_errors)
    fixture_payload = _finance_agent_v2_fixture_payload(store)
    errors.extend(_finance_agent_v2_validate_fixture_payload(fixture_payload))
    canaries = _finance_agent_v2_fixture_canaries(fixture_payload)
    failed_canaries = [name for name, passed in canaries.items() if not passed]
    if failed_canaries:
        errors.append("fixture canaries failed: " + ", ".join(failed_canaries))
    status = {
        "ready": not errors,
        "mode": "deterministic_fixture_finance_agent_v2_tools",
        "fixture_root": str(store),
        "symbol": ticker or "CRWD",
        "manifest_version": manifest.get("version"),
        "validated_files": validated_files,
        "canaries": canaries,
        "searched_roots": searched,
    }
    if errors:
        status["reason"] = errors[0]
        status["errors"] = errors
    return status


def _finance_agent_v2_fixture_roots(root: Path | None = None) -> list[Path]:
    candidates: list[Path] = []
    env_root = os.environ.get(FINANCE_AGENT_V2_FIXTURE_ROOT_ENV)
    if env_root:
        return [Path(env_root).expanduser()]
    if root is not None:
        candidates.append(root / "fixtures" / "finance_agent_v2")
    candidates.append(Path.cwd() / "fixtures" / "finance_agent_v2")
    try:
        candidates.append(Path(__file__).resolve().parents[1] / "fixtures" / "finance_agent_v2")
    except IndexError:
        pass
    candidates.append(Path("/opt/agent-bench/fixtures/finance_agent_v2"))
    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate)
        if key not in seen:
            unique.append(candidate)
            seen.add(key)
    return unique


def _finance_agent_v2_fixture_store(root: Path | None = None) -> Path | None:
    for fixture_root in _finance_agent_v2_fixture_roots(root):
        for candidate in (fixture_root / "crwd", fixture_root):
            if (candidate / "manifest.json").is_file():
                return candidate
    return None


def _read_json_payload(path: Path) -> Any:
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return None


def _finance_agent_v2_validate_manifest_files(
    store: Path,
    manifest: dict[str, Any],
) -> tuple[list[dict[str, str]], list[str]]:
    files = manifest.get("files")
    if not isinstance(files, dict) or not files:
        return [], ["manifest files map is missing"]
    validated: list[dict[str, str]] = []
    errors: list[str] = []
    for relative, expected in sorted(files.items()):
        path = store / str(relative)
        if not path.is_file():
            errors.append(f"fixture file is missing: {relative}")
            continue
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        expected_hash = expected.get("sha256") if isinstance(expected, dict) else expected
        if isinstance(expected_hash, str) and expected_hash and digest != expected_hash:
            errors.append(f"fixture checksum mismatch: {relative}")
        validated.append({"path": str(relative), "sha256": digest})
    return validated, errors


def _finance_agent_v2_fixture_payload(store: Path) -> dict[str, Any]:
    return {
        "manifest": _read_json_payload(store / "manifest.json") or {},
        "search_results": _read_json_payload(store / "search_results.json") or [],
        "edgar_filings": _read_json_payload(store / "edgar_filings_index.json") or [],
        "retrieval_index": _read_json_payload(store / "retrieval_index.json") or [],
        "price_history": _read_json_payload(store / "price_history.json") or {},
        "documents": _finance_agent_v2_documents(store),
    }


def _finance_agent_v2_documents(store: Path) -> list[dict[str, Any]]:
    manifest = _read_json_payload(store / "manifest.json")
    documents = manifest.get("documents") if isinstance(manifest, dict) else []
    if not isinstance(documents, list):
        return []
    loaded: list[dict[str, Any]] = []
    for document in documents:
        if not isinstance(document, dict):
            continue
        parsed_path = str(document.get("parsed_path") or "")
        html_path = str(document.get("html_path") or "")
        parsed_text = ""
        if parsed_path and (store / parsed_path).is_file():
            parsed_text = (store / parsed_path).read_text(encoding="utf-8", errors="replace")
        html_text = ""
        if html_path and (store / html_path).is_file():
            html_text = (store / html_path).read_text(encoding="utf-8", errors="replace")
        loaded.append({**document, "parsed_text": parsed_text, "html_text": html_text})
    return loaded


def _finance_agent_v2_validate_fixture_payload(payload: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not isinstance(payload.get("search_results"), list) or not payload["search_results"]:
        errors.append("search_results.json must contain at least one result")
    filings = payload.get("edgar_filings")
    if not isinstance(filings, list) or not filings:
        errors.append("edgar_filings_index.json must contain filing records")
    else:
        required = {("10-K", "2025-01-31"), ("10-Q", "2025-10-31")}
        present = {
            (str(filing.get("form") or ""), str(filing.get("report_date") or ""))
            for filing in filings
            if isinstance(filing, dict)
        }
        missing = sorted(required - present)
        if missing:
            errors.append("edgar filings missing required periods: " + ", ".join(f"{form} {date}" for form, date in missing))
    if not isinstance(payload.get("retrieval_index"), list) or not payload["retrieval_index"]:
        errors.append("retrieval_index.json must contain evidence snippets")
    if not isinstance(payload.get("price_history"), dict) or "CRWD" not in payload["price_history"]:
        errors.append("price_history.json must contain CRWD history")
    documents = payload.get("documents")
    if not isinstance(documents, list) or len(documents) < 2:
        errors.append("manifest must define the FY2025 10-K and October 2025 10-Q documents")
    elif any(not str(document.get("parsed_text") or "").strip() for document in documents):
        errors.append("all Finance Agent v2 fixture documents must have parsed text")
    return errors


def _finance_agent_v2_fixture_canaries(payload: dict[str, Any]) -> dict[str, bool]:
    evidence_texts = [
        _text_from_value(payload.get("search_results")),
        _text_from_value(payload.get("edgar_filings")),
        _text_from_value(payload.get("retrieval_index")),
    ]
    for document in payload.get("documents", []):
        if isinstance(document, dict):
            evidence_texts.append(str(document.get("parsed_text") or ""))
    combined = normalize_text("\n".join(evidence_texts))
    return {
        name: all(term in combined for term in terms)
        for name, terms in FINANCE_AGENT_V2_CANARY_TERMS
    }


def _finance_agent_v2_matching_records(records: Any, query: str, *, max_results: int) -> list[dict[str, Any]]:
    if not isinstance(records, list):
        return []
    terms = [term for term in re.findall(r"[A-Za-z0-9.$_-]{3,}", normalize_text(query)) if term not in {"the", "and"}]
    scored: list[tuple[int, int, dict[str, Any]]] = []
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            continue
        text = normalize_text(_text_from_value(record))
        score = sum(1 for term in terms if term in text)
        if score or not terms:
            scored.append((score, -index, record))
    scored.sort(reverse=True)
    return [record for _, _, record in scored[:max_results]]


def _finance_agent_v2_document_for_argument(documents: list[dict[str, Any]], arguments: dict[str, Any]) -> dict[str, Any] | None:
    needle = normalize_text(" ".join(_flatten_string_values(arguments)))
    if not needle and documents:
        return documents[0]
    for document in documents:
        haystack = normalize_text(
            " ".join(
                str(document.get(key) or "")
                for key in ("url", "html_path", "parsed_path", "primary_document", "form", "report_date")
            )
        )
        if needle and (needle in haystack or any(part and part in haystack for part in needle.split())):
            return document
    return documents[0] if documents else None


def _looks_like_intermediate_agent_message(content: str) -> bool:
    text = normalize_text(content)
    if not text or _contains_tool_syntax(content):
        return False
    if isinstance(parse_json_object(content), dict):
        return False
    markers = (
        "i will ",
        "i ll ",
        "let me ",
        "let s ",
        "i should ",
        "i need to ",
        "now i will ",
        "i m going to ",
        "i am going to ",
        "try to ",
        "should probably ",
        "need to ",
        "wait ",
        "first i ",
        "next i ",
        "the directory exists",
        "file has been written",
    )
    return any(marker in f" {text} " for marker in markers)


def agent_tool_schemas() -> list[dict[str, Any]]:
    return [
        {
            "type": "function",
            "function": {
                "name": "list_files",
                "description": "List files in the benchmark repository.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Directory path relative to the repository root."},
                        "pattern": {"type": "string", "description": "Optional substring or suffix filter."},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 200},
                    },
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_file",
                "description": "Read a text file from the benchmark repository.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "start_line": {"type": "integer", "minimum": 1},
                        "max_lines": {"type": "integer", "minimum": 1, "maximum": 400},
                        "allow_binary": {"type": "boolean"},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "search_files",
                "description": "Search text files for a literal query.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "path": {"type": "string"},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 100},
                    },
                    "required": ["query"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "write_file",
                "description": "Deprecated alias for write_text_file. Create or replace a UTF-8 text file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "content": {"type": "string"},
                    },
                    "required": ["path", "content"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "write_text_file",
                "description": "Create or replace a UTF-8 text file in the benchmark repository.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "content": {"type": "string"},
                    },
                    "required": ["path", "content"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "write_base64_file",
                "description": "Create or replace a binary file from base64-encoded content.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "base64_content": {"type": "string"},
                    },
                    "required": ["path", "base64_content"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_binary_metadata",
                "description": "Return size, suffix, and basic integrity metadata for a binary or archive file.",
                "parameters": {
                    "type": "object",
                    "properties": {"path": {"type": "string"}},
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "extract_archive",
                "description": "Extract a .zip archive inside the benchmark repository.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "output_dir": {"type": "string"},
                    },
                    "required": ["path", "output_dir"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_spreadsheet",
                "description": "Read a small preview from a CSV or XLSX spreadsheet.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "sheet": {"type": "string"},
                        "max_rows": {"type": "integer", "minimum": 1, "maximum": 50},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_pdf_text",
                "description": "Extract text from a PDF using installed PDF libraries or command-line tools.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "max_chars": {"type": "integer", "minimum": 100, "maximum": 12000},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "list_artifacts",
                "description": "List files currently generated under an output/artifact directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 200},
                    },
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "apply_patch",
                "description": "Apply a unified diff patch to repository files.",
                "parameters": {
                    "type": "object",
                    "properties": {"patch": {"type": "string"}},
                    "required": ["patch"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "run_command",
                "description": "Run a local repository command without a shell and return stdout/stderr.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "argv": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Executable and arguments, for example [\"python3\", \"-c\", \"print('ok')\"].",
                        },
                        "stdin": {"type": "string"},
                        "timeout_seconds": {"type": "integer", "minimum": 1, "maximum": 120},
                    },
                    "required": ["argv"],
                    "additionalProperties": False,
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "final_answer",
                "description": "Submit the final answer to be graded.",
                "parameters": {
                    "type": "object",
                    "properties": {"answer": {"type": "string"}},
                    "required": ["answer"],
                    "additionalProperties": False,
                },
            },
        },
    ]


def handle_native_tool_calls(
    tool_calls: list[dict[str, Any]],
    messages: list[dict[str, Any]],
    tool_trace: list[dict[str, Any]],
) -> str | None:
    for tool_call in tool_calls:
        function = tool_call.get("function") if isinstance(tool_call.get("function"), dict) else {}
        tool_name = function.get("name") if isinstance(function.get("name"), str) else ""
        arguments = _parse_tool_arguments(function.get("arguments"))
        if tool_name == "final_answer":
            candidate = str(arguments.get("answer", ""))
            result = "Final answer accepted."
            tool_trace.append(
                {
                    "tool": tool_name,
                    "arguments": arguments,
                    "result": result,
                    "failed": False,
                }
            )
            messages.append(
                {
                    "role": "tool",
                    "tool_call_id": str(tool_call.get("id", tool_name or "tool_call")),
                    "content": result,
                }
            )
            return candidate
        else:
            result = execute_agent_tool(tool_name, arguments)
        tool_trace.append(
            {
                "tool": tool_name,
                "arguments": arguments,
                "result": result[:1000],
                "failed": _tool_result_failed(tool_name, result),
            }
        )
        messages.append(
            {
                "role": "tool",
                "tool_call_id": str(tool_call.get("id", tool_name or "tool_call")),
                "content": result,
            }
        )
    return None


def parse_text_tool_request(content: str) -> tuple[str, dict[str, Any]] | None:
    parsed = parse_json_object(content)
    if not isinstance(parsed, dict):
        return parse_angle_tool_request(content)
    return _tool_request_from_mapping(parsed)


def _tool_request_from_mapping(parsed: dict[str, Any]) -> tuple[str, dict[str, Any]] | None:
    if isinstance(parsed.get("function"), dict):
        nested = _tool_request_from_mapping(parsed["function"])
        if nested is not None:
            return nested
    if isinstance(parsed.get("tool"), str):
        name = parsed["tool"]
    elif isinstance(parsed.get("tool_name"), str):
        name = parsed["tool_name"]
    elif isinstance(parsed.get("name"), str) and "arguments" in parsed:
        name = parsed["name"]
    elif isinstance(parsed.get("name"), str) and any(key in parsed for key in ("parameters", "params", "args", "input")):
        name = parsed["name"]
    elif "final_answer" in parsed:
        return "final_answer", {"answer": str(parsed.get("final_answer", ""))}
    else:
        return None
    arguments = _first_tool_arguments(parsed)
    if not isinstance(arguments, dict):
        arguments = {
            key: value
            for key, value in parsed.items()
            if key not in {"tool", "tool_name", "name", "function", "arguments", "parameters", "params", "args", "input"}
        }
    return name, arguments


def _first_tool_arguments(parsed: dict[str, Any]) -> dict[str, Any] | None:
    for key in ("arguments", "parameters", "params", "args", "input"):
        if key in parsed:
            value = parsed.get(key)
            if isinstance(value, dict):
                return value
            arguments = _parse_tool_arguments(value)
            return arguments
    return None


def parse_angle_tool_request(content: str) -> tuple[str, dict[str, Any]] | None:
    text = content.strip()
    match = re.match(
        r"(?:<\|?tool_call\|?>\s*)?call:(?P<name>[A-Za-z_][A-Za-z0-9_]*)\{(?P<body>.*)\}(?:\s*<\|?tool_call\|?>)?\s*$",
        text,
        re.DOTALL,
    )
    if not match:
        return None
    name = match.group("name")
    arguments = _parse_angle_tool_arguments(match.group("body"))
    if not arguments:
        return None
    return name, arguments


def _parse_angle_tool_arguments(body: str) -> dict[str, Any]:
    arguments: dict[str, Any] = {}
    index = 0
    while index < len(body):
        key_match = re.search(r"([A-Za-z_][A-Za-z0-9_]*)\s*:", body[index:])
        if not key_match:
            break
        key = key_match.group(1)
        value_start = index + key_match.end()
        value, index = _parse_angle_tool_value(body, value_start)
        arguments[key] = value
        while index < len(body) and body[index] in ", \n\t":
            index += 1
    return arguments


def _parse_angle_tool_value(body: str, index: int) -> tuple[Any, int]:
    while index < len(body) and body[index].isspace():
        index += 1
    if body.startswith("[", index):
        end = _find_matching_bracket(body, index, "[", "]")
        if end == -1:
            return body[index + 1 :].strip(), len(body)
        raw = body[index + 1 : end]
        values = re.findall(r"<\|\"\|>(.*?)<\|\"\|>", raw, flags=re.DOTALL)
        if not values:
            values = [part.strip().strip("'\"") for part in raw.split(",") if part.strip()]
        return values, end + 1
    if body.startswith("<|\"|>", index):
        end = body.find("<|\"|>", index + 5)
        if end == -1:
            return body[index + 5 :].strip(), len(body)
        return body[index + 5 : end], end + 5
    next_comma = body.find(",", index)
    if next_comma == -1:
        return body[index:].strip(), len(body)
    return body[index:next_comma].strip(), next_comma + 1


def _find_matching_bracket(text: str, start: int, opener: str, closer: str) -> int:
    depth = 0
    for index in range(start, len(text)):
        if text[index] == opener:
            depth += 1
        elif text[index] == closer:
            depth -= 1
            if depth == 0:
                return index
    return -1


def execute_agent_tool(name: str, arguments: dict[str, Any]) -> str:
    name = _canonical_tool_name(name)
    try:
        if name == "list_files":
            return tool_list_files(arguments)
        if name == "read_file":
            return tool_read_file(arguments)
        if name == "search_files":
            return tool_search_files(arguments)
        if name == "write_file":
            return tool_write_file(arguments)
        if name == "write_text_file":
            return tool_write_file(arguments)
        if name == "write_base64_file":
            return tool_write_base64_file(arguments)
        if name == "read_binary_metadata":
            return tool_read_binary_metadata(arguments)
        if name == "extract_archive":
            return tool_extract_archive(arguments)
        if name == "read_spreadsheet":
            return tool_read_spreadsheet(arguments)
        if name == "read_pdf_text":
            return tool_read_pdf_text(arguments)
        if name == "list_artifacts":
            return tool_list_artifacts(arguments)
        if name == "apply_patch":
            return tool_apply_patch(arguments)
        if name == "run_command":
            return tool_run_command(arguments)
        benchmark_result = execute_benchmark_tool(name, arguments)
        if benchmark_result is not None:
            return benchmark_result
        return f"Unknown tool: {name}"
    except Exception as exc:
        return f"Tool {name} failed: {exc}"


def execute_benchmark_tool(name: str, arguments: dict[str, Any]) -> str | None:
    if name in FINANCE_AGENT_V2_REQUIRED_TOOLS:
        return _execute_finance_agent_v2_tool(name, arguments)
    fintool_names = _fintoolbench_executable_tool_names()
    if name in fintool_names:
        return _execute_fintoolbench_tool(name, arguments)
    return None


def _fintoolbench_executable_tool_names() -> set[str]:
    raw = os.environ.get(FINTOOLBENCH_EXECUTABLE_TOOLS_ENV, "")
    if raw.strip():
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return {name.strip() for name in raw.split(",") if name.strip()}
        return {str(name).strip() for name in _flatten_string_values(payload) if str(name).strip()}
    return set(_tool_schema_names(_fintoolbench_tool_schemas(Path.cwd())))


def _fintoolbench_valid_backend_tools(root: Path, required_tools: set[str]) -> set[str]:
    canaries = _fintoolbench_backend_canaries(root, required_tools)
    return {tool for tool, canary in canaries.items() if canary.get("passed") is True}


def _fintoolbench_backend_canaries(root: Path, required_tools: set[str]) -> dict[str, dict[str, Any]]:
    tools = sorted(required_tools or FINTOOLBENCH_REQUIRED_CANARY_ARGS)
    canaries: dict[str, dict[str, Any]] = {}
    for tool in tools:
        args = FINTOOLBENCH_REQUIRED_CANARY_ARGS.get(tool)
        if args is None:
            canaries[tool] = {
                "passed": False,
                "reason": "no deterministic fixture canary is defined for this FinToolBench tool",
            }
            continue
        fixture = _load_fintoolbench_fixture(tool, args, root)
        if fixture.get("error"):
            canaries[tool] = {
                "passed": False,
                "reason": fixture["error"],
                "fixture_path": fixture.get("fixture_path", ""),
            }
            continue
        canaries[tool] = _validate_fintoolbench_fixture(tool, fixture.get("payload"), args)
        canaries[tool]["fixture_path"] = fixture.get("fixture_path", "")
    return canaries


def _load_fintoolbench_fixture(name: str, arguments: dict[str, Any], root: Path | None = None) -> dict[str, Any]:
    fixture_key = _fintoolbench_fixture_key(arguments)
    if not fixture_key:
        return {"error": "FinToolBench fixture lookup requires a symbol/ticker argument"}
    for fixture_root in _fintoolbench_fixture_roots(root):
        path = fixture_root / name / f"{fixture_key}.json"
        if not path.is_file():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            return {"error": f"FinToolBench fixture is unreadable: {exc}", "fixture_path": str(path)}
        return {"payload": payload, "fixture_path": str(path)}
    searched = [str(path / name / f"{fixture_key}.json") for path in _fintoolbench_fixture_roots(root)]
    return {"error": "FinToolBench fixture was not found", "searched_paths": searched}


def _fintoolbench_fixture_roots(root: Path | None = None) -> list[Path]:
    candidates: list[Path] = []
    explicit = os.environ.get(FINTOOLBENCH_FIXTURE_ROOT_ENV, "").strip()
    if explicit:
        return [Path(explicit).expanduser()]
    if root is not None:
        candidates.append(root / "fixtures" / "fintoolbench")
    candidates.append(Path.cwd() / "fixtures" / "fintoolbench")
    try:
        candidates.append(Path(__file__).resolve().parents[1] / "fixtures" / "fintoolbench")
    except IndexError:
        pass
    candidates.append(Path("/opt/agent-bench/fixtures/fintoolbench"))
    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate)
        if key not in seen:
            seen.add(key)
            unique.append(candidate)
    return unique


def _fintoolbench_fixture_key(arguments: dict[str, Any]) -> str:
    for key in ("symbol", "ticker", "company", "query"):
        value = arguments.get(key)
        if isinstance(value, str) and value.strip():
            token = value.strip().split()[0].upper()
            return "".join(ch for ch in token if ch.isalnum() or ch in {"_", "-"})
    return ""


def _validate_fintoolbench_fixture(name: str, payload: Any, arguments: dict[str, Any]) -> dict[str, Any]:
    if name != "companies_balance_sheet_statements":
        return {"passed": False, "reason": f"no semantic validator is implemented for {name}"}
    if isinstance(payload, dict) and "matching_records" in payload:
        return {"passed": False, "reason": "fixture contains retrieval snippets, not structured statement records"}
    records = _fintoolbench_fixture_records(payload)
    if not records:
        return {"passed": False, "reason": "fixture does not contain structured statement records"}
    symbol = _fintoolbench_fixture_key(arguments)
    if symbol and not any(str(record.get("symbol") or "").upper() == symbol for record in records):
        return {"passed": False, "reason": f"fixture records do not contain symbol {symbol}"}
    fy2018 = [record for record in records if _fintoolbench_record_year(record) == "2018"]
    if not fy2018:
        return {"passed": False, "reason": "fixture does not contain an FY2018 annual record"}
    for record in fy2018:
        value = _fintoolbench_first_numeric(record, FINTOOLBENCH_NET_PPE_KEYS)
        if value is None:
            continue
        billions = value / 1_000_000_000 if abs(value) > 10_000 else value
        if abs(billions - 8.70) <= 0.01:
            return {
                "passed": True,
                "validated_symbol": symbol,
                "validated_year": "2018",
                "validated_field": _fintoolbench_first_present_key(record, FINTOOLBENCH_NET_PPE_KEYS),
                "validated_value": value,
                "validated_value_billions": round(billions, 4),
            }
    return {"passed": False, "reason": "fixture FY2018 record lacks net PP&E value 8.70B"}


def _fintoolbench_fixture_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("records", "data", "statements", "balance_sheet_statements", "annualReports", "result"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        return [payload]
    return []


def _fintoolbench_record_year(record: dict[str, Any]) -> str:
    for key in ("calendarYear", "fiscalYear", "year"):
        value = record.get(key)
        if isinstance(value, (str, int)) and str(value).strip():
            return str(value).strip()[:4]
    for key in ("date", "fiscalDateEnding", "periodEndDate", "reportedDate"):
        value = record.get(key)
        if isinstance(value, str):
            match = re.search(r"\b(20\d{2}|19\d{2})\b", value)
            if match:
                return match.group(1)
    return ""


def _fintoolbench_first_numeric(record: dict[str, Any], keys: tuple[str, ...]) -> float | None:
    for key in keys:
        value = record.get(key)
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(",", "").replace("$", "").strip()
            try:
                return float(cleaned)
            except ValueError:
                continue
    return None


def _fintoolbench_first_present_key(record: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        if key in record:
            return key
    return ""


def _execute_finance_agent_v2_tool(name: str, arguments: dict[str, Any]) -> str:
    status = _finance_agent_v2_backend_status(Path.cwd())
    if not status["ready"]:
        return _json_tool_result(
            name,
            arguments,
            {
                "tool": name,
                "mode": "deterministic_fixture_finance_agent_v2_tools",
                "fixture_valid": False,
                "error": status.get("reason", "Finance Agent v2 fixture backend is unavailable"),
                "backend_canary": status,
            },
        )
    store = Path(str(status["fixture_root"]))
    fixture = _finance_agent_v2_fixture_payload(store)
    result: dict[str, Any] = {
        "tool": name,
        "mode": "deterministic_fixture_finance_agent_v2_tools",
        "fixture_valid": True,
        "fixture_root": str(store),
        "canary_passed": True,
    }
    query = " ".join(_flatten_string_values(arguments))
    if name == "web_search":
        result["results"] = _finance_agent_v2_matching_records(fixture.get("search_results"), query, max_results=5)
        return _json_tool_result(name, arguments, result)
    if name == "edgar_search":
        result["results"] = _finance_agent_v2_matching_records(fixture.get("edgar_filings"), query or "CRWD 2025", max_results=5)
        return _json_tool_result(name, arguments, result)
    if name == "parse_html_page":
        document = _finance_agent_v2_document_for_argument(fixture.get("documents", []), arguments)
        result["document"] = {
            key: document.get(key)
            for key in ("form", "report_date", "accession", "url", "html_path", "parsed_path")
        } if isinstance(document, dict) else {}
        result["text"] = str(document.get("parsed_text") or "")[:12000] if isinstance(document, dict) else ""
        return _json_tool_result(name, arguments, result)
    if name == "retrieve_information":
        result["results"] = _finance_agent_v2_matching_records(fixture.get("retrieval_index"), query, max_results=8)
        return _json_tool_result(name, arguments, result)
    if name == "price_history":
        symbol = str(arguments.get("symbol") or arguments.get("ticker") or "CRWD").upper()
        result["symbol"] = symbol
        result["history"] = fixture.get("price_history", {}).get(symbol, [])
        return _json_tool_result(name, arguments, result)
    result["error"] = f"Unknown Finance Agent v2 fixture tool: {name}"
    return _json_tool_result(name, arguments, result)


def _execute_fintoolbench_tool(name: str, arguments: dict[str, Any]) -> str:
    fixture = _load_fintoolbench_fixture(name, arguments, Path.cwd())
    if fixture.get("error"):
        return _json_tool_result(
            name,
            arguments,
            {
                "tool": name,
                "mode": "deterministic_fixture_finance_backend",
                "fixture_valid": False,
                "error": fixture["error"],
                "searched_paths": fixture.get("searched_paths", []),
            },
        )
    validation = _validate_fintoolbench_fixture(name, fixture.get("payload"), arguments)
    if validation.get("passed") is not True:
        return _json_tool_result(
            name,
            arguments,
            {
                "tool": name,
                "mode": "deterministic_fixture_finance_backend",
                "fixture_valid": False,
                "error": validation.get("reason", "fixture failed semantic validation"),
                "fixture_path": fixture.get("fixture_path", ""),
            },
        )
    return _json_tool_result(
        name,
        arguments,
        {
            "tool": name,
            "mode": "deterministic_fixture_finance_backend",
            "fixture_valid": True,
            "fixture_path": fixture.get("fixture_path", ""),
            "records": _fintoolbench_fixture_records(fixture.get("payload")),
            "canary": validation,
        },
    )


def _json_tool_result(name: str, arguments: dict[str, Any], payload: dict[str, Any]) -> str:
    return json.dumps(
        {
            "tool": name,
            "arguments": arguments,
            "backend": "agent_bench_local_deterministic",
            "result": payload,
        },
        ensure_ascii=False,
        sort_keys=True,
    )


def _first_argument_path(arguments: dict[str, Any]) -> str:
    for key in ("path", "file", "url"):
        value = arguments.get(key)
        if isinstance(value, str) and value.strip() and not value.startswith(("http://", "https://")):
            return value.strip()
    return ""


def _safe_read_local_text(path_value: str, max_chars: int) -> str:
    path = _safe_workspace_path(path_value)
    if not path.is_file() or _is_binary_like_file(path):
        return ""
    return path.read_text(encoding="utf-8", errors="replace")[:max_chars]


def _local_price_history(arguments: dict[str, Any]) -> dict[str, Any]:
    symbol = str(arguments.get("symbol") or arguments.get("ticker") or arguments.get("query") or "").upper()
    matches: list[dict[str, Any]] = []
    for path in sorted(Path.cwd().rglob("*.json"))[:500]:
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        if symbol and symbol not in path.name.upper():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except Exception:
            continue
        matches.append({"path": str(path.relative_to(Path.cwd())), "data": _compact_json(payload, 2500)})
        if len(matches) >= 3:
            break
    return {"symbol": symbol, "matches": matches}


def _local_retrieval_snippets(arguments: dict[str, Any], *, max_results: int) -> list[dict[str, str]]:
    query = " ".join(_flatten_string_values(arguments))[:200]
    terms = [term.lower() for term in re.findall(r"[A-Za-z0-9.$_-]{3,}", query)[:8]]
    snippets: list[dict[str, str]] = []
    if not terms:
        return snippets
    for path in sorted(Path.cwd().rglob("*")):
        if len(snippets) >= max_results:
            break
        if not path.is_file() or any(part in SKIP_DIRS for part in path.parts) or _is_binary_like_file(path):
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        lowered = text.lower()
        if not any(term in lowered for term in terms):
            continue
        index = min((lowered.find(term) for term in terms if term in lowered), default=0)
        start = max(0, index - 240)
        end = min(len(text), index + 760)
        snippets.append(
            {
                "path": str(path.relative_to(Path.cwd())),
                "snippet": text[start:end],
            }
        )
    return snippets


def _compact_json(value: Any, max_chars: int) -> str:
    return truncate(json.dumps(value, ensure_ascii=False, sort_keys=True), max_chars)


def _canonical_tool_name(name: str) -> str:
    aliases = {
        "cat": "read_file",
        "dir": "list_files",
        "list_dir": "list_files",
        "list_directory": "list_files",
        "ls": "list_files",
    }
    return aliases.get(name, name)


def _tool_schema_names(tools: list[dict[str, Any]]) -> list[str]:
    names: set[str] = set()
    for tool in tools:
        name = _tool_schema_name(tool)
        if name:
            names.add(name)
    return sorted(names)


def _tool_schema_name(tool: dict[str, Any]) -> str:
    if not isinstance(tool, dict):
        return ""
    function = tool.get("function")
    if isinstance(function, dict) and isinstance(function.get("name"), str):
        return function["name"]
    if isinstance(tool.get("name"), str):
        return tool["name"]
    return ""


def _item_required_tools(item: BenchmarkItem) -> list[str]:
    tools: list[str] = _descriptor_required_tools()
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    for key in (
        "required_tools",
        "select_tools",
        "selected_tools",
        "selected_tool",
        "tool_name",
        "tool_names",
        "gold_tools",
        "api_name",
        "api_names",
    ):
        tools.extend(_flatten_string_values(metadata.get(key)))
    nested = metadata.get("tools")
    if isinstance(nested, list):
        for entry in nested:
            if isinstance(entry, str):
                tools.append(entry)
            elif isinstance(entry, dict):
                for key in ("name", "tool", "tool_name", "api_name"):
                    value = entry.get(key)
                    if isinstance(value, str) and value.strip():
                        tools.append(value)
    elif isinstance(nested, dict):
        for key in ("name", "tool", "tool_name", "api_name"):
            value = nested.get(key)
            if isinstance(value, str) and value.strip():
                tools.append(value)
    return sorted({tool.strip() for tool in tools if tool.strip()})


def _missing_required_tools(item: BenchmarkItem, tools: list[dict[str, Any]]) -> list[str]:
    metadata = item.metadata if isinstance(item.metadata, dict) else {}
    if metadata.get("live_tools_required") is False and _item_metadata_may_disable_live_tools(item):
        return []
    exposed = set(_tool_schema_names(tools))
    required = set(_item_required_tools(item))
    if not required:
        return []
    return sorted(required - exposed)


def _item_metadata_may_disable_live_tools(item: BenchmarkItem) -> bool:
    if "tool_call" not in _item_required_capabilities(item):
        return True
    return _static_conversion_live_tools_disabled()


def tool_list_files(arguments: dict[str, Any]) -> str:
    root = Path.cwd()
    start = _safe_workspace_path(str(arguments.get("path") or "."))
    pattern = str(arguments.get("pattern") or "")
    max_results = _bounded_int(arguments.get("max_results"), 80, 1, 200)
    if not start.exists():
        return f"Path does not exist: {start.relative_to(root)}"
    paths = []
    iterator = start.rglob("*") if start.is_dir() else [start]
    for path in iterator:
        if len(paths) >= max_results:
            break
        if not path.is_file() or any(part in SKIP_DIRS for part in path.parts):
            continue
        relative = str(path.relative_to(root))
        if pattern and pattern not in relative and not relative.endswith(pattern):
            continue
        paths.append(relative)
    return "\n".join(paths) if paths else "No files matched."


def tool_read_file(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    if not path.is_file():
        return f"File not found: {arguments.get('path')}"
    if _is_binary_like_file(path) and not bool(arguments.get("allow_binary")):
        return f"Refusing to read binary/archive file as text: {path.relative_to(Path.cwd())}"
    start_line = _bounded_int(arguments.get("start_line"), 1, 1, 1_000_000)
    max_lines = _bounded_int(arguments.get("max_lines"), 160, 1, 400)
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    selected = lines[start_line - 1 : start_line - 1 + max_lines]
    rendered = "\n".join(f"{index}: {line}" for index, line in enumerate(selected, start=start_line))
    return truncate(rendered, TOOL_RESULT_CHARS)


def tool_search_files(arguments: dict[str, Any]) -> str:
    query = str(arguments.get("query") or "")
    if not query:
        return "Missing query."
    start = _safe_workspace_path(str(arguments.get("path") or "."))
    max_results = _bounded_int(arguments.get("max_results"), 50, 1, 100)
    matches: list[str] = []
    iterator = start.rglob("*") if start.is_dir() else [start]
    for path in iterator:
        if len(matches) >= max_results:
            break
        if not path.is_file() or any(part in SKIP_DIRS for part in path.parts):
            continue
        if _is_binary_like_file(path):
            continue
        try:
            for line_number, line in enumerate(path.read_text(encoding="utf-8", errors="replace").splitlines(), 1):
                if query in line:
                    matches.append(f"{path.relative_to(Path.cwd())}:{line_number}: {line[:240]}")
                    if len(matches) >= max_results:
                        break
        except OSError:
            continue
    return "\n".join(matches) if matches else "No matches."


def _is_binary_like_file(path: Path) -> bool:
    if path.suffix.lower() in BINARY_SUFFIXES:
        return True
    try:
        sample = path.read_bytes()[:2048]
    except OSError:
        return True
    return b"\0" in sample


def tool_write_file(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    content = str(arguments.get("content") or "")
    if path.exists() and path.is_dir():
        path = path / "response.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return f"Wrote {len(content.encode('utf-8'))} bytes to {path.relative_to(Path.cwd())}."


def tool_write_base64_file(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    raw = str(arguments.get("base64_content") or "")
    try:
        data = base64.b64decode(raw, validate=True)
    except Exception as exc:
        return f"Invalid base64_content: {exc}"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return f"Wrote {len(data)} binary bytes to {path.relative_to(Path.cwd())}."


def tool_read_binary_metadata(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    if not path.is_file():
        return f"File not found: {arguments.get('path')}"
    metadata = _binary_metadata(path)
    return json.dumps(metadata, ensure_ascii=False, sort_keys=True)


def tool_extract_archive(arguments: dict[str, Any]) -> str:
    archive = _safe_workspace_path(str(arguments.get("path") or ""))
    output_dir = _safe_workspace_path(str(arguments.get("output_dir") or ""))
    if not archive.is_file():
        return f"Archive not found: {arguments.get('path')}"
    if archive.suffix.lower() != ".zip":
        return "Only .zip archives are supported by extract_archive."
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(archive) as handle:
            bad_member = handle.testzip()
            if bad_member:
                return f"Archive integrity check failed at member: {bad_member}"
            handle.extractall(output_dir)
    except zipfile.BadZipFile as exc:
        return f"Invalid zip archive: {exc}"
    return f"Extracted {archive.relative_to(Path.cwd())} to {output_dir.relative_to(Path.cwd())}."


def tool_read_spreadsheet(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    if not path.is_file():
        return f"File not found: {arguments.get('path')}"
    max_rows = _bounded_int(arguments.get("max_rows"), 20, 1, 50)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            rows = list(csv.reader(handle))[:max_rows]
        return _render_table_preview(rows)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError:
            try:
                return _render_table_preview(_read_xlsx_preview_zip(path, max_rows))
            except Exception as exc:
                return f"Unable to read spreadsheet: {exc}"
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return f"Unable to read spreadsheet: {exc}"
        sheet_name = str(arguments.get("sheet") or "")
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = []
        for row in worksheet.iter_rows(max_row=max_rows, values_only=True):
            rows.append(["" if value is None else str(value) for value in row])
        return _render_table_preview(rows)
    return f"Unsupported spreadsheet type: {path.suffix}"


def _read_xlsx_preview_zip(path: Path, max_rows: int) -> list[list[str]]:
    if not zipfile.is_zipfile(path):
        raise ValueError("XLSX file is not a valid ZIP container")
    with zipfile.ZipFile(path) as handle:
        shared_strings = _xlsx_shared_strings(handle)
        workbook = _xlsx_first_sheet_path(handle)
        if workbook is None:
            raise ValueError("XLSX workbook does not contain a worksheet")
        root = ET.fromstring(handle.read(workbook))
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list[str]] = []
    for row_node in root.findall(".//x:sheetData/x:row", namespace):
        row_values: list[str] = []
        for cell in row_node.findall("x:c", namespace):
            value_node = cell.find("x:v", namespace)
            inline_node = cell.find("x:is/x:t", namespace)
            if inline_node is not None and inline_node.text is not None:
                row_values.append(inline_node.text)
                continue
            raw = "" if value_node is None or value_node.text is None else value_node.text
            if cell.attrib.get("t") == "s":
                try:
                    row_values.append(shared_strings[int(raw)])
                except (ValueError, IndexError):
                    row_values.append(raw)
            else:
                row_values.append(raw)
        rows.append(row_values)
        if len(rows) >= max_rows:
            break
    return rows


def _xlsx_shared_strings(handle: zipfile.ZipFile) -> list[str]:
    try:
        raw = handle.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(raw)
    values: list[str] = []
    for item in root.findall(".//x:si", namespace):
        parts = [node.text or "" for node in item.findall(".//x:t", namespace)]
        values.append("".join(parts))
    return values


def _xlsx_first_sheet_path(handle: zipfile.ZipFile) -> str | None:
    for name in sorted(handle.namelist()):
        if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
            return name
    return None


def tool_read_pdf_text(arguments: dict[str, Any]) -> str:
    path = _safe_workspace_path(str(arguments.get("path") or ""))
    if not path.is_file():
        return f"File not found: {arguments.get('path')}"
    max_chars = _bounded_int(arguments.get("max_chars"), 4000, 100, 12000)
    try:
        import pypdf
    except ImportError:
        pypdf = None
    if pypdf is not None:
        try:
            reader = pypdf.PdfReader(str(path))
            text = "\n".join(page.extract_text() or "" for page in reader.pages[:10])
            return truncate(text, max_chars) if text.strip() else "No extractable PDF text found."
        except Exception:
            pass
    pdftotext = shutil.which("pdftotext")
    if pdftotext:
        completed = subprocess.run(
            [pdftotext, "-layout", str(path), "-"],
            text=True,
            capture_output=True,
            cwd=Path.cwd(),
            timeout=30,
            check=False,
        )
        if completed.returncode == 0:
            return truncate(completed.stdout, max_chars)
        return f"pdftotext failed: {completed.stderr[-500:]}"
    fallback_text = _fallback_pdf_text(path)
    if fallback_text.strip():
        return truncate(fallback_text, max_chars)
    return "Unable to read PDF: no text extractor is installed and fallback extraction found no text."


def _fallback_pdf_text(path: Path) -> str:
    try:
        raw = path.read_bytes()
    except OSError:
        return ""
    if not raw.startswith(b"%PDF-"):
        return ""
    chunks: list[str] = []
    for match in re.finditer(rb"\((.*?)\)", raw, flags=re.DOTALL):
        value = match.group(1)
        value = value.replace(rb"\(", b"(").replace(rb"\)", b")").replace(rb"\\n", b"\n")
        try:
            text = value.decode("utf-8")
        except UnicodeDecodeError:
            text = value.decode("latin-1", errors="replace")
        if text.strip():
            chunks.append(text)
    return "\n".join(chunks)


def tool_list_artifacts(arguments: dict[str, Any]) -> str:
    start = _safe_workspace_path(str(arguments.get("path") or "agent_bench_outputs"))
    max_results = _bounded_int(arguments.get("max_results"), 80, 1, 200)
    if not start.exists():
        return f"Path does not exist: {start.relative_to(Path.cwd())}"
    rows = []
    iterator = start.rglob("*") if start.is_dir() else [start]
    for path in iterator:
        if len(rows) >= max_results:
            break
        if not path.is_file():
            continue
        try:
            size = path.stat().st_size
        except OSError:
            size = 0
        rows.append(f"{path.relative_to(Path.cwd())}\t{size} bytes")
    return "\n".join(rows) if rows else "No artifacts found."


def _render_table_preview(rows: list[list[Any]]) -> str:
    if not rows:
        return "No rows found."
    return "\n".join("\t".join(str(value) for value in row) for row in rows)


def _binary_metadata(path: Path) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "path": str(path.relative_to(Path.cwd())),
        "suffix": path.suffix.lower(),
        "size_bytes": path.stat().st_size,
        "is_zip": zipfile.is_zipfile(path),
    }
    if path.suffix.lower() in {".zip", ".xlsx", ".xlsm", ".docx"} and zipfile.is_zipfile(path):
        try:
            with zipfile.ZipFile(path) as handle:
                metadata["zip_member_count"] = len(handle.namelist())
                metadata["zip_test_ok"] = handle.testzip() is None
        except zipfile.BadZipFile:
            metadata["zip_test_ok"] = False
    if path.suffix.lower() == ".pdf":
        try:
            metadata["pdf_header"] = path.read_bytes()[:5] == b"%PDF-"
        except OSError:
            metadata["pdf_header"] = False
    return metadata


def tool_apply_patch(arguments: dict[str, Any]) -> str:
    patch = str(arguments.get("patch") or "")
    if not patch.strip():
        return "Missing patch."
    completed = subprocess.run(
        ["patch", "-p1"],
        input=patch,
        text=True,
        capture_output=True,
        cwd=Path.cwd(),
        timeout=30,
        check=False,
    )
    output = f"exit_code={completed.returncode}\n{completed.stdout}{completed.stderr}"
    return truncate(output, TOOL_RESULT_CHARS)


def tool_run_command(arguments: dict[str, Any]) -> str:
    argv = arguments.get("argv")
    if isinstance(argv, list) and all(isinstance(item, str) and item for item in argv):
        args = argv
    else:
        if arguments.get("command"):
            return "Invalid run_command arguments: legacy command field is disabled; use argv."
        return "Missing argv."
    if not args:
        return "Missing argv."
    timeout_seconds = _bounded_int(arguments.get("timeout_seconds"), 30, 1, 120)
    stdin = arguments.get("stdin")
    if stdin is not None and not isinstance(stdin, str):
        return "stdin must be a string."
    completed = subprocess.run(
        args,
        input=stdin,
        text=True,
        capture_output=True,
        cwd=Path.cwd(),
        timeout=timeout_seconds,
        check=False,
    )
    output = f"exit_code={completed.returncode}\nSTDOUT:\n{completed.stdout}\nSTDERR:\n{completed.stderr}"
    return truncate(output, TOOL_RESULT_CHARS)


def _contains_shell_syntax(command: str) -> bool:
    return bool(re.search(r"(^|\s)(?:&&|\|\||[|<>;`])|<<|\$\(", command))


def _safe_workspace_path(raw_path: str) -> Path:
    root = Path.cwd().resolve()
    if not raw_path:
        raise ValueError("path is required")
    candidate = (root / raw_path).resolve()
    if candidate != root and root not in candidate.parents:
        raise ValueError(f"path escapes repository root: {raw_path}")
    return candidate


@contextmanager
def pushd(path: Path):
    previous = Path.cwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(previous)


@contextmanager
def wall_clock_timeout(timeout_seconds: float):
    if not hasattr(signal, "SIGALRM"):
        yield
        return

    def _handle_timeout(signum: int, frame: Any) -> None:
        raise ChatCompletionTimeoutError(timeout_seconds)

    previous_handler = signal.getsignal(signal.SIGALRM)
    signal.signal(signal.SIGALRM, _handle_timeout)
    previous_timer = signal.setitimer(signal.ITIMER_REAL, timeout_seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous_handler)
        if previous_timer[0] > 0:
            signal.setitimer(signal.ITIMER_REAL, previous_timer[0], previous_timer[1])


def _agent_run_from_dict(result: dict[str, Any]) -> AgentRun:
    return AgentRun(
        answer=str(result.get("answer", "")),
        content=str(result.get("content", "")),
        usage=result.get("usage", {}) if isinstance(result.get("usage"), dict) else {},
        tool_trace=result.get("tool_trace", []) if isinstance(result.get("tool_trace"), list) else [],
        diagnostics=result.get("diagnostics", {}) if isinstance(result.get("diagnostics"), dict) else {},
        status_code=int(result.get("status_code", 200)) if isinstance(result.get("status_code", 200), int) else 200,
    )


def _isolated_workspace_root(item: BenchmarkItem) -> Path:
    root = _item_output_dir(item) / "workspace"
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True, exist_ok=True)
    return root


def _write_model_visible_task_files(root: Path, item: BenchmarkItem) -> None:
    lines = [
        "Task",
        "====",
        "",
        _model_visible_question(item).strip(),
    ]
    if item.choices:
        lines.extend(["", "Choices"])
        lines.extend(f"{label}. {text}" for label, text in sorted(item.choices.items()))
    lines.extend(
        [
            "",
            "Final answer format",
            "Return the final answer as compact JSON when finished.",
        ]
    )
    (root / "TASK.md").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def _model_visible_question(item: BenchmarkItem) -> str:
    question = item.question.strip()
    visible_context = item.metadata.get("visible_context")
    if not isinstance(visible_context, str) or not visible_context.strip():
        return question
    return f"{question}\n\nBenchmark record context:\n{visible_context.strip()}"


def _expose_input_files_at_workspace_root(root: Path, relative_paths: list[str]) -> None:
    by_name: dict[str, str] = {}
    collisions: set[str] = set()
    for relative_path in relative_paths:
        name = Path(relative_path).name
        if not name:
            continue
        if name in by_name and by_name[name] != relative_path:
            collisions.add(name)
        else:
            by_name[name] = relative_path
    for name in collisions:
        by_name.pop(name, None)
    for name, relative_path in by_name.items():
        source = root / relative_path
        target = root / name
        if target.exists() or target.is_symlink() or not source.is_file():
            continue
        try:
            target.symlink_to(source.relative_to(target.parent))
        except OSError:
            shutil.copy2(source, target)


def _write_task_file_manifest(root: Path, relative_paths: list[str]) -> None:
    rows = []
    for relative_path in sorted(relative_paths):
        path = root / relative_path
        if not path.is_file():
            continue
        rows.append(
            {
                "display_filename": path.name,
                "tool_path": relative_path,
                "size_bytes": path.stat().st_size,
            }
        )
    manifest_dir = root / "agent_bench_task_inputs"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    (manifest_dir / "manifest.json").write_text(
        json.dumps({"files": rows}, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    table = ["| Display filename | Tool path | Size |", "|---|---|---:|"]
    table.extend(
        f"| {row['display_filename']} | {row['tool_path']} | {row['size_bytes']} |"
        for row in rows
    )
    (root / "TASK_FILES.md").write_text("\n".join(table) + "\n", encoding="utf-8")


def _item_to_dict(item: BenchmarkItem) -> dict[str, Any]:
    return {
        "id": _item_id(item),
        "question": item.question,
        "expected": item.expected,
        "source": item.source,
        "choices": item.choices,
        "metadata": item.metadata,
    }


def _workspace_to_dict(workspace: TaskWorkspace) -> dict[str, Any]:
    return {
        "root": str(workspace.root),
        "output_dir": str(workspace.output_dir),
        "manifest": workspace.manifest,
        "metadata": workspace.metadata,
    }


def _agent_run_to_dict(run: AgentRun) -> dict[str, Any]:
    return {
        "answer": run.answer,
        "content": run.content,
        "usage": run.usage,
        "tool_trace": run.tool_trace,
        "diagnostics": run.diagnostics,
        "status_code": run.status_code,
    }


def _output_bundle_to_dict(outputs: OutputBundle) -> dict[str, Any]:
    return {
        "answer": outputs.answer,
        "patch": outputs.patch,
        "artifact_paths": outputs.artifact_paths,
        "tool_trace": outputs.tool_trace,
        "metadata": outputs.metadata,
    }


def _write_item_json(item_dir: Path, name: str, payload: dict[str, Any]) -> None:
    item_dir.mkdir(parents=True, exist_ok=True)
    (item_dir / name).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _item_output_dir(item: BenchmarkItem) -> Path:
    output_root = Path(os.environ.get("AGENT_BENCH_OUTPUT_DIR") or (Path.cwd() / "agent_bench_outputs"))
    slug = _safe_slug(item.source)[:80]
    output_dir = output_root / "items" / f"{slug}__sha256-{_item_hash(item)[:16]}"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _item_id(item: BenchmarkItem) -> str:
    explicit = item.metadata.get("item_id")
    if isinstance(explicit, str) and explicit.strip():
        return explicit.strip()
    return f"{_safe_slug(item.source)}__sha256-{_item_hash(item)[:16]}"


def _item_hash(item: BenchmarkItem) -> str:
    digest = hashlib.sha256()
    payload = {
        "benchmark": os.environ.get("AGENT_BENCH_BENCHMARK_NAME", ""),
        "source": item.source,
        "question": item.question,
        "expected": item.expected,
        "choices": item.choices,
    }
    digest.update(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8"))
    return digest.hexdigest()


def _workspace_manifest(root: Path, limit: int = 200) -> list[str]:
    try:
        paths = [
            str(path.relative_to(root))
            for path in root.rglob("*")
            if path.is_file() and not any(part in SKIP_DIRS for part in path.parts)
        ]
    except OSError:
        return []
    return sorted(paths)[:limit]


def _artifact_asset_paths(item: BenchmarkItem, root: Path) -> list[Path]:
    candidates: list[str] = []
    for key in (
        "input_assets",
        "input_files",
        "asset_paths",
        "attachments",
        "files",
        "reference_files",
        "documents",
        "dataset_files",
    ):
        candidates.extend(_flatten_string_values(item.metadata.get(key)))
    paths: list[Path] = []
    root = root.resolve()
    for raw in candidates:
        candidate = (root / raw).resolve()
        if candidate == root or root not in candidate.parents:
            continue
        if candidate.exists():
            paths.append(candidate)
    paths.extend(_inferred_artifact_assets(item, root))
    paths.extend(_cached_artifact_assets(item))
    return sorted(set(paths))


def _inferred_artifact_assets(item: BenchmarkItem, root: Path) -> list[Path]:
    source_path = item.source.split(":", 1)[0]
    inferred: list[Path] = []
    candidate_source = (root / source_path).resolve()
    if candidate_source.exists():
        parent = candidate_source.parent if candidate_source.is_file() else candidate_source
        for name in (
            "assets",
            "input",
            "inputs",
            "data",
            "reference_files",
            "deliverable_files",
            "paper.pdf",
            "paper.md",
            "addendum.md",
            "config.yaml",
        ):
            candidate = parent / name
            if candidate.exists():
                inferred.append(candidate)
    for name in ("reference_files", "deliverable_files"):
        candidate = root / name
        if candidate.exists():
            inferred.append(candidate)
    return inferred


def _cached_artifact_assets(item: BenchmarkItem) -> list[Path]:
    cache = _asset_root()
    if not cache.exists():
        return []
    slug = _safe_slug(item.source)
    benchmark = _safe_slug(os.environ.get("AGENT_BENCH_BENCHMARK_NAME", "benchmark"))
    candidates = [cache / benchmark / slug, cache / benchmark, cache / slug]
    for path in candidates:
        if path.exists():
            return [path]
    return []


def _asset_root() -> Path:
    return Path(os.environ.get("AGENT_BENCH_ASSET_ROOT", DEFAULT_ASSET_ROOT)).expanduser()


def _allow_asset_download() -> bool:
    return os.environ.get("AGENT_BENCH_ALLOW_ASSET_DOWNLOAD", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def _materialize_artifact_assets(asset_paths: list[Path], root: Path, item: BenchmarkItem) -> list[str]:
    input_root = root / "agent_bench_task_inputs" / _safe_slug(item.source)
    if input_root.exists():
        shutil.rmtree(input_root)
    input_root.mkdir(parents=True, exist_ok=True)
    materialized: list[str] = []
    for source in asset_paths:
        if source.is_dir():
            for child in sorted(source.iterdir()):
                target = input_root / child.name
                if child.is_dir():
                    shutil.copytree(child, target, dirs_exist_ok=True)
                elif child.is_file():
                    shutil.copy2(child, target)
            for path in sorted(input_root.rglob("*")):
                if path.is_file():
                    materialized.append(str(path.relative_to(root)))
        elif source.is_file():
            target = input_root / source.name
            shutil.copy2(source, target)
            materialized.append(str(target.relative_to(root)))
    return sorted(set(materialized))


def _collect_artifacts_to_output(source_dir: Path, artifact_dir: Path) -> list[str]:
    artifact_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    if not source_dir.exists():
        return copied
    if source_dir.is_file():
        target = artifact_dir / source_dir.name
        shutil.copy2(source_dir, target)
        return [str(target)]
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file():
            continue
        relative = path.relative_to(source_dir)
        target = artifact_dir / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, target)
        copied.append(str(target))
    return copied


def _artifact_integrity_errors(
    artifact_paths: list[str],
    *,
    allowed_root: Path | None = None,
) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    resolved_allowed_root: Path | None = None
    if allowed_root is not None:
        try:
            resolved_allowed_root = allowed_root.resolve()
        except OSError:
            resolved_allowed_root = allowed_root
    for raw_path in artifact_paths:
        path = Path(raw_path)
        suffix = path.suffix.lower()
        if resolved_allowed_root is not None:
            try:
                resolved_path = path.resolve()
            except OSError:
                resolved_path = path
            if resolved_path != resolved_allowed_root and resolved_allowed_root not in resolved_path.parents:
                errors.append({"path": raw_path, "error": "path outside allowed output directory"})
                continue
        if not path.exists():
            errors.append({"path": raw_path, "error": "missing file"})
            continue
        if path.is_file() and path.stat().st_size == 0:
            errors.append({"path": raw_path, "error": "empty file"})
            continue
        if suffix in {".patch", ".diff"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            if not _looks_like_unified_diff(text):
                errors.append({"path": raw_path, "error": "invalid patch format"})
            continue
        if suffix in {".xlsx", ".xlsm", ".docx", ".zip"}:
            if not zipfile.is_zipfile(path):
                errors.append({"path": raw_path, "error": f"{suffix} file is not a valid ZIP container"})
                continue
            try:
                with zipfile.ZipFile(path) as handle:
                    bad_member = handle.testzip()
                    if bad_member:
                        errors.append({"path": raw_path, "error": f"ZIP member failed CRC: {bad_member}"})
            except zipfile.BadZipFile as exc:
                errors.append({"path": raw_path, "error": str(exc)})
        elif suffix == ".pdf":
            try:
                if path.read_bytes()[:5] != b"%PDF-":
                    errors.append({"path": raw_path, "error": "PDF header is missing"})
            except OSError as exc:
                errors.append({"path": raw_path, "error": str(exc)})
    return errors


def _looks_like_unified_diff(text: str) -> bool:
    return bool(
        text.strip()
        and (
            re.search(r"^diff --git a/.+ b/.+", text, flags=re.MULTILINE)
            or (
                re.search(r"^---\s+", text, flags=re.MULTILINE)
                and re.search(r"^\+\+\+\s+", text, flags=re.MULTILINE)
                and re.search(r"^@@\s", text, flags=re.MULTILINE)
            )
        )
    )


def _artifact_previews(artifact_paths: list[str]) -> str:
    previews: list[str] = []
    for raw_path in artifact_paths[:5]:
        path = Path(raw_path)
        header = f"## {path.name}"
        if not path.is_file():
            previews.append(f"{header}\nmissing")
            continue
        if _is_binary_like_file(path):
            previews.append(f"{header}\n{json.dumps(_binary_metadata(path), ensure_ascii=False, sort_keys=True)}")
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            previews.append(f"{header}\nUnable to read artifact: {exc}")
            continue
        previews.append(f"{header}\n{truncate(text, 2500)}")
    return "\n\n".join(previews) if previews else "(none)"


def _file_artifact_canary() -> dict[str, Any]:
    canary_root = Path(os.environ.get("AGENT_BENCH_OUTPUT_DIR", "/outputs")) / "canaries" / "file_artifact"
    if canary_root.exists():
        shutil.rmtree(canary_root)
    canary_root.mkdir(parents=True, exist_ok=True)
    try:
        _write_minimal_xlsx(canary_root / "input.xlsx", [["label", "value"], ["canary", "ok"]])
        _write_minimal_pdf(canary_root / "input.pdf", "agent bench pdf canary")
        with pushd(canary_root):
            spreadsheet = tool_read_spreadsheet({"path": "input.xlsx", "max_rows": 5})
            if _tool_result_failed("read_spreadsheet", spreadsheet) or "canary" not in spreadsheet:
                return {"passed": False, "reason": "read_spreadsheet canary failed", "result": spreadsheet[:500]}
            pdf = tool_read_pdf_text({"path": "input.pdf", "max_chars": 1000})
            if _tool_result_failed("read_pdf_text", pdf) or "agent bench pdf canary" not in pdf:
                return {"passed": False, "reason": "read_pdf_text canary failed", "result": pdf[:500]}
            output_dir = Path("agent_bench_outputs") / "canary"
            output_dir.mkdir(parents=True, exist_ok=True)
            _write_minimal_xlsx(output_dir / "report.xlsx", [["status"], ["ok"]])
            (output_dir / "notes.txt").write_text("artifact canary ok\n", encoding="utf-8")
            listing = tool_list_artifacts({"path": str(output_dir), "max_results": 20})
            if _tool_result_failed("list_artifacts", listing) or "report.xlsx" not in listing or "notes.txt" not in listing:
                return {"passed": False, "reason": "list_artifacts canary failed", "result": listing[:500]}
            collected = _collect_artifacts_to_output(output_dir, canary_root / "collected")
        if len(collected) < 2:
            return {"passed": False, "reason": "artifact collection canary did not collect expected files"}
        integrity_errors = _artifact_integrity_errors(collected)
        if integrity_errors:
            return {
                "passed": False,
                "reason": "artifact integrity canary failed",
                "artifact_errors": integrity_errors,
            }
        return {"passed": True, "root": str(canary_root), "collected_count": len(collected)}
    except Exception as exc:
        return {"passed": False, "reason": f"file_artifact canary failed: {exc}"}


def _repo_patch_canary() -> dict[str, Any]:
    canary_root = Path(os.environ.get("AGENT_BENCH_OUTPUT_DIR", "/outputs")) / "canaries" / "repo_patch"
    if canary_root.exists():
        shutil.rmtree(canary_root)
    repo = canary_root / "repo"
    patch_output_dir = canary_root / "patch"
    repo.mkdir(parents=True, exist_ok=True)
    patch_output_dir.mkdir(parents=True, exist_ok=True)
    try:
        patch_binary, patch_error = _verify_patch_binary()
        if patch_error:
            return {
                "passed": False,
                "reason": patch_error,
                "blocker_type": "repo_patch_harness_setup",
                "patch_output_dir": str(patch_output_dir),
            }
        checks: list[dict[str, Any]] = []

        def run(args: list[str], **kwargs: Any) -> subprocess.CompletedProcess[str]:
            completed = subprocess.run(
                args,
                text=True,
                capture_output=True,
                cwd=repo,
                timeout=30,
                check=False,
                **kwargs,
            )
            checks.append({"argv": args, "exit_code": completed.returncode})
            return completed

        for command in (
            ["git", "init"],
            ["git", "config", "user.email", "agent-bench@example.invalid"],
            ["git", "config", "user.name", "Agent Bench"],
        ):
            completed = run(command)
            if completed.returncode != 0:
                return {"passed": False, "reason": "repo_patch canary git init failed", "stderr": completed.stderr}
        (repo / "canary.txt").write_text("before\n", encoding="utf-8")
        for command in (["git", "add", "canary.txt"], ["git", "commit", "-m", "base"]):
            completed = run(command)
            if completed.returncode != 0:
                return {"passed": False, "reason": "repo_patch canary base commit failed", "stderr": completed.stderr}
        base = run(["git", "rev-parse", "HEAD"])
        if base.returncode != 0 or not base.stdout.strip():
            return {"passed": False, "reason": "repo_patch canary could not identify base commit"}
        checkout = run(["git", "checkout", "--detach", base.stdout.strip()])
        if checkout.returncode != 0:
            return {"passed": False, "reason": "repo_patch canary checkout failed", "stderr": checkout.stderr}
        patch = (
            "--- a/canary.txt\n"
            "+++ b/canary.txt\n"
            "@@ -1 +1 @@\n"
            "-before\n"
            "+after\n"
        )
        apply = subprocess.run(
            [patch_binary, "-p1"],
            input=patch,
            text=True,
            capture_output=True,
            cwd=repo,
            timeout=30,
            check=False,
        )
        checks.append({"argv": [patch_binary, "-p1"], "exit_code": apply.returncode})
        if apply.returncode != 0:
            return {"passed": False, "reason": "repo_patch canary patch apply failed", "stderr": apply.stderr}
        (patch_output_dir / "model.patch").write_text(patch, encoding="utf-8")
        if not (patch_output_dir / "model.patch").is_file():
            return {
                "passed": False,
                "reason": "repo_patch canary could not write patch artifact",
                "blocker_type": "repo_patch_harness_setup",
                "patch_output_dir": str(patch_output_dir),
            }
        diff = run(["git", "diff", "--binary"])
        if diff.returncode != 0 or "+after" not in diff.stdout:
            return {"passed": False, "reason": "repo_patch canary diff collection failed", "stderr": diff.stderr}
        test = run(["python", "-c", "from pathlib import Path; assert Path('canary.txt').read_text() == 'after\\n'"])
        if test.returncode != 0:
            return {"passed": False, "reason": "repo_patch canary trivial test failed", "stderr": test.stderr}
        return {
            "passed": True,
            "root": str(repo),
            "base_commit": base.stdout.strip(),
            "patch_binary": patch_binary,
            "patch_output_dir": str(patch_output_dir),
            "checks": checks,
        }
    except Exception as exc:
        return {"passed": False, "reason": f"repo_patch canary failed: {exc}"}


def _verify_patch_binary() -> tuple[str, str]:
    patch_binary = shutil.which("patch")
    if not patch_binary:
        return "", "repo_patch requires the `patch` executable, but it was not found in PATH"
    try:
        completed = subprocess.run(
            [patch_binary, "--version"],
            text=True,
            capture_output=True,
            timeout=10,
            check=False,
        )
    except Exception as exc:
        return "", f"`patch --version` failed: {exc}"
    if completed.returncode not in (0, 1):
        return "", f"`patch --version` failed: {completed.stderr or completed.stdout}"
    return patch_binary, ""


def _write_minimal_xlsx(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    sheet_rows: list[str] = []
    for row_index, row in enumerate(rows, 1):
        cells: list[str] = []
        for col_index, value in enumerate(row, 1):
            ref = f"{_excel_column(col_index)}{row_index}"
            text = _xml_escape(str(value))
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>')
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData></worksheet>'
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as handle:
        handle.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            "</Types>",
        )
        handle.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        handle.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        handle.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>",
        )
        handle.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def _write_minimal_pdf(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    path.write_bytes(
        (
            "%PDF-1.4\n"
            "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n"
            "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n"
            "3 0 obj << /Type /Page /Parent 2 0 R /Resources << /Font << /F1 4 0 R >> >> "
            "/MediaBox [0 0 300 144] /Contents 5 0 R >> endobj\n"
            "4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n"
            f"5 0 obj << /Length {len(escaped) + 48} >> stream\n"
            f"BT /F1 12 Tf 36 100 Td ({escaped}) Tj ET\n"
            "endstream endobj\n"
            "trailer << /Root 1 0 R >>\n%%EOF\n"
        ).encode("utf-8")
    )


def _excel_column(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _flatten_string_values(value: Any) -> list[str]:
    if isinstance(value, str) and value.strip():
        return [value.strip()]
    if isinstance(value, list):
        flattened: list[str] = []
        for item in value:
            flattened.extend(_flatten_string_values(item))
        return flattened
    if isinstance(value, dict):
        flattened = []
        for item in value.values():
            flattened.extend(_flatten_string_values(item))
        return flattened
    return []


def _repo_patch_target_repo(item: BenchmarkItem) -> str:
    for key in ("target_repo", "repo", "repo_url", "git_repo", "repository"):
        value = item.metadata.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _repo_patch_base_commit(item: BenchmarkItem) -> str:
    for key in ("base_commit", "commit", "base_sha", "revision"):
        value = item.metadata.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _prepare_target_repo_checkout(item: BenchmarkItem, target_repo: str, base_commit: str) -> Path:
    existing = _find_materialized_target_repo(item, target_repo)
    if existing is not None:
        return _prepare_isolated_repo_checkout(item, target_repo, base_commit, str(existing))
    local_source = Path(target_repo).expanduser()
    if local_source.exists():
        return _prepare_isolated_repo_checkout(item, target_repo, base_commit, str(local_source.resolve()))
    if os.environ.get("AGENT_BENCH_ALLOW_TARGET_CHECKOUT", "").strip().lower() not in {"1", "true", "yes", "on"}:
        raise AdapterSetupError(
            "failed_harness_setup",
            "repo_patch target repository checkout was not materialized",
            {
                "target_repo": target_repo,
                "base_commit": base_commit,
                "hint": f"set {TARGET_REPO_ROOT_ENV} or AGENT_BENCH_ALLOW_TARGET_CHECKOUT=1",
            },
        )
    return _prepare_isolated_repo_checkout(item, target_repo, base_commit, _repo_url(target_repo))


def _prepare_isolated_repo_checkout(
    item: BenchmarkItem,
    target_repo: str,
    base_commit: str,
    clone_source: str,
) -> Path:
    checkout_root = Path(os.environ.get("AGENT_BENCH_OUTPUT_DIR", "/outputs")) / "target_repos" / _safe_slug(target_repo)
    checkout_root.mkdir(parents=True, exist_ok=True)
    target = checkout_root / _safe_slug(_item_id(item))
    if target.exists():
        shutil.rmtree(target)
    if not target.exists():
        clone = subprocess.run(
            ["git", "clone", clone_source, str(target)],
            text=True,
            capture_output=True,
            timeout=_git_timeout_seconds(),
            check=False,
        )
        if clone.returncode != 0:
            raise AdapterSetupError(
                "failed_harness_setup",
                f"unable to clone repo_patch target repository: {(clone.stderr or clone.stdout).strip()}",
                {"target_repo": target_repo},
            )
    _checkout_commit(target, base_commit)
    return target


def _repo_patch_checkout_available(items: list[BenchmarkItem]) -> bool:
    if not items:
        return False
    if os.environ.get("AGENT_BENCH_ALLOW_TARGET_CHECKOUT", "").strip().lower() in {"1", "true", "yes", "on"}:
        return True
    for item in items:
        target_repo = _repo_patch_target_repo(item)
        if not target_repo:
            return False
        if Path(target_repo).expanduser().exists():
            continue
        if _find_materialized_target_repo(item, target_repo) is None:
            return False
    return True


def _find_materialized_target_repo(item: BenchmarkItem, target_repo: str) -> Path | None:
    root_value = os.environ.get(TARGET_REPO_ROOT_ENV, "").strip()
    if not root_value:
        return None
    root = Path(root_value).expanduser().resolve()
    candidates = [
        item.metadata.get("instance_id", ""),
        target_repo.rstrip("/").split("/")[-1].removesuffix(".git"),
        _safe_slug(target_repo),
    ]
    for candidate in candidates:
        if not isinstance(candidate, str) or not candidate:
            continue
        path = (root / candidate).resolve()
        if (path / ".git").exists():
            return path
    if (root / ".git").exists():
        return root
    return None


def _checkout_commit(root: Path, commit: str) -> None:
    verify = subprocess.run(
        ["git", "-C", str(root), "rev-parse", "--verify", f"{commit}^{{commit}}"],
        text=True,
        capture_output=True,
        timeout=30,
        check=False,
    )
    if verify.returncode != 0:
        raise AdapterSetupError(
            "failed_harness_setup",
            f"target repo does not contain base_commit {commit}",
            {"target_repo_path": str(root), "base_commit": commit},
        )
    checkout = subprocess.run(
        ["git", "-C", str(root), "checkout", "--detach", commit],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if checkout.returncode != 0:
        raise AdapterSetupError(
            "failed_harness_setup",
            f"unable to checkout base_commit {commit}: {(checkout.stderr or checkout.stdout).strip()}",
            {"target_repo_path": str(root), "base_commit": commit},
        )
    reset = subprocess.run(
        ["git", "-C", str(root), "reset", "--hard", commit],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if reset.returncode != 0:
        raise AdapterSetupError(
            "failed_harness_setup",
            f"unable to reset target repo to base_commit {commit}: {(reset.stderr or reset.stdout).strip()}",
            {"target_repo_path": str(root), "base_commit": commit},
        )
    clean = subprocess.run(
        ["git", "-C", str(root), "clean", "-fdx"],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if clean.returncode != 0:
        raise AdapterSetupError(
            "failed_harness_setup",
            f"unable to clean target repo checkout: {(clean.stderr or clean.stdout).strip()}",
            {"target_repo_path": str(root), "base_commit": commit},
        )


def _repo_url(value: str) -> str:
    if value.startswith(("http://", "https://", "git@", "ssh://")):
        return value
    if "/" in value:
        return f"https://github.com/{value.removesuffix('.git')}.git"
    return value


def _git_timeout_seconds() -> int:
    return _bounded_int(os.environ.get("AGENT_BENCH_GIT_TIMEOUT"), 900, 30, 7200)


def _git_diff(root: Path) -> str:
    completed = subprocess.run(
        ["git", "-C", str(root), "diff", "--binary"],
        text=True,
        capture_output=True,
        timeout=120,
        check=False,
    )
    if completed.returncode != 0:
        raise AdapterSetupError(
            "failed_harness_setup",
            f"unable to collect git diff: {(completed.stderr or completed.stdout).strip()}",
            {"target_repo_path": str(root)},
        )
    return completed.stdout


def _repo_patch_official_grader_config(benchmark: str, item: BenchmarkItem) -> dict[str, Any]:
    grader_command = os.environ.get(REPO_PATCH_GRADER_ENV, "").strip()
    if grader_command:
        return {
            "configured": True,
            "kind": "external_repo_patch_grader",
            "official_equivalent": True,
            "score_mode": "official_repo_patch_grader",
            "command": grader_command,
        }
    swelancer = _swelancer_official_grader_for_item(item, benchmark=benchmark)
    if swelancer.get("available"):
        return {
            "configured": True,
            "kind": "swelancer_task_tests",
            "official_equivalent": True,
            "score_mode": "official_swelancer_task_tests",
            **swelancer,
        }
    return {
        "configured": False,
        "kind": "",
        "official_equivalent": False,
        "score_mode": "",
    }


def _repo_patch_official_grade_fields(config: dict[str, Any], *, fallback_score_mode: str) -> dict[str, Any]:
    configured = bool(config.get("configured"))
    return {
        "official_grader": configured,
        "official_grader_configured": configured,
        "official_equivalent": bool(config.get("official_equivalent")) if configured else False,
        "score_mode": str(config.get("score_mode") or fallback_score_mode),
        "included_in_official_score": configured,
    }


def _repo_patch_artifact_check_grade_fields(
    config: dict[str, Any],
    *,
    fallback_score_mode: str,
    not_run_reason: str,
) -> dict[str, Any]:
    fields = _repo_patch_official_grade_fields(config, fallback_score_mode=fallback_score_mode)
    if fields["official_grader_configured"]:
        fields["method"] = "pre_grader_artifact_check"
        fields["official_grader_not_run_reason"] = not_run_reason
    else:
        fields["method"] = "repo_patch_output_presence"
    return fields


def _swelancer_official_grader_for_items(items: list[BenchmarkItem]) -> dict[str, Any]:
    if not items:
        return {"available": False, "reason": "no SWE-Lancer items were extracted"}
    infos = [_swelancer_official_grader_for_item(item) for item in items]
    missing = [info for info in infos if not info.get("available")]
    if missing:
        return {
            "available": False,
            "reason": "SWE-Lancer official task tests are missing for one or more items",
            "missing_count": len(missing),
            "missing_sources": [info.get("source", "") for info in missing[:20]],
            "missing_reasons": [info.get("reason", "") for info in missing[:20]],
        }
    commands = sorted({str(info.get("command", "")) for info in infos if str(info.get("command", ""))})
    return {
        "available": True,
        "mode": "builtin_swelancer_task_tests",
        "command": commands[0] if len(commands) == 1 else "per-item SWE-Lancer task test",
        "item_count": len(items),
        "issue_ids": [str(info.get("issue_id", "")) for info in infos],
    }


def _swelancer_official_grader_for_item(item: BenchmarkItem, *, benchmark: str = "") -> dict[str, Any]:
    if not _is_swelancer_item(item, benchmark=benchmark):
        return {"available": False, "source": item.source, "reason": "item is not a SWE-Lancer task"}
    issue_dir = _swelancer_issue_dir(item, Path.cwd())
    if issue_dir is None:
        return {
            "available": False,
            "source": item.source,
            "reason": "SWE-Lancer item does not identify an issue directory",
        }
    test_path = issue_dir / "test.py"
    if not test_path.is_file():
        return {
            "available": False,
            "source": item.source,
            "issue_dir": str(issue_dir),
            "reason": "SWE-Lancer issue test.py was not found",
        }
    issue_id = str(item.metadata.get("instance_id") or issue_dir.name)
    command = os.environ.get(SWELANCER_TEST_COMMAND_ENV, "").strip()
    if not command:
        command = f"python -m pytest {shlex.quote(str(test_path))}"
    return {
        "available": True,
        "source": item.source,
        "mode": "builtin_swelancer_task_tests",
        "issue_id": issue_id,
        "issue_dir": str(issue_dir),
        "test_path": str(test_path),
        "command": command,
    }


def _is_swelancer_item(item: BenchmarkItem, *, benchmark: str = "") -> bool:
    values = [
        benchmark,
        os.environ.get("AGENT_BENCH_BENCHMARK_NAME", ""),
        str(item.metadata.get("benchmark", "")),
        str(item.metadata.get("metadata_association", "")),
        str(item.metadata.get("upstream_public_repository", "")),
        str(item.source),
    ]
    for value in values:
        normalized = normalize_text(value).replace("-", " ")
        if "swe lancer" in normalized or "swelancer" in normalized:
            return True
    return False


def _swelancer_issue_dir(item: BenchmarkItem, catalog_root: Path) -> Path | None:
    raw_issue_dir = item.metadata.get("issue_dir")
    raw_instance_id = item.metadata.get("instance_id")
    raw_catalog_root = item.metadata.get("catalog_root")
    root = Path(str(raw_catalog_root)).expanduser() if isinstance(raw_catalog_root, str) and raw_catalog_root else catalog_root
    if isinstance(raw_issue_dir, str) and raw_issue_dir.strip():
        issue_dir = Path(raw_issue_dir.strip()).expanduser()
        if not issue_dir.is_absolute():
            issue_dir = root / issue_dir
        return issue_dir
    if isinstance(raw_instance_id, str) and raw_instance_id.strip():
        return root / "issues" / raw_instance_id.strip()
    return None


def _run_swelancer_official_grader(
    item: BenchmarkItem,
    outputs: OutputBundle,
    grader: dict[str, Any],
) -> tuple[float, dict[str, Any]]:
    patch_path = Path(str(outputs.metadata.get("patch_path") or "")).expanduser()
    source_checkout = Path(str(outputs.metadata.get("target_checkout_path") or "")).expanduser()
    base_commit = str(outputs.metadata.get("base_commit") or _repo_patch_base_commit(item))
    if not patch_path.is_file():
        return 0.0, _swelancer_official_harness_failure(
            "model patch file was not found for SWE-Lancer official grading",
            grader,
            {"patch_path": str(patch_path)},
        )
    if not (source_checkout / ".git").exists():
        return 0.0, _swelancer_official_harness_failure(
            "target checkout was not found for SWE-Lancer official grading",
            grader,
            {"target_checkout_path": str(source_checkout)},
        )
    grader_root = patch_path.parent / "official_swelancer_grader_checkout"
    if grader_root.exists():
        shutil.rmtree(grader_root)
    clone = subprocess.run(
        ["git", "clone", str(source_checkout), str(grader_root)],
        text=True,
        capture_output=True,
        timeout=_git_timeout_seconds(),
        check=False,
    )
    if clone.returncode != 0:
        return 0.0, _swelancer_official_harness_failure(
            "unable to create SWE-Lancer official grader checkout",
            grader,
            {"stderr": (clone.stderr or clone.stdout)[-2000:]},
        )
    checkout = subprocess.run(
        ["git", "-C", str(grader_root), "checkout", "--detach", base_commit],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if checkout.returncode != 0:
        return 0.0, _swelancer_official_harness_failure(
            "unable to checkout SWE-Lancer base commit for official grading",
            grader,
            {"base_commit": base_commit, "stderr": (checkout.stderr or checkout.stdout)[-2000:]},
        )
    apply = subprocess.run(
        ["git", "-C", str(grader_root), "apply", "--binary", "--whitespace=nowarn", str(patch_path)],
        text=True,
        capture_output=True,
        timeout=120,
        check=False,
    )
    if apply.returncode != 0:
        return 0.0, {
            "method": "official_swelancer_task_tests",
            "score": 0.0,
            "status": "failed_model_tool_use",
            "reason": "model.patch did not apply cleanly to the SWE-Lancer base checkout",
            "patch_apply_exit_code": apply.returncode,
            "patch_apply_output": (apply.stderr or apply.stdout)[-2000:],
            "grader_checkout_path": str(grader_root),
            **_repo_patch_official_grade_fields(grader, fallback_score_mode="official_swelancer_task_tests"),
        }
    completed = _run_swelancer_task_test_command(item, grader, grader_root, patch_path)
    if completed.get("launch_error"):
        return 0.0, _swelancer_official_harness_failure(
            str(completed["launch_error"]),
            grader,
            {"grader_checkout_path": str(grader_root)},
        )
    passed = completed.get("exit_code") == 0
    return (1.0 if passed else 0.0), {
        "method": "official_swelancer_task_tests",
        "score": 1.0 if passed else 0.0,
        "status": "passed" if passed else "failed_model_answer",
        "reason": "" if passed else "SWE-Lancer official task tests failed",
        "grader_checkout_path": str(grader_root),
        "issue_id": grader.get("issue_id", item.metadata.get("instance_id", "")),
        "test_path": grader.get("test_path", ""),
        "official_test_command": completed.get("command", ""),
        "official_test_exit_code": completed.get("exit_code"),
        "official_test_stdout_tail": completed.get("stdout_tail", ""),
        "official_test_stderr_tail": completed.get("stderr_tail", ""),
        "timed_out": bool(completed.get("timed_out")),
        **_repo_patch_official_grade_fields(grader, fallback_score_mode="official_swelancer_task_tests"),
    }


def _run_swelancer_task_test_command(
    item: BenchmarkItem,
    grader: dict[str, Any],
    grader_root: Path,
    patch_path: Path,
) -> dict[str, Any]:
    raw_command = os.environ.get(SWELANCER_TEST_COMMAND_ENV, "").strip()
    if raw_command:
        argv = shlex.split(raw_command)
        command_for_report = raw_command
    else:
        argv = ["python", "-m", "pytest", str(grader.get("test_path", ""))]
        command_for_report = " ".join(shlex.quote(part) for part in argv)
    env = os.environ.copy()
    issue_id = str(grader.get("issue_id") or item.metadata.get("instance_id") or "")
    env.update(
        {
            "AGENT_BENCH_SWELANCER_GRADER_CHECKOUT": str(grader_root),
            "AGENT_BENCH_SWELANCER_ISSUE_ID": issue_id,
            "AGENT_BENCH_SWELANCER_TEST_PATH": str(grader.get("test_path", "")),
            "AGENT_BENCH_MODEL_PATCH_PATH": str(patch_path),
            "ISSUE_ID": issue_id,
        }
    )
    existing_pythonpath = env.get("PYTHONPATH", "")
    issue_dir = Path(str(grader.get("issue_dir", ""))).expanduser()
    pythonpath_parts = [str(issue_dir.parent.parent), str(issue_dir.parent)]
    if existing_pythonpath:
        pythonpath_parts.append(existing_pythonpath)
    env["PYTHONPATH"] = os.pathsep.join(part for part in pythonpath_parts if part)
    try:
        completed = subprocess.run(
            argv,
            cwd=grader_root,
            text=True,
            capture_output=True,
            timeout=_bounded_int(os.environ.get(SWELANCER_GRADER_TIMEOUT_ENV), 1800, 5, 7200),
            check=False,
            env=env,
        )
        return {
            "command": command_for_report,
            "exit_code": completed.returncode,
            "stdout_tail": (completed.stdout or "")[-4000:],
            "stderr_tail": (completed.stderr or "")[-4000:],
            "timed_out": False,
        }
    except subprocess.TimeoutExpired as exc:
        return {
            "command": command_for_report,
            "exit_code": None,
            "stdout_tail": (exc.stdout or "")[-4000:] if isinstance(exc.stdout, str) else "",
            "stderr_tail": (exc.stderr or "")[-4000:] if isinstance(exc.stderr, str) else "",
            "timed_out": True,
        }
    except OSError as exc:
        return {"command": command_for_report, "launch_error": f"unable to run SWE-Lancer official tests: {exc}"}


def _swelancer_official_harness_failure(
    reason: str,
    grader: dict[str, Any],
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "method": "official_swelancer_task_tests",
        "score": 0.0,
        "status": "failed_grader",
        "reason": reason,
        **(details or {}),
        **_repo_patch_official_grade_fields(grader, fallback_score_mode="official_swelancer_task_tests"),
    }


def _run_repo_patch_grader(command: str, item: BenchmarkItem, outputs: OutputBundle) -> tuple[float, dict[str, Any]]:
    env = os.environ.copy()
    env["AGENT_BENCH_MODEL_PATCH_PATH"] = str(outputs.metadata.get("patch_path", ""))
    env["AGENT_BENCH_TARGET_REPO"] = str(outputs.metadata.get("target_repo", ""))
    env["AGENT_BENCH_BASE_COMMIT"] = str(outputs.metadata.get("base_commit", ""))
    env["AGENT_BENCH_INSTANCE_ID"] = str(item.metadata.get("instance_id", ""))
    completed = subprocess.run(
        shlex.split(command),
        text=True,
        capture_output=True,
        timeout=_bounded_int(os.environ.get("AGENT_BENCH_REPO_PATCH_GRADER_TIMEOUT"), 1800, 30, 7200),
        check=False,
        env=env,
    )
    raw = (completed.stdout or "") + (completed.stderr or "")
    parsed = parse_json_object(raw)
    if not isinstance(parsed, dict):
        return 0.0, {
            "method": "official_patch_tests",
            "score": 0.0,
            "status": "failed_harness_setup",
            "reason": "official repo_patch grader did not return a JSON object",
            "grader_exit_code": completed.returncode,
            "grader_output_sample": raw[-1000:],
        }
    score = coerce_unit_score(parsed.get("score"))
    passed = bool(parsed.get("passed")) if "passed" in parsed else score >= 1.0
    return score, {
        "method": "official_patch_tests",
        "score": score,
        "passed": passed,
        "status": "passed" if passed else "failed_model_answer",
        "reason": str(parsed.get("reason", ""))[:500],
        "grader_exit_code": completed.returncode,
        "grader_raw_json": parsed,
    }


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9_.-]+", "-", value).strip("-._")
    return slug[:120] or "item"


def _parse_tool_arguments(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        if isinstance(parsed, dict):
            return parsed
    return {}


def _bounded_int(value: Any, default: int, lower: int, upper: int) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = default
    return max(lower, min(upper, number))


def _bounded_float(value: Any, default: float, lower: float, upper: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = default
    return max(lower, min(upper, number))


def _max_agent_turns() -> int:
    raw = os.environ.get("AGENT_BENCH_AGENT_TURNS", str(MAX_AGENT_TURNS))
    return _bounded_int(raw, MAX_AGENT_TURNS, 1, 24)


def _merge_usage(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    merged = dict(left)
    for key, value in right.items():
        if isinstance(value, (int, float)) and isinstance(merged.get(key), (int, float)):
            merged[key] += value
        else:
            merged[key] = value
    return merged


def _max_answer_tokens(item: BenchmarkItem) -> int:
    raw = os.environ.get("AGENT_BENCH_MAX_TOKENS", str(DEFAULT_AGENT_MAX_TOKENS))
    upper_bound = _bounded_int(raw, DEFAULT_AGENT_MAX_TOKENS, 512, 65536)
    grading = str(item.metadata.get("grading", "exact"))
    if item.choices:
        target = 512
    elif item.metadata.get("_file_artifact_task"):
        target = 4096
    elif _looks_like_patch_task(item):
        target = 4096
    elif grading in {"rubric", "task_compliance"}:
        target = 2048
    elif grading == "numeric":
        target = 4096
    else:
        target = 1024
    return min(upper_bound, target)


def _is_direct_answer_item(item: BenchmarkItem) -> bool:
    grading = str(item.metadata.get("grading", "exact"))
    if bool(item.choices) or grading == "numeric":
        return True
    if grading == "exact" and str(item.metadata.get("expected_key", "")) == "answer_rubric":
        return True
    return grading == "exact" and (
        str(item.metadata.get("expected_key", "")) == "fixture_needle"
        or "codeneedle" in f"{item.source} {item.question[:200]}".lower()
    )


def _needs_direct_answer_retry(answer: str, item: BenchmarkItem) -> bool:
    if not answer.strip():
        return True
    if isinstance(parse_json_object(answer), dict):
        return False
    if item.choices and normalize_answer_label(answer, item.choices):
        return False
    if len(answer) > 240:
        return True
    return "\n" in answer.strip()


def _direct_answer_retry_instruction(item: BenchmarkItem) -> str:
    if item.choices:
        labels = ", ".join(sorted(item.choices))
        values = ", ".join(str(value) for _, value in sorted(item.choices.items()))
        return (
            "Your previous response was not valid final-answer JSON. Extract only your final choice. "
            f"Return compact JSON only, using one of these labels ({labels}) or values ({values}), "
            'for example {"answer":"A","confidence":0.0}.'
        )
    return (
        "Your previous response was not valid final-answer JSON. Extract only the exact final answer. "
        'Return compact JSON only, for example {"answer":"value","confidence":0.0}.'
    )


def _agent_system_prompt(direct_answer: bool) -> str:
    if direct_answer:
        return (
            "You answer benchmark questions. Do not use tools. Do not include reasoning, derivations, "
            "markdown, or scratch work. Return only compact JSON with keys answer and confidence."
        )
    return (
        "You are an agent running inside a cloned benchmark repository. Use tools when they are useful. "
        "Do not include hidden reasoning or scratch work. If native tool calls are unavailable, request "
        'one tool by returning only JSON like {"tool":"read_file","arguments":{"path":"README.md"}}. '
        "After observing tool results, provide the final benchmark answer as compact JSON."
    )


def _looks_like_patch_task(item: BenchmarkItem) -> bool:
    expected = item.expected.lstrip()
    if expected.startswith(("diff --git ", "--- ", "Index: ")):
        return True
    text = f"{item.source} {item.question[:500]}".lower()
    return any(marker in text for marker in ("swe-bench", "patch")) or bool(
        re.search(r"\b(?:bug|fix|repo|repository)\b", text)
    )


def grade_answer(benchmark: str, item: BenchmarkItem, answer: str) -> tuple[float, dict[str, Any]]:
    grading = item.metadata.get("grading", "exact")
    if normalize_text(benchmark).replace("-", " ") == "stockbench":
        extracted, extraction_status, extraction_error = _extract_stockbench_model_answer(answer)
        if extracted is None:
            return 0.0, {
                "method": "stockbench_label_rationale",
                "score": 0.0,
                "status": "failed_model_format",
                "reason": extraction_error,
                "raw_model_response": answer,
                "extracted_answer": {},
                "extraction_status": extraction_status,
            }
        score, grade = judge_answer(
            benchmark,
            item,
            json.dumps(extracted, ensure_ascii=False, sort_keys=True),
            "stockbench_label_rationale",
        )
        grade["raw_model_response"] = answer
        grade["extracted_answer"] = extracted
        grade["extraction_status"] = extraction_status
        return score, grade
    if grading in {"exact", "numeric"}:
        if _contains_tool_syntax(answer):
            return 0.0, {
                "method": str(grading),
                "score": 0.0,
                "status": "failed_model_format",
                "reason": "answer contained tool-call syntax instead of the requested final answer",
            }
        score = score_answer(answer, item.expected, item.choices)
        return score, {
            "method": str(grading),
            "score": score,
            "status": "passed" if score >= 1.0 else "failed_model_answer",
        }
    return judge_answer(benchmark, item, answer, str(grading))


def _extract_stockbench_model_answer(answer: str) -> tuple[dict[str, str] | None, str, str]:
    try:
        parsed = json.loads(strip_thinking_blocks(answer).strip())
    except json.JSONDecodeError:
        parsed = None
    if not isinstance(parsed, dict):
        return None, "failed", "StockBench answer must be JSON with label and rationale"
    label = str(parsed.get("label") or "").strip().lower()
    rationale = str(parsed.get("rationale") or "").strip()
    if label not in {"stronger", "weaker", "mixed"}:
        return None, "failed", "StockBench JSON is missing a valid label"
    if not rationale:
        return None, "failed", "StockBench JSON is missing rationale"
    return {"label": label, "rationale": rationale}, "parsed", ""


def judge_answer(benchmark: str, item: BenchmarkItem, answer: str, method: str) -> tuple[float, dict[str, Any]]:
    if not answer:
        return 0.0, {"method": method, "score": 0.0, "reason": "empty candidate answer"}
    base_url = os.environ.get("AGENT_BENCH_JUDGE_BASE_URL", os.environ.get("AGENT_BENCH_BASE_URL", "")).rstrip("/")
    model = os.environ.get("AGENT_BENCH_JUDGE_MODEL", os.environ.get("AGENT_BENCH_MODEL", ""))
    rubric = item.expected
    if method == "task_compliance":
        rubric = (
            "Grade whether the candidate answer directly satisfies the benchmark task prompt. "
            "Award partial credit only for concrete, correct, task-relevant work."
        )
    candidate_answer = _judge_visible_answer(item, answer)
    choices_text = _judge_visible_choices(item)
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are grading a benchmark response. Do not include hidden reasoning or scratch work. "
                    "Do not require derivations, intermediate work, or explanations unless the task prompt or "
                    "rubric explicitly asks for them; a concise final answer can be sufficient. "
                    "Return only compact JSON with "
                    '{"score":0.0,"passed":false,"reason":"brief concrete reason"}. '
                    "The score must be between 0 and 1. Keep reason under 120 characters."
                ),
            },
            {
                "role": "user",
                "content": (
                    "/no_think\n"
                    f"Benchmark: {benchmark}\n"
                    f"Grading method: {method}\n\n"
                    f"Task prompt:\n{item.question}\n\n"
                    f"{choices_text}"
                    f"Rubric or expected behavior:\n{rubric}\n\n"
                    f"Candidate answer:\n{candidate_answer}\n"
                ),
            },
        ],
        "temperature": _bounded_float(os.environ.get("AGENT_BENCH_JUDGE_TEMPERATURE"), 0.0, 0.0, 2.0),
        "top_p": 1,
        "max_tokens": _bounded_int(os.environ.get("AGENT_BENCH_JUDGE_MAX_TOKENS"), 2048, 256, 8192),
        "stream": False,
        "response_format": {"type": "json_object"},
    }
    headers = chat_completion_headers("judge")
    budget_failure = _token_budget_failure(payload)
    if budget_failure is not None:
        return 0.0, {
            "method": method,
            "score": 0.0,
            "status": "failed_token_budget",
            "reason": str(budget_failure["reason"]),
            "token_budget": budget_failure,
        }
    schema = _judge_schema_for(benchmark, method)
    attempts: list[dict[str, Any]] = []
    max_retries = _bounded_int(os.environ.get("AGENT_BENCH_JUDGE_MAX_RETRIES"), 2, 0, 5)
    parsed_response: dict[str, Any] = {}
    grade: dict[str, Any] | None = None
    repaired = False
    judge_text = ""
    for attempt in range(max_retries + 1):
        request_payload = payload if attempt == 0 else _judge_repair_payload(payload, attempts[-1]["raw"], schema)
        try:
            response_body = post_chat_completion(base_url, request_payload, headers)
        except ChatCompletionTimeoutError as exc:
            return 0.0, {
                "method": method,
                "score": 0.0,
                "status": "timed_out",
                "reason": str(exc),
                "timed_out": True,
                "judge_retry_count": attempt,
            }
        except ChatCompletionHTTPError as exc:
            status = "failed_token_budget" if _looks_like_context_error(exc.body) else "failed_harness_setup"
            return 0.0, {
                "method": method,
                "score": 0.0,
                "status": status,
                "reason": f"judge request failed: HTTP {exc.code}: {exc.body[-500:]}",
                "judge_retry_count": attempt,
            }
        except Exception as exc:
            return 0.0, {
                "method": method,
                "score": 0.0,
                "status": "failed_harness_setup",
                "reason": f"judge request failed: {exc}",
                "judge_retry_count": attempt,
            }
        try:
            parsed_response = json.loads(response_body)
            content = extract_openai_content(parsed_response)
            judge_text = content or response_body
            if not content:
                reasoning = extract_openai_reasoning(parsed_response)
                judge_text = reasoning or response_body
            candidate, repaired = parse_judge_json_with_repair(judge_text)
            error_message = _judge_schema_error(candidate, schema)
            if error_message:
                raise ValueError(error_message)
            grade = candidate
            attempts.append({"raw": judge_text, "error": ""})
            break
        except (json.JSONDecodeError, ValueError) as exc:
            raw = response_body
            if isinstance(parsed_response, dict):
                raw = extract_openai_content(parsed_response) or extract_openai_reasoning(parsed_response) or response_body
            attempts.append({"raw": raw, "error": str(exc)})
            judge_text = raw
    if grade is None:
        for attempt_entry in reversed(attempts):
            raw = str(attempt_entry.get("raw") or "")
            prose_grade = parse_prose_judge_grade(raw) or parse_qualitative_prose_judge_grade(raw)
            if prose_grade is None:
                continue
            score = coerce_unit_score(prose_grade.get("score"))
            score, choice_reason = _apply_choice_task_partial_credit(method, item, answer, score)
            passed = bool(prose_grade.get("passed")) if "passed" in prose_grade else score >= 1.0
            status = "passed" if passed else "failed_model_answer"
            reason = (choice_reason or str(prose_grade.get("reason", "")))[:500]
            return score, {
                "method": method,
                "score": score,
                "passed": passed,
                "status": status,
                "reason": reason,
                "judge_raw_text": raw,
                "judge_parsed_json": prose_grade,
                "judge_parse_repaired": True,
                "judge_parser_status": "prose_score",
                "judge_sample": raw[:500],
                "judge_attempts": attempts,
                "judge_retry_count": max(0, len(attempts) - 1),
                "usage": parsed_response.get("usage") if isinstance(parsed_response.get("usage"), dict) else {},
            }
        return 0.0, {
            "method": method,
            "score": 0.0,
            "passed": False,
            "status": "failed_grader",
            "reason": "judge_parse_error: judge did not return valid JSON after retries",
            "setup_details": {
                "blocker_type": "judge_parse_error",
                "judge_attempts": attempts,
                "judge_retry_count": max(0, len(attempts) - 1),
            },
            "judge_raw_text": judge_text,
            "judge_parsed_json": {},
            "judge_parser_status": "judge_parse_error",
            "judge_sample": judge_text[:300],
            "judge_attempts": attempts,
            "judge_retry_count": max(0, len(attempts) - 1),
            "usage": parsed_response.get("usage") if isinstance(parsed_response.get("usage"), dict) else {},
        }
    score = coerce_unit_score(grade.get("score"))
    score, choice_reason = _apply_choice_task_partial_credit(method, item, answer, score)
    passed = bool(grade.get("passed")) if "passed" in grade else score >= 1.0
    status = "passed" if passed else "failed_model_answer"
    reason = (choice_reason or str(grade.get("reason", "")))[:500]
    if normalize_text(reason) == "short reason":
        status = "failed_grader"
        score = 0.0
        passed = False
        reason = "judge returned placeholder reason; excluded from primary scoring"
    return score, {
        "method": method,
        "score": score,
        "passed": passed,
        "status": status,
        "reason": reason,
        "judge_raw_text": judge_text,
        "judge_parsed_json": grade,
        "judge_parse_repaired": repaired,
        "judge_parser_status": "parsed_after_retry" if len(attempts) > 1 else ("repaired" if repaired else "parsed"),
        "judge_sample": judge_text[:500],
        "judge_attempts": attempts,
        "judge_retry_count": max(0, len(attempts) - 1),
        "usage": parsed_response.get("usage") if isinstance(parsed_response.get("usage"), dict) else {},
    }


def _apply_choice_task_partial_credit(
    method: str,
    item: BenchmarkItem,
    answer: str,
    score: float,
) -> tuple[float, str]:
    if method != "task_compliance" or score > 0.0 or not item.choices:
        return score, ""
    label = _choice_answer_label(answer, item.choices)
    if not label:
        return score, ""
    return 0.5, f"candidate provided a valid choice label: {label} ({item.choices[label]})"


def _judge_schema_for(benchmark: str, method: str) -> dict[str, Any]:
    normalized = normalize_text(benchmark).replace("-", " ")
    required = {"score": (int, float), "passed": bool, "reason": str}
    if normalized == "stockbench" or method == "stockbench_label_rationale":
        required.update(
            {
                "label_correct": bool,
                "rationale_present": bool,
                "rationale_quality": str,
            }
        )
    return {"required": required}


def _judge_schema_error(grade: dict[str, Any], schema: dict[str, Any]) -> str:
    required = schema.get("required") if isinstance(schema, dict) else {}
    if not isinstance(required, dict):
        return ""
    for key, expected_type in required.items():
        if key not in grade:
            return f"judge JSON missing required key: {key}"
        value = grade.get(key)
        if expected_type is bool and not isinstance(value, bool):
            return f"judge JSON key {key} must be boolean"
        if expected_type is str and not isinstance(value, str):
            return f"judge JSON key {key} must be string"
        if expected_type == (int, float) and not isinstance(value, (int, float)):
            return f"judge JSON key {key} must be numeric"
    return ""


def _judge_repair_payload(payload: dict[str, Any], previous_raw: str, schema: dict[str, Any]) -> dict[str, Any]:
    required = schema.get("required") if isinstance(schema, dict) else {}
    required_keys = ", ".join(str(key) for key in required) if isinstance(required, dict) else "score, passed, reason"
    repaired = json.loads(json.dumps(payload))
    original_messages = payload.get("messages") if isinstance(payload.get("messages"), list) else []
    original_prompt = ""
    if len(original_messages) > 1 and isinstance(original_messages[1], dict):
        original_prompt = str(original_messages[1].get("content") or "")
    repaired["messages"] = [
        {
            "role": "system",
            "content": (
                "Return valid compact JSON only. No prose, markdown, code fences, or hidden reasoning. "
                f"The object must include these keys: {required_keys}."
            ),
        },
        {
            "role": "user",
            "content": (
                "The previous judge response was invalid for the required schema.\n\n"
                f"Original grading request:\n{truncate_middle(original_prompt, 7000)}\n\n"
                f"Previous response:\n{truncate(previous_raw, 2000)}\n\n"
                "Re-grade the original candidate and return only a valid JSON object."
            ),
        },
    ]
    repaired["response_format"] = {"type": "json_object"}
    return repaired


def _judge_visible_choices(item: BenchmarkItem) -> str:
    if not item.choices:
        return ""
    rows = "\n".join(f"{label}. {text}" for label, text in sorted(item.choices.items()))
    return f"Choices:\n{rows}\n\n"


def _judge_visible_answer(item: BenchmarkItem, answer: str) -> str:
    if not item.choices:
        return answer
    label = _choice_answer_label(answer, item.choices)
    if not label:
        return answer
    expanded = f"{label} ({item.choices[label]})"
    if normalize_text(answer) == normalize_text(label) or normalize_text(answer) == normalize_text(item.choices[label]):
        return expanded
    return f"{expanded}\n\nOriginal response:\n{answer}"


def extract_openai_content(parsed: Any) -> str:
    message = extract_openai_message(parsed)
    content = _content_to_text(message.get("content"))
    if content:
        return content
    return ""


def extract_openai_reasoning(parsed: Any) -> str:
    message = extract_openai_message(parsed)
    for key in ("reasoning_content", "reasoning"):
        content = _content_to_text(message.get(key))
        if content:
            return content
    return ""


def _message_has_hidden_reasoning(message: dict[str, Any]) -> bool:
    for key in ("reasoning_content", "reasoning"):
        if _content_to_text(message.get(key)):
            return True
    return False


def extract_openai_finish_reason(parsed: Any) -> str:
    choices = parsed.get("choices") if isinstance(parsed, dict) else None
    if not isinstance(choices, list) or not choices:
        return ""
    first = choices[0]
    if not isinstance(first, dict):
        return ""
    value = first.get("finish_reason")
    return value if isinstance(value, str) else ""


def extract_openai_message(parsed: Any) -> dict[str, Any]:
    choices = parsed.get("choices") if isinstance(parsed, dict) else None
    if not isinstance(choices, list) or not choices:
        return {}
    first = choices[0]
    if not isinstance(first, dict):
        return {}
    message = first.get("message")
    if isinstance(message, dict):
        return message
    text = first.get("text")
    return {"content": text if isinstance(text, str) else ""}


def extract_openai_tool_calls(parsed: Any) -> list[dict[str, Any]]:
    message = extract_openai_message(parsed)
    tool_calls = message.get("tool_calls")
    if isinstance(tool_calls, list):
        return [item for item in tool_calls if isinstance(item, dict)]
    function_call = message.get("function_call")
    if isinstance(function_call, dict):
        return [{"id": "function_call", "type": "function", "function": function_call}]
    return []


def _content_to_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, str):
                parts.append(item)
            elif isinstance(item, dict):
                for key in ("text", "content"):
                    text = item.get(key)
                    if isinstance(text, str):
                        parts.append(text)
                        break
        return "\n".join(part for part in parts if part)
    return ""


def post_chat_completion(base_url: str, payload: dict[str, Any], headers: dict[str, str]) -> str:
    response_body, _ = post_chat_completion_with_variant(base_url, payload, headers)
    return response_body


def chat_completion_headers(label: str) -> dict[str, str]:
    headers = {
        "Content-Type": "application/json",
        "X-Agent-Bench-Proxy-Label": label,
        "X-Agent-Bench-Benchmark-Id": os.environ.get("AGENT_BENCH_BENCHMARK_ID", ""),
        "X-Agent-Bench-Task-Id": os.environ.get("AGENT_BENCH_TASK_ID", ""),
    }
    api_key = os.environ.get("AGENT_BENCH_API_KEY", "")
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    return {key: value for key, value in headers.items() if value}


def post_chat_completion_with_variant(
    base_url: str,
    payload: dict[str, Any],
    headers: dict[str, str],
) -> tuple[str, dict[str, Any]]:
    variants = _chat_completion_payload_variants(payload)
    last_error: ChatCompletionHTTPError | None = None
    timeout_seconds = _model_request_timeout()
    for index, variant in enumerate(variants):
        body = json.dumps(variant).encode("utf-8")
        probe = request.Request(f"{base_url}/chat/completions", data=body, headers=headers, method="POST")
        try:
            with wall_clock_timeout(timeout_seconds):
                with request.urlopen(probe, timeout=timeout_seconds) as response:
                    return response.read().decode("utf-8", errors="replace"), variant
        except error.HTTPError as exc:
            response_body = exc.read().decode("utf-8", errors="replace")
            last_error = ChatCompletionHTTPError(exc.code, response_body)
            if exc.code in HTTP_FALLBACK_STATUS_CODES and index < len(variants) - 1:
                continue
            raise last_error from exc
        except (TimeoutError, socket.timeout) as exc:
            raise ChatCompletionTimeoutError(timeout_seconds) from exc
        except error.URLError as exc:
            if _is_url_timeout(exc):
                raise ChatCompletionTimeoutError(timeout_seconds) from exc
            raise
    if last_error is not None:
        raise last_error
    raise RuntimeError("no chat completion payload variants were available")


def _is_url_timeout(exc: error.URLError) -> bool:
    reason = getattr(exc, "reason", None)
    if isinstance(reason, (TimeoutError, socket.timeout)):
        return True
    return "timed out" in str(reason).lower()


def _looks_like_context_error(text: str) -> bool:
    lowered = text.lower()
    return any(
        marker in lowered
        for marker in (
            "context length",
            "context window",
            "maximum context",
            "too many tokens",
            "token limit",
            "prompt is too long",
            "maximum number of tokens",
        )
    )


def _chat_completion_payload_variants(payload: dict[str, Any]) -> list[dict[str, Any]]:
    variants = [payload]
    if "response_format" in payload:
        without_response_format = dict(payload)
        without_response_format.pop("response_format", None)
        variants.append(without_response_format)
    if "tools" in payload:
        without_tools = dict(payload)
        without_tools.pop("tools", None)
        without_tools.pop("tool_choice", None)
        without_tools.pop("response_format", None)
        variants.append(without_tools)
    return variants


def evaluation_payload(
    item: BenchmarkItem,
    answer: str,
    score: float,
    error_message: str,
    **extra: Any,
) -> dict[str, Any]:
    status = normalize_status(extra.pop("status", None))
    if not isinstance(status, str) or not status:
        status = "passed" if score >= 1.0 else "failed_model_answer"
    if status == "completed":
        status = "passed" if score >= 1.0 else "failed_model_answer"
    if status not in STRICT_STATUSES:
        status = "failed_harness_setup"
    required_capabilities = extra.pop("required_capabilities", sorted(_item_required_capabilities(item)))
    supported_capabilities = extra.pop("supported_capabilities", [])
    required_tools = extra.pop("required_tools", _item_required_tools(item))
    exposed_tools = extra.pop("exposed_tools", [])
    missing_tools = extra.pop("missing_tools", [])
    payload = {
        "source": item.source,
        "question": item.question,
        "answer": answer,
        "expected": item.expected,
        "choices": item.choices,
        "metadata": item.metadata,
        "score": score,
        "passed": score >= 1.0 and status == "passed",
        "status": status,
        "error": error_message,
        "included_in_official_score": status not in INVALID_EVALUATION_STATUSES,
        "required_capabilities": required_capabilities,
        "supported_capabilities": supported_capabilities,
        "required_tools": required_tools,
        "exposed_tools": exposed_tools,
        "missing_tools": missing_tools,
    }
    payload.update(extra)
    grade = payload.get("grade")
    if isinstance(grade, dict):
        explicit_inclusion = grade.get("included_in_official_score")
        if isinstance(explicit_inclusion, bool):
            payload["included_in_official_score"] = explicit_inclusion
        for key in ("official_equivalent", "score_mode", "official_grader", "fallback_grader"):
            if key in grade and key not in payload:
                payload[key] = grade[key]
    if status in INVALID_EVALUATION_STATUSES:
        payload["included_in_official_score"] = False
    if payload.get("capabilities_verified") is False:
        payload["included_in_official_score"] = False
    return payload


def summarize_evaluations(evaluations: list[dict[str, Any]]) -> dict[str, Any]:
    if not evaluations:
        return {}
    valid = _valid_evaluations(evaluations)
    passed = sum(1 for item in evaluations if item.get("passed"))
    total = len(evaluations)
    first = evaluations[0]
    grading_methods = sorted(
        {
            str(item.get("metadata", {}).get("grading") or item.get("grade", {}).get("method"))
            for item in evaluations
            if isinstance(item.get("metadata"), dict) or isinstance(item.get("grade"), dict)
        }
        - {""}
    )
    return {
        "ok": passed == total,
        "score": _average_score(valid),
        "raw_score": _average_score(evaluations),
        "valid_task_count": len(valid),
        "answer": f"{passed}/{total}",
        "expected": f"{total}/{total}",
        "question": first.get("question", ""),
        "content_sample": first.get("content_sample", ""),
        "grading_methods": grading_methods,
        "status_counts": _status_counts(evaluations),
        "grader_failure_count": sum(1 for item in evaluations if normalize_status(item.get("status")) == "failed_grader"),
        "judge_parse_failure_count": sum(1 for item in evaluations if normalize_status(item.get("status")) == "failed_grader"),
        "judge_retry_count": sum(int(item.get("judge_retry_count") or 0) for item in evaluations),
        "judge_parse_repaired_count": sum(1 for item in evaluations if item.get("judge_parse_repaired")),
        "usage": first.get("usage", {}),
        "error": "" if passed == total else f"{passed}/{total} benchmark records passed",
    }


def _valid_evaluations(evaluations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [item for item in evaluations if normalize_status(item.get("status")) not in INVALID_EVALUATION_STATUSES]


def _overall_status_and_error(evaluations: list[dict[str, Any]], setup_error: str) -> tuple[str, str]:
    if setup_error:
        lowered_error = setup_error.lower()
        if "disabled until" in setup_error:
            return "skipped_unsupported_capability", setup_error
        if "required securities-report task files are missing" in setup_error:
            return "skipped_unsupported_capability", setup_error
        if (
            any(marker in lowered_error for marker in ("missing", "not found", "not present"))
            and any(
                marker in lowered_error
                for marker in ("asset", "file", "directory", "cache", "task input", "securities-report")
            )
        ):
            return "failed_missing_assets", setup_error
        if setup_error.startswith("No benchmark task records"):
            return "failed_dataset_extraction", setup_error
        return "failed_harness_setup", setup_error
    if not evaluations:
        return "completed", ""
    valid_count = len(_valid_evaluations(evaluations))
    if valid_count > 0:
        return "completed", ""
    status_counts = _status_counts(evaluations)
    for status in (
        "failed_harness_setup",
        "failed_dataset_extraction",
        "failed_grader",
        "failed_token_budget",
            "failed_missing_required_tool",
            "failed_invalid_task_context",
            "failed_model_endpoint",
            "timed_out",
        "failed_missing_assets",
        "skipped_unsupported_capability",
    ):
        if status_count(status_counts, status):
            return status, _overall_error_message(status, status_counts)
    return "completed", ""


def _overall_error_message(status: str, status_counts: dict[str, int]) -> str:
    total = sum(status_counts.values())
    status_label = status.replace("_", " ")
    return f"All {total} benchmark record evaluation(s) were invalid: {status_label}"


def _status_counts(evaluations: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in evaluations:
        status = normalize_status(item.get("status"))
        if not isinstance(status, str) or not status:
            status = "passed" if item.get("passed") else "failed_model_answer"
        counts[status] = counts.get(status, 0) + 1
    return dict(sorted(counts.items()))


def score_answer(answer: str, expected: str, choices: dict[str, str]) -> float:
    if not answer:
        return 0.0
    answer = normalize_exact_answer(answer)
    expected = normalize_exact_answer(expected)
    if choices:
        answer_label = _choice_answer_label(answer, choices)
        expected_label = normalize_answer_label(expected, choices)
        if answer_label and expected_label and answer_label == expected_label:
            return 1.0
    for numeric_answer in _numeric_candidates(answer):
        for numeric_expected in _numeric_candidates(expected):
            tolerance = _numeric_tolerance(expected, numeric_expected)
            if abs(numeric_answer - numeric_expected) <= tolerance:
                return 1.0
    if normalize_text(answer) == normalize_text(expected):
        return 1.0
    if _exact_answer_embedded(answer, expected):
        return 1.0
    return 0.0


def normalize_exact_answer(value: str) -> str:
    parsed = parse_json_object(value)
    if isinstance(parsed, dict):
        answer = None
        for key in ("answer", "final_answer", "final", "value"):
            if key in parsed:
                answer = parsed.get(key)
                break
        if isinstance(answer, str):
            return answer.strip()
        if isinstance(answer, (int, float, bool)):
            return str(answer)
        if isinstance(answer, list) and len(answer) == 1:
            return normalize_exact_answer(str(answer[0]))
    return value.strip()


def coerce_unit_score(value: Any) -> float:
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return max(0.0, min(1.0, float(value)))
    if isinstance(value, str):
        match = re.search(r"-?\d+(?:\.\d+)?", value)
        if match:
            numeric = float(match.group(0))
            if numeric > 1.0:
                numeric /= 100.0
            return max(0.0, min(1.0, numeric))
    return 0.0


def normalize_answer_label(value: str, choices: dict[str, str]) -> str:
    stripped = value.strip()
    upper = stripped.upper()
    if upper in choices:
        return upper
    match = re.search(r"\b([A-Z])\b", upper)
    if match and match.group(1) in choices:
        return match.group(1)
    normalized = normalize_text(stripped)
    for label, choice in choices.items():
        if normalized == normalize_text(choice):
            return label
    if stripped.isdigit():
        return normalize_choice_label(stripped, 0)
    return ""


def _choice_answer_label(answer: str, choices: dict[str, str]) -> str:
    if _looks_like_concise_answer(answer):
        direct = normalize_answer_label(answer, choices)
        if direct:
            return direct

    labels = {label.upper(): label.upper() for label in choices}
    values = {normalize_text(text): label.upper() for label, text in choices.items()}
    candidates: list[str] = []
    tail = "\n".join(answer.strip().splitlines()[-20:])
    patterns = (
        r"\b(?:final\s+answer|answer|action|decision|recommendation|choice)\s*(?:is|=|:|-)\s*(?:choice\s*)?([A-Za-z0-9_ .+-]+)",
        r"\b(?:choose|select|picked?)\s+(?:choice\s*)?([A-Za-z0-9_ .+-]+)",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, tail, flags=re.IGNORECASE):
            value = match.group(1).strip().strip("`'\".。")
            label = normalize_answer_label(value, choices)
            if label:
                candidates.append(label)
                continue
            first_word = value.split()[0] if value.split() else ""
            label = normalize_answer_label(first_word, choices)
            if label:
                candidates.append(label)

    for line in reversed(answer.strip().splitlines()):
        stripped = line.strip().strip("*-• \t`'\".")
        if not stripped:
            continue
        label = normalize_answer_label(stripped, choices)
        if label:
            candidates.append(label)
            break
        normalized = normalize_text(stripped)
        if normalized in values:
            candidates.append(values[normalized])
            break
        if stripped.upper() in labels:
            candidates.append(stripped.upper())
            break

    return candidates[-1] if candidates else ""


def _looks_like_concise_answer(answer: str) -> bool:
    stripped = answer.strip()
    return len(stripped) <= 80 and "\n" not in stripped


def _exact_answer_embedded(answer: str, expected: str) -> bool:
    normalized_expected = normalize_text(expected)
    if not normalized_expected:
        return False
    if len(normalized_expected) >= 12 and normalized_expected in normalize_text(answer):
        return True
    for candidate in _inline_answer_candidates(answer):
        if normalize_text(candidate) == normalized_expected:
            return True
    return False


def _inline_answer_candidates(answer: str) -> list[str]:
    candidates: list[str] = []
    candidates.extend(match.group(1).strip() for match in re.finditer(r"`([^`\n]{1,300})`", answer))
    patterns = (
        r"\b(?:final\s+answer|answer|line|value)\s*(?:is|=|:|-)\s*([^\n]{1,300})",
        r"\b(?:return|choose)\s+([^\n]{1,300})",
    )
    for pattern in patterns:
        candidates.extend(match.group(1).strip().strip("`'\".") for match in re.finditer(pattern, answer, re.IGNORECASE))
    return [candidate for candidate in candidates if candidate]


def _numeric_value(value: str) -> float | None:
    candidates = _numeric_candidates(value)
    return candidates[0] if candidates else None


def _numeric_tolerance(expected: str, numeric_expected: float) -> float:
    decimal_tolerance = 0.0
    match = re.search(r"-?\$?\s*[0-9][0-9,]*(?:\.(\d+))?", normalize_exact_answer(expected))
    if match and match.group(1):
        decimal_tolerance = 0.5 * (10 ** -len(match.group(1)))
    return max(1e-6, decimal_tolerance, abs(numeric_expected) * 0.005)


def _numeric_candidates(value: str) -> list[float]:
    cleaned = normalize_exact_answer(value)
    multiplier = 1.0
    lowered = cleaned.lower()
    if re.search(r"\b(billion|bn)\b", lowered):
        multiplier = 1_000_000_000.0
    elif re.search(r"\b(million|mn|mm)\b", lowered):
        multiplier = 1_000_000.0
    elif re.search(r"\b(thousand|k)\b", lowered):
        multiplier = 1_000.0
    percent = "%" in cleaned or "percent" in lowered
    basis_points = bool(re.search(r"\b(bps|basis points?)\b", lowered))
    candidates: list[float] = []
    pattern = re.compile(r"(?P<paren>\(\s*)?(?P<sign>-?)\$?\s*(?P<number>[0-9][0-9,]*(?:\.\d+)?(?:e[+-]?\d+)?)", re.IGNORECASE)
    for match in pattern.finditer(cleaned):
        try:
            number = float(match.group("number").replace(",", "")) * multiplier
        except ValueError:
            continue
        if match.group("sign") == "-" or match.group("paren"):
            number = -number
        candidates.append(number)
        if percent:
            candidates.append(number / 100.0)
        if basis_points:
            candidates.append(number / 10_000.0)
    return candidates


def extract_answer(content: str) -> str:
    text = content.strip()
    if not text:
        return ""
    payload = parse_json_object(text)
    if isinstance(payload, dict):
        value = payload.get("answer")
        if isinstance(value, str) and value.strip():
            return value.strip()
        if isinstance(value, (int, float, bool)):
            return str(value)
    match = re.search(r"\b(?:answer|final)\s*[:\-]\s*([A-Za-z0-9_.+-]+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return text[:4000]


def parse_json_object(text: str) -> object | None:
    text = _strip_json_wrappers(strip_thinking_blocks(text).strip())
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        for candidate in extract_json_objects(text):
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                continue
        return None


def parse_judge_json(text: str) -> dict[str, Any]:
    parsed, _ = parse_judge_json_with_repair(text)
    return parsed


def parse_prose_judge_grade(text: str) -> dict[str, Any] | None:
    cleaned = strip_thinking_blocks(text).strip()
    score_match = re.search(
        r"\bscore\s*(?:is|=|:)?\s*"
        r"(?P<numerator>\d+(?:\.\d+)?)\s*/\s*(?P<denominator>\d+(?:\.\d+)?)\b",
        cleaned,
        re.IGNORECASE,
    )
    score_index = -1
    if score_match:
        numerator = float(score_match.group("numerator"))
        denominator = float(score_match.group("denominator"))
        score = 0.0 if denominator <= 0 else max(0.0, min(1.0, numerator / denominator))
        score_index = score_match.start()
    else:
        score_match = re.search(
            r"\bscore\s*(?:is|=|:)?\s*([01](?:\.\d+)?|100(?:\.0+)?|\d{1,2}(?:\.\d+)?)\b",
            cleaned,
            re.IGNORECASE,
        )
        if not score_match:
            return None
        raw_score = score_match.group(1)
        score = coerce_unit_score(raw_score)
        score_index = score_match.start()
    passed_match = re.search(r"\bpassed?\s*(?:is|=|:)?\s*(true|false|yes|no)\b", cleaned, re.IGNORECASE)
    if passed_match:
        passed = passed_match.group(1).lower() in {"true", "yes"}
    else:
        passed = score >= 1.0
    reason = _prose_judge_reason(cleaned, score_index)
    return {"score": score, "passed": passed, "reason": reason}


def parse_qualitative_prose_judge_grade(text: str) -> dict[str, Any] | None:
    cleaned = strip_thinking_blocks(text).strip()
    qualitative = _qualitative_prose_judge_grade(cleaned)
    if qualitative is not None:
        return qualitative
    return None


def _qualitative_prose_judge_grade(text: str) -> dict[str, Any] | None:
    lowered = text.lower()
    negative_patterns = (
        r"candidate (?:answer|response).{0,120}(?:does not|doesn't|fails?|failed|incomplete|incorrect|not provide)",
        r"(?:answer|response).{0,80}(?:not a solution|only contains a thought process|internal monologue)",
        r"(?:candidate )?(?:answer|response).{0,120}not a direct (?:answer|response)",
        r"(?:candidate )?(?:answer|response).{0,120}(?:meta-analysis|reasoning process|thought process)",
        r"fails? to provide",
    )
    if any(re.search(pattern, lowered, flags=re.DOTALL) for pattern in negative_patterns):
        return {
            "score": 0.0,
            "passed": False,
            "reason": "judge prose described the candidate as not directly satisfying the task",
        }
    positive_patterns = (
        r"candidate answer.{0,120}(?:accurate|correct|reasonable|concise|satisfies|addresses)",
        r"candidate response.{0,120}(?:accurate|correct|reasonable|concise|satisfies|addresses)",
        r"reasonable response",
        r"accurate summary",
    )
    if any(re.search(pattern, lowered, flags=re.DOTALL) for pattern in positive_patterns):
        return {
            "score": 0.5,
            "passed": False,
            "reason": "judge prose described the candidate as partially task-relevant",
        }
    return None


def _prose_judge_reason(text: str, score_index: int) -> str:
    before = text[:score_index].strip()
    after = text[score_index:].strip()
    for source in (after, before):
        sentences = [part.strip() for part in re.split(r"(?<=[.!?])\s+", source) if part.strip()]
        for sentence in sentences:
            if len(sentence) > 12 and "score" not in sentence.lower():
                return truncate(sentence, 200)
    return "judge returned a numeric prose score"


def parse_judge_json_with_repair(text: str) -> tuple[dict[str, Any], bool]:
    cleaned = _strip_json_wrappers(strip_thinking_blocks(text).strip()).strip()
    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        parsed = None
    if isinstance(parsed, dict) and JUDGE_REQUIRED_KEYS <= parsed.keys():
        return parsed, cleaned != text.strip()

    for candidate in extract_json_objects(cleaned):
        try:
            recovered = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(recovered, dict) and JUDGE_REQUIRED_KEYS <= recovered.keys():
            return recovered, True

    raise ValueError("judge content was not a JSON object")


def strip_thinking_blocks(text: str) -> str:
    without_blocks = re.sub(r"<think\b[^>]*>.*?</think>", "", text, flags=re.IGNORECASE | re.DOTALL)
    return re.sub(r"</think>", "", without_blocks, flags=re.IGNORECASE)


def extract_json_objects(text: str) -> list[str]:
    objects: list[str] = []
    start: int | None = None
    depth = 0
    in_string = False
    escaped = False
    for index, char in enumerate(text):
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
            continue
        if char == "{":
            if depth == 0:
                start = index
            depth += 1
            continue
        if char == "}" and depth:
            depth -= 1
            if depth == 0 and start is not None:
                objects.append(text[start : index + 1])
                start = None
    return objects


def _strip_json_wrappers(text: str) -> str:
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    return text


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9.+-]+", " ", value.lower())).strip()


def truncate(value: str, limit: int = MAX_FIELD_CHARS) -> str:
    value = value.strip()
    if len(value) <= limit:
        return value
    return value[: limit - 20].rstrip() + " ...[truncated]"


def truncate_middle(value: str, limit: int, marker: str = "\n...[truncated middle]...\n") -> str:
    value = value.strip()
    if len(value) <= limit:
        return value
    if limit <= len(marker) + 2:
        return truncate(value, limit)
    head = (limit - len(marker)) // 2
    tail = limit - len(marker) - head
    return f"{value[:head].rstrip()}{marker}{value[-tail:].lstrip()}"


def _sample_limit() -> int:
    raw = os.environ.get("AGENT_BENCH_SAMPLE_LIMIT") or os.environ.get("AGENT_BENCH_LIMIT") or "3"
    try:
        return max(1, min(20, int(raw)))
    except ValueError:
        return 3


def _model_request_timeout() -> float:
    raw = os.environ.get("AGENT_BENCH_MODEL_REQUEST_TIMEOUT", "1800")
    try:
        return max(1.0, float(raw))
    except ValueError:
        return 1800.0


def _average_score(evaluations: list[dict[str, Any]]) -> float:
    if not evaluations:
        return 0.0
    return sum(float(item.get("score", 0.0)) for item in evaluations) / len(evaluations)


def _is_remote_provider() -> bool:
    return (
        os.environ.get("AGENT_BENCH_PROVIDER") == "openai-compatible"
        and bool(os.environ.get("AGENT_BENCH_BASE_URL", "").strip())
        and bool(os.environ.get("AGENT_BENCH_MODEL", "").strip())
    )


if __name__ == "__main__":
    raise SystemExit(main())
